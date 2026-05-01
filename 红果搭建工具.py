"""
红果搭建工具 - 统一 GUI
合并安卓/IOS × 七留/每留 四套自定义素材搭建脚本
"""
import re, random, functools, time, logging, sys, os, threading, queue, json, webbrowser, subprocess, glob
from datetime import datetime
from pathlib import Path
try:
    import openpyxl
    from openpyxl import Workbook
except Exception:
    openpyxl = None
    Workbook = None
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout, expect

# ═══════════════════════════════════════════════════════════════
#  配置表：4 种组合的差异参数
# ═══════════════════════════════════════════════════════════════
PROFILES = {
    "安卓-每留": dict(
        strategy="安卓-每留",
        material_account_id="1855367293890569",
        audience_keyword="红果通用",
        monitor_btn_text="选择分包和链接组",
        name_prefix="安卓-站内-短剧-每留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\安卓\搭建\每留\ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\安卓\搭建\每留\logs"),
        wait_scale=0.6,
    ),
    "安卓-七留": dict(
        strategy="安卓-七留",
        material_account_id="1855367293890569",
        audience_keyword="红果通用",
        monitor_btn_text="选择分包和链接组",
        name_prefix="安卓-站内-短剧-七留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\安卓\搭建\七留\ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\安卓\搭建\七留\logs"),
        wait_scale=1.0,  # 七留用较大等待
    ),
    "IOS-每留": dict(
        strategy="IOS-每留",
        material_account_id="1859509275615367",
        audience_keyword="IOS定向",
        monitor_btn_text="选择链接组",
        name_prefix="IOS-站内-短剧-每留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\IOS\每留\ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\IOS\每留\logs"),
        wait_scale=0.6,
    ),
    "IOS-七留": dict(
        strategy="IOS-七留",
        material_account_id="1859509275615367",
        audience_keyword="IOS定向",
        monitor_btn_text="选择链接组",
        name_prefix="IOS-站内-短剧-七留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\IOS\七留\ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\IOS\七留\logs"),
        wait_scale=1.0,
    ),
}

INCENTIVE_PROFILES = {
    "安卓-激励每留": dict(
        strategy="安卓-激励-每留",
        material_account_id="1855641147536460",
        audience_keyword="通用激励",
        monitor_btn_text="选择分包和链接组",
        name_prefix="安卓-站内-激励-每留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\激励\安卓搭建\激励每留\激励ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\激励\安卓搭建\激励每留\logs"),
        wait_scale=0.6,
        build_mode="incentive",
        pages_per_round=3,
        push_account_id="1855641147536460",
    ),
    "安卓-激励七留": dict(
        strategy="安卓-激励-七留",
        material_account_id="1855641147536460",
        audience_keyword="通用激励",
        monitor_btn_text="选择分包和链接组",
        name_prefix="安卓-站内-激励-七留",
        ids_file=Path(r"C:\Users\Administrator\Desktop\红果\激励\安卓搭建\激励七留\激励ids.txt"),
        log_dir=Path(r"C:\Users\Administrator\Desktop\红果\激励\安卓搭建\激励七留\logs"),
        wait_scale=1.0,
        build_mode="incentive",
        pages_per_round=3,
        push_account_id="1855641147536460",
    ),
}

ALL_PROFILES = {**PROFILES, **INCENTIVE_PROFILES}

PROFILE_CATEGORIES = [
    ("短剧单本", list(PROFILES.keys())),
    ("短剧激励", list(INCENTIVE_PROFILES.keys())),
]

# ═══════════════════════════════════════════════════════════════
#  内置配置（config.json）层级化数据 —— 一站式 ids/链接/素材管理
# ═══════════════════════════════════════════════════════════════
CONFIG_FILE = Path(__file__).parent / "config.json"
BUILD_RECORD_FILE = Path(__file__).parent / "build_records.json"
MATERIAL_HISTORY_FILE = Path(__file__).parent / "material_history.json"

# 仅保存到 config.json 的可编辑字段
PROFILE_EDITABLE_FIELDS = (
    "strategy",
    "material_account_id",
    "audience_keyword",
    "monitor_btn_text",
    "name_prefix",
    "wait_scale",
)


def _profile_defaults(key: str) -> dict:
    p = ALL_PROFILES[key]
    return {
        "strategy": p["strategy"],
        "material_account_id": p["material_account_id"],
        "audience_keyword": p["audience_keyword"],
        "monitor_btn_text": p["monitor_btn_text"],
        "name_prefix": p["name_prefix"],
        "wait_scale": p["wait_scale"],
        "groups": [],
    }


def _default_config() -> dict:
    return {
        "common": {"cdp_endpoint": "http://localhost:9222", "drama_titles": []},
        "profiles": {key: _profile_defaults(key) for key in ALL_PROFILES},
    }


def _empty_drama() -> dict:
    return {"name": "", "click": "", "show": "", "video": "", "material_ids": []}


def _empty_group() -> dict:
    return {"account_ids": [], "dramas": [_empty_drama()]}


def load_config() -> dict:
    """读取 config.json，缺失字段用默认值兜底。"""
    cfg = _default_config()
    if CONFIG_FILE.exists():
        try:
            data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            data = {}
        if isinstance(data, dict):
            common = data.get("common") or {}
            if isinstance(common, dict):
                if isinstance(common.get("cdp_endpoint"), str):
                    cfg["common"]["cdp_endpoint"] = common.get("cdp_endpoint")
                titles = common.get("drama_titles") or []
                if isinstance(titles, str):
                    titles = [x.strip() for x in titles.splitlines() if x.strip()]
                if isinstance(titles, list):
                    seen_titles = set()
                    norm_titles = []
                    for title in titles:
                        title = str(title).strip()
                        key = re.sub(r"\W+", "", title, flags=re.UNICODE).lower()
                        if title and key and key not in seen_titles:
                            seen_titles.add(key)
                            norm_titles.append(title)
                    cfg["common"]["drama_titles"] = norm_titles
            profiles = data.get("profiles") or {}
            if isinstance(profiles, dict):
                for key in ALL_PROFILES:
                    src = profiles.get(key) or {}
                    if not isinstance(src, dict):
                        continue
                    dst = cfg["profiles"][key]
                    for f in PROFILE_EDITABLE_FIELDS:
                        if f in src and src[f] not in (None, ""):
                            dst[f] = src[f]
                    groups = src.get("groups")
                    if isinstance(groups, list):
                        norm_groups = []
                        for g in groups:
                            if not isinstance(g, dict):
                                continue
                            acc = g.get("account_ids") or []
                            if isinstance(acc, str):
                                acc = [x.strip() for x in re.split(r"[\s,]+", acc) if x.strip()]
                            acc = [str(x).strip() for x in acc if str(x).strip()]
                            dramas = g.get("dramas") or []
                            norm_dramas = []
                            if isinstance(dramas, list):
                                for d in dramas:
                                    if not isinstance(d, dict):
                                        continue
                                    mids = d.get("material_ids") or []
                                    if isinstance(mids, str):
                                        mids = [x.strip() for x in re.split(r"[\s,]+", mids) if x.strip()]
                                    mids = [str(x).strip() for x in mids if str(x).strip()]
                                    norm_dramas.append({
                                        "name": str(d.get("name", "")).strip(),
                                        "click": str(d.get("click", "")).strip(),
                                        "show": str(d.get("show", "")).strip(),
                                        "video": str(d.get("video", "")).strip(),
                                        "material_ids": mids,
                                    })
                            norm_groups.append({
                                "account_ids": acc,
                                "dramas": norm_dramas,
                                "group_name": str(g.get("group_name", "")).strip(),
                                "click_url": str(g.get("click_url", "")).strip(),
                                "show_url": str(g.get("show_url", "")).strip(),
                                "play_url": str(g.get("play_url", "")).strip(),
                            })
                        dst["groups"] = norm_groups
    try:
        wait_scale = cfg["profiles"][next(iter(PROFILES))]["wait_scale"]
        float(wait_scale)
    except Exception:
        pass
    return cfg


def save_config(cfg: dict) -> None:
    """保存 config.json（UTF-8、缩进、保留中文）。"""
    CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(
        json.dumps(cfg, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_build_records() -> dict:
    if BUILD_RECORD_FILE.exists():
        try:
            data = json.loads(BUILD_RECORD_FILE.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def save_build_records(records: dict) -> None:
    BUILD_RECORD_FILE.parent.mkdir(parents=True, exist_ok=True)
    BUILD_RECORD_FILE.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def record_build_success(account_count: int, project_count: int) -> None:
    today = datetime.now().strftime("%Y-%m-%d")
    records = load_build_records()
    day = records.get(today) or {"accounts": 0, "projects": 0}
    day["accounts"] = day.get("accounts", 0) + account_count
    day["projects"] = day.get("projects", 0) + project_count
    records[today] = day
    save_build_records(records)


def load_material_history() -> list:
    if MATERIAL_HISTORY_FILE.exists():
        try:
            data = json.loads(MATERIAL_HISTORY_FILE.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return data
        except Exception:
            pass
    return []


def save_material_history(history: list) -> None:
    MATERIAL_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    MATERIAL_HISTORY_FILE.write_text(
        json.dumps(history, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def add_material_history(names: list[str]) -> None:
    if not names:
        return
    history = load_material_history()
    existing = {r["name"] for r in history}
    date_tag = datetime.now().strftime("%m%d")
    for name in names:
        if name and name not in existing:
            history.insert(0, {"date": date_tag, "name": name})
            existing.add(name)
    save_material_history(history)


def get_used_material_names() -> set:
    return {r["name"] for r in load_material_history() if r.get("name")}


def _parse_ids_txt_groups(ids_file: Path):
    """解析 ids.txt → [(account_ids, [drama_dict])]。失败则返回空列表。"""
    if not ids_file.exists():
        return []
    try:
        raw_text = ids_file.read_text(encoding="utf-8")
    except Exception:
        return []
    raw_text = "\n".join(
        line for line in raw_text.splitlines()
        if not line.lstrip().startswith("#")
    )
    chunks = re.split(r"(?:═{3,}[^\n]*═{3,}|={3,})", raw_text.strip())
    _silent = logging.getLogger("_silent_migrate")
    _silent.addHandler(logging.NullHandler())
    _silent.propagate = False
    groups = []
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk:
            continue
        ids, dramas = _parse_single_group(chunk, _silent)
        if ids or dramas:
            groups.append((ids, dramas))
    return groups


def _parse_incentive_ids_txt(ids_file: Path):
    if not ids_file.exists():
        return []
    try:
        raw_text = ids_file.read_text(encoding="utf-8")
    except Exception:
        return []
    lines = [l.rstrip() for l in raw_text.splitlines()]
    groups = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not re.match(r"^组\d+", line):
            i += 1
            continue
        group_name = line
        i += 1
        account_ids = []
        while i < len(lines) and lines[i].strip():
            val = lines[i].strip()
            if val.isdigit() and len(val) > 10:
                account_ids.append(val)
            i += 1
        while i < len(lines) and not lines[i].strip():
            i += 1
        track_urls = []
        while i < len(lines) and lines[i].strip().startswith("http"):
            track_urls.append(lines[i].strip())
            i += 1
        click_url = ""
        show_url = ""
        play_url = ""
        for url in track_urls:
            if "action_type=click" in url:
                click_url = url
            elif "action_type=view" in url:
                show_url = url
            elif "action_type=effective_play" in url:
                play_url = url
        if account_ids:
            groups.append({
                "group_name": group_name,
                "account_ids": account_ids,
                "click_url": click_url,
                "show_url": show_url,
                "play_url": play_url,
            })
    return groups


def migrate_ids_txt_to_config() -> bool:
    cfg = load_config()
    changed = False
    for key, p in ALL_PROFILES.items():
        prof = cfg["profiles"].get(key)
        if not prof:
            continue
        if prof.get("groups"):
            continue
        is_incentive = p.get("build_mode") == "incentive"
        if is_incentive:
            inc_groups = _parse_incentive_ids_txt(p["ids_file"])
            if not inc_groups:
                continue
            new_groups = []
            for g in inc_groups:
                new_groups.append({
                    "account_ids": g["account_ids"],
                    "group_name": g["group_name"],
                    "click_url": g.get("click_url", ""),
                    "show_url": g.get("show_url", ""),
                    "play_url": g.get("play_url", ""),
                    "dramas": [],
                })
            prof["groups"] = new_groups
        else:
            groups = _parse_ids_txt_groups(p["ids_file"])
            if not groups:
                continue
            new_groups = []
            for ids, dramas in groups:
                new_groups.append({
                    "account_ids": [str(x) for x in ids],
                    "dramas": [
                        {
                            "name": d.get("name", ""),
                            "click": d.get("click", ""),
                            "show": d.get("show", ""),
                            "video": d.get("video", ""),
                            "material_ids": list(d.get("material_ids", [])),
                        }
                        for d in dramas
                    ],
                })
            prof["groups"] = new_groups
        changed = True
    if changed:
        save_config(cfg)
    return changed


def sanitize_config_groups(groups):
    cleaned_groups = []
    changed = False
    for group in groups or []:
        ids = [str(x).strip() for x in (group.get("account_ids") or []) if str(x).strip().isdigit()]
        dramas = []
        for d in group.get("dramas") or []:
            name = str(d.get("name") or "").strip()
            if not name or is_separator_line(name):
                changed = True
                continue
            if re.search(r"https?://", name, re.I):
                _, recovered = _parse_single_group(name, None)
                if recovered:
                    dramas.extend(recovered)
                changed = True
                continue
            item = {
                "name": name.splitlines()[0].strip(),
                "click": sanitize_link_text(d.get("click") or ""),
                "show": sanitize_link_text(d.get("show") or ""),
                "video": sanitize_link_text(d.get("video") or ""),
                "material_ids": [str(x).strip() for x in (d.get("material_ids") or []) if str(x).strip()],
            }
            if item["name"] != name or item["click"] != (d.get("click") or "") or item["show"] != (d.get("show") or "") or item["video"] != (d.get("video") or ""):
                changed = True
            dramas.append(item)
        if ids or dramas:
            entry = {"account_ids": ids, "dramas": dramas}
            for extra_key in ("group_name", "click_url", "show_url", "play_url"):
                if extra_key in group:
                    entry[extra_key] = str(group[extra_key]).strip()
            cleaned_groups.append(entry)
    return cleaned_groups, changed


def build_runtime_profile_config(profile_key: str, app_cfg: dict | None = None) -> dict:
    app_cfg = app_cfg or load_config()
    cfg = dict(ALL_PROFILES[profile_key])
    user_prof = (app_cfg.get("profiles") or {}).get(profile_key) or {}
    for f in PROFILE_EDITABLE_FIELDS:
        if f in user_prof and user_prof[f] not in (None, ""):
            cfg[f] = user_prof[f]
    try:
        cfg["wait_scale"] = float(cfg["wait_scale"])
    except Exception:
        cfg["wait_scale"] = ALL_PROFILES[profile_key]["wait_scale"]
    return cfg


def profile_groups_from_config(cfg: dict, profile_key: str):
    prof = (cfg.get("profiles") or {}).get(profile_key) or {}
    groups = prof.get("groups") or []
    is_incentive = ALL_PROFILES.get(profile_key, {}).get("build_mode") == "incentive"
    groups, _ = sanitize_config_groups(groups)
    out = []
    for g in groups:
        ids = [str(x).strip() for x in (g.get("account_ids") or []) if str(x).strip()]
        if is_incentive:
            meta = {
                "group_name": g.get("group_name", ""),
                "click_url": g.get("click_url", ""),
                "show_url": g.get("show_url", ""),
                "play_url": g.get("play_url", ""),
            }
            if ids:
                out.append((ids, [], meta))
        else:
            dramas = []
            for d in (g.get("dramas") or []):
                name = (d.get("name") or "").strip()
                if not name:
                    continue
                dramas.append({
                    "name": name,
                    "click": (d.get("click") or "").strip(),
                    "show": (d.get("show") or "").strip(),
                    "video": (d.get("video") or "").strip(),
                    "material_ids": [str(x).strip() for x in (d.get("material_ids") or []) if str(x).strip()],
                })
            if ids or dramas:
                out.append((ids, dramas))
    return out


# ═══════════════════════════════════════════════════════════════
#  ids.txt 模板内容
# ═══════════════════════════════════════════════════════════════
IDS_TEMPLATE = """\
# ════════════════════════════════════════════════════════════
# 红果搭建工具 - ids.txt 数据模板
# ════════════════════════════════════════════════════════════
# 使用说明：
#   1. 顶部依次填写媒体账户ID（每行一个纯数字）
#   2. 空行后开始第一组数据：
#        - 剧名
#        - 点击监测链接（含 action_type=click 或 /display/）
#        - 展示监测链接（含 impression 或 action_type=view）
#        - 视频播放监测链接（含 action_type=effective_play）
#        - 素材ID（多个用空格分隔）
#   3. 多组数据之间用 === 分隔
#   4. 以 # 开头的行为注释，会被忽略
# ════════════════════════════════════════════════════════════

# ===== 在下方填写媒体账户ID（每行一个）=====
1234567890123456
1234567890123457

# ===== 第 1 组数据 =====
示例剧名

https://example.com/click?action_type=click&xxx

https://example.com/show?action_type=view&xxx

https://example.com/play?action_type=effective_play&xxx

7000000000001 7000000000002 7000000000003

===

# ===== 第 2 组数据（可继续添加，删除示例后保留格式即可）=====
"""


# ═══════════════════════════════════════════════════════════════
#  初始化数据目录与模板文件
# ═══════════════════════════════════════════════════════════════
def init_data_dirs():
    """首次运行时自动创建 4 个方向的文件夹和 ids.txt 模板。

    返回创建的文件列表，便于在 GUI 中提示用户。
    """
    created = []
    for key, cfg in ALL_PROFILES.items():
        ids_file: Path = cfg["ids_file"]
        log_dir: Path = cfg["log_dir"]
        # 创建 ids.txt 所在目录
        ids_file.parent.mkdir(parents=True, exist_ok=True)
        # 创建 logs 目录
        log_dir.mkdir(parents=True, exist_ok=True)
        # 不存在时写入模板
        if not ids_file.exists():
            ids_file.write_text(IDS_TEMPLATE, encoding="utf-8")
            created.append(str(ids_file))
    return created


# ═══════════════════════════════════════════════════════════════
#  通用常量
# ═══════════════════════════════════════════════════════════════
TIMEOUT = 60_000
RE_CONFIRM = re.compile(r"确\s*定|确定|确\s*认|确认")
RE_MMDD = re.compile(r'(?<!\d)(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)')
TODAY_STR = datetime.now().strftime("%m%d")


# ═══════════════════════════════════════════════════════════════
#  等待时间（基准值，运行时乘以 wait_scale）
# ═══════════════════════════════════════════════════════════════
BASE_WAITS = dict(
    TINY=200, SHORT=300, MEDIUM=500, NORMAL=800,
    LONG=1000, LONGER=1200, EXTRA=1500, LOAD=2000,
    HEAVY=2500, SEARCH=4000,
)


class WaitTimes:
    def __init__(self, scale=1.0):
        for k, v in BASE_WAITS.items():
            setattr(self, k, max(80, int(v * scale)))


# ═══════════════════════════════════════════════════════════════
#  异常
# ═══════════════════════════════════════════════════════════════
class AccountsMissingError(Exception):
    def __init__(self, missing_ids, found_count, input_count):
        self.missing_ids = list(missing_ids)
        self.found_count = found_count
        self.input_count = input_count
        super().__init__(
            f"媒体账户缺失: 输入 {input_count} 个，搜到 {found_count} 个，"
            f"缺失 {len(self.missing_ids)} 个: {self.missing_ids}"
        )


class StopRequested(Exception):
    pass


def check_stop(stop_event):
    if stop_event and stop_event.is_set():
        raise StopRequested()


# ═══════════════════════════════════════════════════════════════
#  日志
# ═══════════════════════════════════════════════════════════════
def setup_logger(log_dir: Path):
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"搭建_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    lg = logging.getLogger("build")
    lg.setLevel(logging.DEBUG)
    lg.handlers.clear()
    lg.propagate = False

    console_fmt = logging.Formatter("[%(asctime)s] [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
    file_fmt = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] [%(funcName)s:%(lineno)d] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(console_fmt)
    lg.addHandler(ch)

    try:
        fh = logging.FileHandler(log_file, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(file_fmt)
        lg.addHandler(fh)
        lg.info(f"📝 日志文件: {log_file}")
    except Exception as e:
        lg.warning(f"⚠️ 日志文件初始化失败: {e}")

    return lg


def fmt_duration(seconds: float) -> str:
    if seconds < 60:
        return f"{seconds:.1f}秒"
    m, s = divmod(int(seconds), 60)
    if m < 60:
        return f"{m}分{s}秒"
    h, m = divmod(m, 60)
    return f"{h}时{m}分{s}秒"


# ═══════════════════════════════════════════════════════════════
#  数据解析
# ═══════════════════════════════════════════════════════════════
def is_separator_line(text: str) -> bool:
    value = (text or "").strip()
    if not value:
        return False
    if re.fullmatch(r"[=═\-—＿_\s]{3,}", value):
        return True
    if re.fullmatch(r"[=═\-—＿_\s]*第\s*\d+\s*组.*[=═\-—＿_\s]*", value):
        return True
    if re.fullmatch(r"[=═\-—＿_\s]*链接分配之后.*[=═\-—＿_\s]*", value):
        return True
    return False


def sanitize_link_text(text: str) -> str:
    value = (text or "").replace("\r", "").replace("\n", "").strip()
    value = value.strip("`'\" ")
    value = re.split(r"\s*(?:={3,}|═{3,})", value, maxsplit=1)[0].strip()
    value = value.strip("`'\" ")
    return value


def normalize_link(text: str) -> str:
    return sanitize_link_text(text)


def classify_link(url: str) -> str:
    u = normalize_link(url).lower()
    if "action_type=effective_play" in u: return "video"
    if "action_type=click" in u: return "click"
    if "action_type=view" in u: return "show"
    if "effective_play" in u: return "video"
    if "/display/" in u: return "click"
    if "impression" in u: return "show"
    return "unknown"


@functools.lru_cache(maxsize=64)
def _compile_sequel_pattern(drama_name: str):
    return re.compile(
        rf"{re.escape(drama_name)}\s*("
        rf"[0-9一二三四五六七八九十]+|"
        rf"第[一二三四五六七八九十]+(部|季)|"
        rf"[二三四五六七八九十]+(部|季)|"
        rf"之)"
    )


def _normalize_material_text(text: str) -> str:
    return re.sub(r"\s+", "", text or "")


def _material_name_has_exact_drama_segment(drama_name: str, material_name: str) -> bool:
    drama = _normalize_material_text(drama_name)
    if not drama:
        return False
    stem = Path(material_name or "").stem
    parts = [p for p in re.split(r"[-－—–_＿]+", stem) if p.strip()]
    for part in parts:
        cleaned = re.sub(r"[(\（]\d+[)\）]$", "", part.strip())
        if _normalize_material_text(cleaned) == drama:
            return True
    return False


def is_valid_material_name(drama_name: str, material_name: str) -> bool:
    if not _material_name_has_exact_drama_segment(drama_name, material_name):
        return False
    if _compile_sequel_pattern(drama_name).search(material_name):
        return False
    return True


def extract_mmdd(name: str):
    m = RE_MMDD.search(name or "")
    if not m:
        return None
    return f"{m.group(1)}{m.group(2)}"


def _is_url_line(line: str) -> bool:
    return bool(re.match(r"^`?\s*https?://", (line or "").strip(), re.I))


def _clean_plain_line(line: str) -> str:
    return (line or "").strip().strip("`'\" ")


def _is_material_ids_line(line: str) -> bool:
    parts = (line or "").split()
    return bool(parts) and len(parts) > 1 and all(p.isdigit() for p in parts)


def _parse_single_group(raw_text, logger):
    raw_text = re.sub(r"(?:={3,}|═{3,})", "\n", raw_text or "")
    lines = []
    for line in raw_text.splitlines():
        clean = _clean_plain_line(line)
        if not clean or is_separator_line(clean):
            continue
        lines.append(clean)
    if not lines:
        return [], []

    ids = []
    idx = 0
    while idx < len(lines):
        clean = _clean_plain_line(lines[idx])
        if clean.isdigit():
            ids.append(clean)
            idx += 1
            continue
        break

    drama_list = []
    current = None

    def log_info(msg):
        if logger:
            logger.info(msg)

    def log_warning(msg):
        if logger:
            logger.warning(msg)

    def flush_current():
        nonlocal current
        if not current:
            return
        missing_types = [k for k in ("click", "show", "video") if not current[k]]
        if missing_types:
            log_warning(f"⚠️ 链接类型缺失: {current['name']} | 缺少: {', '.join(missing_types)}")
        log_info(f"📘 剧名解析: {current['name']}")
        drama_list.append(current)
        current = None

    while idx < len(lines):
        clean = _clean_plain_line(lines[idx])
        if not clean or is_separator_line(clean):
            idx += 1
            continue

        if _is_url_line(clean):
            if not current:
                log_warning(f"⚠️ 遇到无剧名链接，已跳过: {sanitize_link_text(clean)[:80]}")
                idx += 1
                continue
            link = sanitize_link_text(clean)
            link_type = classify_link(link)
            if link_type == "unknown":
                log_warning(f"⚠️ 未识别链接类型，已跳过: {link[:80]}")
            elif current[link_type]:
                log_warning(f"⚠️ 重复{link_type}链接: {current['name']}")
            else:
                current[link_type] = link
            idx += 1
            continue

        if current and _is_material_ids_line(clean):
            current["material_ids"] = clean.split()
            log_info(f"📦 {current['name']} 素材ID数量: {len(current['material_ids'])}")
            idx += 1
            continue

        if clean.isdigit():
            if current:
                current["material_ids"].append(clean)
            idx += 1
            continue

        flush_current()
        current = {"name": clean, "click": "", "show": "", "video": "", "material_ids": []}
        idx += 1

    flush_current()
    return ids, drama_list


def read_data(id_file: Path, logger):
    if not id_file.exists():
        raise FileNotFoundError(f"找不到文件: {id_file}")
    raw_text = id_file.read_text(encoding="utf-8")
    # 过滤掉以 # 开头的注释行（保留空行用于段落分隔）
    raw_text = "\n".join(
        line for line in raw_text.splitlines()
        if not line.lstrip().startswith("#")
    )
    chunks = re.split(r'\s*\n\s*={3,}\s*\n\s*', raw_text.strip())
    groups = []
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk: continue
        ids, dramas = _parse_single_group(chunk, logger)
        if ids or dramas:
            groups.append((ids, dramas))
    total_ids = sum(len(g[0]) for g in groups)
    total_drama = sum(len(g[1]) for g in groups)
    logger.info(f"✅ 读取到 {len(groups)} 组数据，共 {total_ids} 个账号, {total_drama} 部剧")
    return groups


# ═══════════════════════════════════════════════════════════════
#  Playwright 通用工具
# ═══════════════════════════════════════════════════════════════
def safe_click(popup, locator, *, timeout=TIMEOUT, retries=3, desc="", logger=None, W=None):
    for attempt in range(1, retries + 1):
        try:
            locator.wait_for(state="visible", timeout=timeout)
            locator.scroll_into_view_if_needed()
            popup.wait_for_timeout(W.TINY if W else 200)
            locator.click(force=True)
            if attempt > 1 and logger:
                logger.info(f"🔁 点击在第{attempt}次重试后成功({desc})")
            return
        except PlaywrightTimeout:
            if logger: logger.warning(f"⏰ 点击超时({desc}) 第{attempt}/{retries}次")
        except Exception as e:
            if logger: logger.warning(f"⚠️ 点击失败({desc}) 第{attempt}/{retries}次: {e}")
        if attempt < retries:
            popup.wait_for_timeout(W.LONG if W else 1000)
    raise Exception(f"❌ 点击最终失败({desc})，已重试{retries}次")


def _locator_count(locator):
    try: return locator.count()
    except: return 0


def wait_loading_gone(popup, container, *, timeout=30_000):
    mask = container.locator(".el-loading-mask").first
    if mask.count() > 0:
        try: mask.wait_for(state="hidden", timeout=timeout)
        except PlaywrightTimeout: pass


def wait_locator_ready(popup, locator, *, timeout=TIMEOUT, desc="元素", W=None):
    locator.wait_for(state="visible", timeout=timeout)
    try: locator.scroll_into_view_if_needed()
    except: pass
    wait_small(popup, W.TINY if W else 200)
    return locator


def wait_idle(target, *, mask_timeout=8_000, network=False, network_timeout=3_000):
    try:
        mask = target.locator(".el-loading-mask:visible")
        if mask.count() > 0:
            mask.first.wait_for(state="hidden", timeout=mask_timeout)
    except: pass
    if network:
        try: target.wait_for_load_state("networkidle", timeout=network_timeout)
        except: pass


def wait_small(popup, ms=300):
    try:
        mask = popup.locator(".el-loading-mask:visible")
        if mask.count() > 0:
            mask.first.wait_for(state="hidden", timeout=max(ms, 1500))
            return
    except: pass
    popup.wait_for_timeout(min(ms, max(80, ms // 2)))


def _safe_page_title(page):
    try:
        return page.title()
    except Exception:
        return ""


def _safe_page_url(page):
    try:
        return page.url
    except Exception:
        return ""


def _is_browser_internal_page(url):
    value = (url or "").lower()
    return value.startswith(("about:", "chrome:", "devtools:", "edge:", "trae:"))


def select_build_page(context, logger):
    pages = list(context.pages)
    logger.info(f"🔎 当前 CDP 连接到 {len(pages)} 个页面")
    usable_pages = []
    for idx, pg in enumerate(pages, 1):
        url = _safe_page_url(pg)
        title = _safe_page_title(pg)
        logger.info(f"  页面{idx}: {title or '无标题'} | {url or '无URL'}")
        if not _is_browser_internal_page(url):
            usable_pages.append(pg)

    candidates = []
    for pg in usable_pages:
        url = _safe_page_url(pg).lower()
        title = _safe_page_title(pg).lower()
        score = 0
        if "qianchuan" in url or "ad.oceanengine" in url or "巨量" in title or "千川" in title:
            score += 80
        if "promotion" in url or "campaign" in url or "project" in url or "广告" in title or "计划" in title or "推广" in title:
            score += 30
        try:
            if pg.locator("button:has-text('批量新建')").count() > 0:
                score += 70
            elif pg.locator("button:has-text('新建')").count() > 0:
                score += 30
        except Exception:
            pass
        candidates.append((score, pg))

    candidates.sort(key=lambda item: item[0], reverse=True)
    if candidates and candidates[0][0] > 0:
        page = candidates[0][1]
    elif usable_pages:
        page = usable_pages[-1]
    elif pages:
        page = pages[-1]
    else:
        raise RuntimeError("没有检测到可控制的浏览器页面，请确认浏览器已用 9222 调试端口启动")

    page.set_default_timeout(15_000)
    try:
        page.bring_to_front()
    except Exception:
        pass
    logger.info(f"🎯 已选择操作页面：{_safe_page_title(page) or '无标题'} | {_safe_page_url(page) or '无URL'}")
    return page


def get_visible_layer(popup, *, desc="弹窗", timeout=15_000, logger=None, W=None):
    selectors = [
        "div.el-dialog__wrapper:visible",
        "div.el-dialog:visible",
        "div.mg-dialog-wrapper:visible",
        "div.cl-drawer:visible",
        "div.drawer-content:visible",
        "div[role='dialog']:visible",
        ".arco-modal:visible",
        ".arco-drawer:visible",
    ]
    deadline = time.time() + timeout / 1000
    while time.time() < deadline:
        for sel in selectors:
            layers = popup.locator(sel)
            n = _locator_count(layers)
            if n > 0:
                if logger:
                    logger.info(f"✅ 已识别{desc}：{sel}，数量 {n}")
                return layers.nth(n - 1)
        wait_small(popup, W.TINY if W else 200)
    details = []
    for sel in selectors:
        try:
            details.append(f"{sel}={popup.locator(sel).count()}")
        except Exception:
            details.append(f"{sel}=?")
    raise PlaywrightTimeout(f"等待{desc}超时，未匹配到可见弹层；{' | '.join(details)}")


def get_visible_drawer(popup):
    drawer = popup.locator("div.drawer-content:visible").last
    wrap = drawer.locator("div.el-scrollbar__wrap").first
    return drawer, wrap


def scroll_wrap_to_bottom(popup, wrap, W):
    if wrap.count() > 0:
        try: wrap.evaluate("wrap => wrap.scrollTop = wrap.scrollHeight")
        except: pass
    wait_small(popup, W.SHORT)


def scroll_to_module(popup, wrap, module_id, W):
    if wrap.count() > 0:
        wrap.evaluate(
            """(wrap, id) => {
                const el = wrap.querySelector('#' + id);
                if (el) wrap.scrollTop = el.offsetTop - 80;
            }""", module_id)
    else:
        popup.locator(f"#{module_id}").scroll_into_view_if_needed()
    wait_small(popup, 250)


def click_top_confirm(popup, scope=None, *, desc="确认按钮", timeout=TIMEOUT, wait_close=False, logger=None, W=None):
    deadline = time.time() + timeout / 1000
    while time.time() < deadline:
        scopes = []
        if scope is not None:
            scopes.append(scope)
            pd = scope.locator("xpath=ancestor::div[contains(concat(' ', normalize-space(@class), ' '), ' cl-drawer ')][1]")
            if _locator_count(pd) > 0: scopes.append(pd.first)
            pw2 = scope.locator("xpath=ancestor::div[contains(concat(' ', normalize-space(@class), ' '), ' mg-dialog-wrapper ')][1]")
            if _locator_count(pw2) > 0: scopes.append(pw2.first)
        else:
            for sel in ["div.el-dialog__wrapper:visible", "div.mg-dialog-wrapper:visible",
                        "div.cl-drawer:visible", "div.el-dialog:visible"]:
                layers = popup.locator(sel)
                n = _locator_count(layers)
                if n > 0: scopes.append(layers.nth(n - 1))

        for sc in scopes:
            for candidates in [
                sc.locator("button.el-button--primary:not(.is-disabled):visible").filter(has_text=RE_CONFIRM),
                sc.locator("button:not(.is-disabled):visible").filter(has_text=RE_CONFIRM),
            ]:
                if _locator_count(candidates) > 0:
                    btn = candidates.last
                    safe_click(popup, btn, desc=desc, logger=logger, W=W)
                    if wait_close:
                        try: sc.wait_for(state="hidden", timeout=4_000)
                        except: pass
                    return
        wait_small(popup, W.TINY if W else 200)
    raise Exception(f"❌ 当前范围内未找到可点击的确认按钮: {desc}")


def _visible_confirm_count(popup):
    selectors = [
        "div.operate-button:visible button.el-button--primary:visible",
        "div.operate-button:visible button:visible",
        "button.el-button--primary.el-button--small:visible",
        "button.el-button--primary:visible",
        "button:visible",
    ]
    total = 0
    for sel in selectors:
        total += _locator_count(popup.locator(sel).filter(has_text=RE_CONFIRM))
    return total


def _click_confirm_button_hard(popup, button, *, desc, logger=None, W=None):
    try:
        button.wait_for(state="visible", timeout=2_000)
    except Exception:
        pass
    try:
        button.scroll_into_view_if_needed()
    except Exception:
        pass
    wait_small(popup, W.TINY if W else 200)
    methods = ["普通点击", "强制点击", "JS点击", "坐标点击"]
    last_error = None
    for method in methods:
        try:
            if method == "普通点击":
                button.click(timeout=3_000)
            elif method == "强制点击":
                button.click(force=True, timeout=3_000)
            elif method == "JS点击":
                button.evaluate("el => el.click()")
            else:
                box = button.bounding_box()
                if not box:
                    continue
                popup.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
            if logger:
                logger.info(f"🖱️ {desc}已执行{method}")
            wait_small(popup, W.NORMAL if W else 500)
            if _visible_confirm_count(popup) == 0:
                return True
        except Exception as e:
            last_error = e
            if logger:
                logger.warning(f"⚠️ {desc}{method}失败: {e}")
    if logger and last_error:
        logger.warning(f"⚠️ {desc}所有点击方式后仍未关闭: {last_error}")
    return _visible_confirm_count(popup) == 0


def click_optional_confirm(popup, *, desc="可选确认按钮", timeout=6_000, logger=None, W=None):
    deadline = time.time() + timeout / 1000
    direct_selectors = [
        "div.operate-button:visible button.el-button--primary:visible",
        "div.operate-button:visible button:visible",
        "button.el-button--primary.el-button--small:visible",
        "button.el-button--primary:visible",
        "button:visible",
    ]
    while time.time() < deadline:
        for sel in direct_selectors:
            candidates = popup.locator(sel).filter(has_text=RE_CONFIRM)
            count = _locator_count(candidates)
            if count > 0:
                btn = candidates.last
                ok = _click_confirm_button_hard(popup, btn, desc=desc, logger=logger, W=W)
                wait_idle(popup, mask_timeout=5_000)
                if ok:
                    if logger:
                        logger.info(f"✅ 已点击{desc}并确认弹层消失")
                    wait_small(popup, W.NORMAL if W else 500)
                    return True
                if logger:
                    logger.warning(f"⚠️ 已尝试点击{desc}，但按钮仍可见，继续重试")
        try:
            click_top_confirm(popup, desc=desc, timeout=500, wait_close=True, logger=logger, W=W)
            wait_idle(popup, mask_timeout=5_000)
            if _visible_confirm_count(popup) == 0:
                if logger:
                    logger.info(f"✅ 已点击{desc}并确认弹层消失")
                wait_small(popup, W.NORMAL if W else 500)
                return True
        except Exception:
            pass
        wait_small(popup, W.TINY if W else 200)
    if logger:
        logger.warning(f"⚠️ {desc}点击后仍未消失，后续会继续等待素材弹窗关闭")
    return False


def safe_select_option(popup, trigger_locator, option_text, *, desc="", logger=None, W=None):
    try:
        safe_click(popup, trigger_locator, desc=f"{desc}下拉触发", logger=logger, W=W)
        dropdown = popup.locator("ul.el-select-dropdown__list:visible").last
        dropdown.wait_for(state="visible", timeout=TIMEOUT)
        option = dropdown.locator("li.el-select-dropdown__item").filter(has_text=option_text).first
        if option.count() > 0:
            safe_click(popup, option, desc=f"{desc}选择'{option_text}'", logger=logger, W=W)
            return True
        return False
    except:
        return False


# ═══════════════════════════════════════════════════════════════
#  搭建步骤函数（参数化）
# ═══════════════════════════════════════════════════════════════
def step_select_strategy(popup, cfg, logger, W):
    strategy_name = cfg["strategy"]
    logger.info(f"🔍 查找“选择策略”按钮，目标策略：{strategy_name}")
    strategy_buttons = popup.locator("button:has-text('选择策略')")
    try:
        logger.info(f"🔘 匹配到“选择策略”按钮：{strategy_buttons.count()} 个")
    except Exception:
        pass
    safe_click(popup, strategy_buttons.first, timeout=15_000, desc="选择策略按钮", logger=logger, W=W)
    wait_small(popup, W.MEDIUM)
    strategy_dlg = get_visible_layer(popup, desc="策略弹窗", timeout=15_000, logger=logger, W=W)
    wait_small(popup, W.LOAD)

    rows = strategy_dlg.locator("tbody tr.el-table__row, tbody tr, .el-table__row, tr")
    strategy_row = None
    last_seen = []
    deadline = time.time() + TIMEOUT / 1000
    while time.time() < deadline and strategy_row is None:
        try: rows.first.wait_for(state="visible", timeout=3_000)
        except PlaywrightTimeout:
            wait_small(popup, W.MEDIUM); continue
        row_count = rows.count()
        last_seen = []
        for i in range(row_count):
            row = rows.nth(i)
            cells = row.locator("td")
            try:
                row_text = re.sub(r"\s+", " ", row.inner_text()).strip()
                if row_text: last_seen.append(row_text)
            except: pass
            for j in range(cells.count()):
                if cells.nth(j).inner_text().strip() == strategy_name:
                    strategy_row = row; break
            if strategy_row: break
        if not strategy_row: wait_small(popup, W.NORMAL)

    if not strategy_row:
        try:
            snapshot = re.sub(r"\s+", " ", strategy_dlg.inner_text(timeout=2_000)).strip()
            logger.warning(f"⚠️ 策略弹窗内容片段：{snapshot[:500]}")
        except Exception:
            pass
        raise Exception(f"❌ 未找到策略: {strategy_name}")

    strategy_row.wait_for(state="visible", timeout=TIMEOUT)
    strategy_row.locator("label.el-radio").first.click(force=True)
    wait_small(popup, W.SHORT)
    click_top_confirm(popup, strategy_dlg, desc="策略确认", wait_close=True, logger=logger, W=W)
    wait_small(popup, W.NORMAL)
    logger.info(f"📌 已选择策略: {strategy_name}")


def step_select_media_accounts(popup, ids, cfg, logger, W):
    popup.locator("div.selector:has(label:has-text('媒体账户')) button:has-text('更改')").click()
    wait_small(popup, W.LOAD)
    popup.locator("div.selected-card-header").first.wait_for(state="visible", timeout=TIMEOUT)
    popup.locator("div.selected-card-header button:has-text('清空')").first.click(force=True)
    wait_small(popup, W.MEDIUM)
    popup.get_by_label("选择媒体账户").get_by_title("批量搜索").click(force=True)

    id_input = popup.locator("input[placeholder='请粘贴或输入账户ID，回车可换行']")
    id_input.wait_for(state="visible", timeout=TIMEOUT)
    id_input.click()
    for _id in ids:
        popup.keyboard.type(_id)
        popup.keyboard.press("Enter")
    logger.info(f"📝 已输入 {len(ids)} 个媒体账户ID")

    popup.locator("button.el-button--primary:visible").filter(has_text="搜索").last.click(force=True)
    wait_small(popup, W.SEARCH)

    dlg_acc = popup.locator("div.el-dialog__wrapper:visible").last
    try: dlg_acc.locator("tbody tr").first.wait_for(state="visible", timeout=TIMEOUT)
    except PlaywrightTimeout: raise AccountsMissingError(list(ids), 0, len(ids))
    wait_small(popup, W.MEDIUM)

    # 检查账户行
    rows = dlg_acc.locator("div.el-table__fixed-body-wrapper tbody tr.el-table__row")
    row_count = rows.count()
    found_ids = set()
    for i in range(row_count):
        row = rows.nth(i)
        id_el = row.locator("td:nth-child(2) p").filter(has_text="ID：")
        if id_el.count() > 0:
            try:
                row_text = id_el.first.inner_text().strip()
                match = re.search(r'ID[：:]\s*(\d+)', row_text)
                if match: found_ids.add(match.group(1))
            except: pass
    missing = [_id for _id in ids if _id not in found_ids]
    if row_count < len(ids) or missing:
        if not missing: missing = [f"(未知缺失{len(ids) - row_count}个)"]
        raise AccountsMissingError(missing, row_count, len(ids))

    dlg_acc.locator("thead label.el-checkbox").first.click(force=True)
    click_top_confirm(popup, dlg_acc, desc="媒体账户确认", wait_close=True, logger=logger, W=W)
    wait_small(popup, W.LOAD)
    try: dlg_acc.wait_for(state="hidden", timeout=10_000)
    except: pass
    wait_small(popup, W.NORMAL)


def step_link_product(popup, drama_name, cfg, logger, W):
    try:
        popup.wait_for_load_state("networkidle", timeout=5_000)
    except PlaywrightTimeout:
        logger.warning("⚠️ 关联产品前等待网络空闲超时，继续执行")
    wait_small(popup, W.LONG)
    edit_btn = popup.locator("tfoot td button:has-text('编辑')").nth(0)
    edit_btn.wait_for(state="visible", timeout=TIMEOUT)
    edit_btn.scroll_into_view_if_needed()
    wait_small(popup, W.MEDIUM)
    edit_btn.click()
    wait_small(popup, W.LOAD)

    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "link-product", W)
    safe_click(popup, popup.locator("#link-product button:has-text('选择产品')"), desc="选择产品按钮", logger=logger, W=W)
    prod_dlg = popup.locator("div.el-dialog__wrapper:visible").last
    prod_dlg.wait_for(state="visible", timeout=TIMEOUT)
    safe_click(popup, prod_dlg.locator("button.el-button--text.el-button--mini:has-text('清空')").first, desc="清空产品", logger=logger, W=W)
    wait_small(popup, W.LONG)
    p_search = prod_dlg.locator("input[placeholder='请输入关键词']").first
    p_search.fill(drama_name)
    wait_small(popup, W.MEDIUM)
    safe_click(popup, prod_dlg.locator("button:has-text('查询')"), desc="查询产品", logger=logger, W=W)
    wait_small(popup, W.HEAVY)

    target_l = prod_dlg.locator(f"div.el-checkbox-group label.el-checkbox:has(div.product-name:has-text('{drama_name}'))")
    if target_l.count() == 0:
        p_search.fill("")
        safe_click(popup, prod_dlg.locator("button:has-text('查询')"), desc="查询全部产品", logger=logger, W=W)
        wait_small(popup, W.HEAVY)
        all_p = prod_dlg.locator("div.el-checkbox-group label.el-checkbox")
        total = all_p.count()
        if total == 0:
            logger.warning(f"⚠️ 未找到任何产品: {drama_name}"); target = None
        else:
            random_idx = random.randint(0, total - 1)
            target = all_p.nth(random_idx)
            logger.info(f"🎯 未匹配剧名『{drama_name}』，随机选中第 {random_idx + 1}/{total} 个产品")
    else:
        target = target_l.first
        logger.info(f"🎯 已匹配到剧名对应产品: {drama_name}")

    if target is not None:
        safe_click(popup, target, desc="选择产品项", logger=logger, W=W)
        click_top_confirm(popup, prod_dlg, desc="产品确认", wait_close=True, logger=logger, W=W)
        wait_small(popup, W.LONG)


def step_fill_monitor_links(popup, drama, cfg, logger, W):
    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "goal", W)

    safe_click(popup,
        popup.locator(f"#select-track-group button:has-text('{cfg['monitor_btn_text']}')"),
        desc=cfg['monitor_btn_text'], logger=logger, W=W)

    popup.locator("text=手动输入监测链接").first.wait_for(state="visible", timeout=TIMEOUT)
    monitor_drawer, monitor_wrap = get_visible_drawer(popup)

    def switch_on(label):
        item = monitor_drawer.locator(f"div.el-form-item:has(label.el-form-item__label:text-is('{label}'))").first
        if item.count() > 0:
            item.locator("label.el-radio-button:has-text('开启')").first.click(force=True)
            wait_small(popup, W.MEDIUM)

    def fill_link(label, value):
        item = monitor_drawer.locator(f"div.el-form-item:has(label.el-form-item__label:has-text('{label}'))").first
        input_box = item.locator("input.el-input__inner:visible").first
        input_box.wait_for(state="visible", timeout=TIMEOUT)
        input_box.click(force=True)
        wait_small(popup, W.TINY)
        input_box.fill(sanitize_link_text(value))
        wait_small(popup, W.SHORT)
        try:
            actual = input_box.input_value().strip()
        except Exception:
            actual = ""
        if not actual:
            raise Exception(f"{label}未填写成功")

    switch_on('展示链接'); switch_on('有效触点链接'); switch_on('视频有效播放链接')
    fill_link('请输入展示链接', drama["show"])
    logger.info("  展示链接☑️")
    fill_link('请输入有效触点链接', drama["click"])
    logger.info("  监测链接☑️")
    fill_link('请输入视频有效播放链接', drama["video"])
    logger.info("  播放链接☑️")
    logger.info(f"  ✅ 三个链接已填写完成")

    scroll_wrap_to_bottom(popup, monitor_wrap, W)
    click_top_confirm(popup, logger=logger, W=W)
    wait_small(popup, W.LONG)


def step_select_audience_package(popup, cfg, logger, W):
    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "audience-package", W)
    safe_click(popup, popup.locator("#audience-package button:has-text('选择定向包')"), desc="选择定向包按钮", logger=logger, W=W)

    search_box = popup.locator(".cl-search-input:visible").first
    wait_locator_ready(popup, search_box, desc="定向包搜索区", W=W)
    wait_loading_gone(popup, popup.locator("body"))
    keyword_input = search_box.locator("input.el-input__inner[placeholder='请输入关键词']")

    if keyword_input.count() > 0:
        try:
            wait_locator_ready(popup, keyword_input, desc="定向包关键词输入框", W=W)
            keyword_input.click(force=True)
            wait_small(popup, W.TINY)
            keyword_input.fill(cfg["audience_keyword"])
            wait_small(popup, W.SHORT)
            search_btn = search_box.locator("button[title='搜索']").first
            if search_btn.count() > 0:
                wait_locator_ready(popup, search_btn, desc="定向包搜索按钮", W=W)
                search_btn.click(force=True)
                wait_loading_gone(popup, popup.locator("body"))
                wait_small(popup, W.NORMAL)
        except Exception as e:
            logger.warning(f"⚠️ 定向包搜索出错: {e}")

    btn_m = popup.locator("button:has-text('多账户快速选择'):visible").last
    wait_locator_ready(popup, btn_m, desc="多账户快速选择按钮", W=W)
    safe_click(popup, btn_m, desc="多账户快速选择(第1次)", logger=logger, W=W)
    wait_small(popup, W.MEDIUM)
    wait_locator_ready(popup, btn_m, desc="多账户快速选择按钮", W=W)
    safe_click(popup, btn_m, desc="多账户快速选择(第2次)", logger=logger, W=W)
    wait_loading_gone(popup, popup.locator("body"))
    wait_small(popup, W.EXTRA)
    click_top_confirm(popup, desc="定向包确认", logger=logger, W=W)
    wait_small(popup, W.LONG)


def step_fill_project_name(popup, drama_name, cfg, logger, W):
    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "project-name", W)
    project_name = f"{cfg['name_prefix']}-{drama_name}-lzp-<日期>"
    name_input = popup.locator("#project-name input.el-input__inner").first
    name_input.wait_for(state="visible", timeout=TIMEOUT)
    name_input.click(force=True); name_input.fill("")
    wait_small(popup, W.TINY)
    name_input.fill(project_name)
    wait_small(popup, W.SHORT)

    try: actual_value = name_input.input_value().strip()
    except: actual_value = ""
    if actual_value != project_name:
        name_input.click(force=True); name_input.fill("")
        wait_small(popup, W.TINY); name_input.fill(project_name)
        wait_small(popup, W.SHORT)

    logger.info(f"📝 项目名称: {project_name}")
    name_input.press("Tab"); wait_small(popup, W.SHORT)
    scroll_wrap_to_bottom(popup, wrap, W)

    drawer_confirm = drawer.locator("button.el-button--primary:not(.is-disabled):visible").filter(has_text=RE_CONFIRM).last
    if _locator_count(drawer_confirm) == 0:
        drawer_confirm = popup.locator("div.drawer-content:visible").last.locator("button:not(.is-disabled):visible").filter(has_text=RE_CONFIRM).last
    if _locator_count(drawer_confirm) > 0:
        safe_click(popup, drawer_confirm, desc="项目名称确定按钮", logger=logger, W=W)
    else:
        click_top_confirm(popup, desc="项目名称确定按钮", logger=logger, W=W)
    wait_idle(popup, mask_timeout=W.LOAD)


def step_fill_ad_name(popup, drama_name, cfg, logger, W):
    promo_block = popup.locator("div.module-container#promotion-name")
    edit_btn = popup.locator("tfoot td button:has-text('编辑')").nth(1)
    edit_btn.wait_for(state="visible", timeout=TIMEOUT)
    edit_btn.scroll_into_view_if_needed()

    last_err = None
    for _attempt in range(5):
        try: popup.keyboard.press("Escape")
        except: pass
        popup.wait_for_timeout(W.SHORT)
        try: edit_btn.click(force=True)
        except Exception as e:
            last_err = e; popup.wait_for_timeout(W.LONG); continue
        try: promo_block.first.wait_for(state="visible", timeout=8_000); break
        except Exception as e:
            last_err = e; popup.wait_for_timeout(W.LONG)
    else:
        raise Exception(f"点击'广告编辑'后始终未出现广告名称模块 (最后错误: {last_err})")
    wait_small(popup, W.NORMAL)

    ad_name = f"{cfg['name_prefix']}-{drama_name}-lzp-<日期>-<动态标号>"
    popup.locator("div.module-container#promotion-name").locator(
        "div.el-form-item:has(label:has-text('广告名称')) input.el-input__inner"
    ).first.fill(ad_name)
    logger.info(f"📝 广告名称: {ad_name}")

    drawer, wrap = get_visible_drawer(popup)
    scroll_wrap_to_bottom(popup, wrap, W)
    click_top_confirm(popup, logger=logger, W=W)
    wait_small(popup, W.LONGER)


def _get_material_pager(pane):
    pagers = pane.locator("div.el-pagination")
    if pagers.count() == 0:
        return None
    return pagers.last


def _get_material_total(pane):
    total_el = pane.locator("span.el-pagination__total").last
    if total_el.count() == 0:
        return None
    try:
        text = total_el.inner_text().strip()
    except Exception:
        return None
    m = re.search(r"共\s*(\d+)\s*条", text)
    return int(m.group(1)) if m else None


def _get_active_material_page(pane):
    pager = _get_material_pager(pane)
    if pager is None:
        return 1
    active = pager.locator("li.number.active").first
    if active.count() == 0:
        return 1
    try:
        return int(active.inner_text().strip())
    except Exception:
        return 1


def _has_next_material_page(pane):
    pager = _get_material_pager(pane)
    if pager is None:
        return False
    next_btn = pager.locator("button.btn-next").first
    if next_btn.count() == 0:
        return False
    disabled = next_btn.get_attribute("disabled")
    cls = next_btn.get_attribute("class") or ""
    return disabled is None and "disabled" not in cls


def _go_to_next_material_page(popup, pane, logger, W):
    pager = _get_material_pager(pane)
    if pager is None:
        return False
    next_btn = pager.locator("button.btn-next").first
    if next_btn.count() == 0:
        return False
    disabled = next_btn.get_attribute("disabled")
    cls = next_btn.get_attribute("class") or ""
    if disabled is not None or "disabled" in cls:
        return False
    old_page = _get_active_material_page(pane)
    try:
        next_btn.scroll_into_view_if_needed()
        popup.wait_for_timeout(W.TINY if W else 200)
        next_btn.click(force=True)
    except Exception as e:
        if logger:
            logger.warning(f"⚠️ 点击素材下一页失败: {e}")
        return False
    popup.wait_for_timeout(W.LOAD if W else 1500)
    material_wrapper = pane.locator("div.material-wrapper").first
    if material_wrapper.count() > 0:
        wait_loading_gone(popup, material_wrapper, timeout=10_000)
    for _ in range(10):
        new_page = _get_active_material_page(pane)
        if new_page != old_page:
            return True
        popup.wait_for_timeout(500)
    return False


def _go_to_material_page(popup, pane, page_no, logger, W):
    pager = _get_material_pager(pane)
    if pager is None:
        return _get_active_material_page(pane) == page_no
    for _ in range(20):
        current_page = _get_active_material_page(pane)
        if current_page == page_no:
            return True
        visible_target = pager.locator("li.number").filter(has_text=str(page_no)).first
        if visible_target.count() > 0:
            safe_click(popup, visible_target, desc=f"素材页码{page_no}", logger=logger, W=W)
            wait_small(popup, W.LOAD)
            wait_loading_gone(popup, pane)
            continue
        if current_page < page_no and _has_next_material_page(pane):
            _go_to_next_material_page(popup, pane, logger, W)
            continue
        prev_btn = pager.locator("button.btn-prev").first
        if current_page > page_no and prev_btn.count() > 0:
            disabled = prev_btn.get_attribute("disabled")
            cls = prev_btn.get_attribute("class") or ""
            if disabled is None and "disabled" not in cls:
                safe_click(popup, prev_btn, desc="素材上一页", logger=logger, W=W)
                wait_small(popup, W.LOAD)
                wait_loading_gone(popup, pane)
                continue
        break
    return _get_active_material_page(pane) == page_no


def _find_material_card_on_current_page(popup, material_dlg, pane, material_name, W):
    material_wrapper = pane.locator("div.material-wrapper").first
    seen_rounds = 0
    try:
        material_wrapper.evaluate("el => el.scrollTop = 0")
    except Exception:
        pass
    wait_small(popup, W.SHORT)
    for _ in range(80):
        names = material_wrapper.locator("div.material-name")
        count = names.count()
        for i in range(count):
            name_el = names.nth(i)
            try:
                current_name = name_el.inner_text().strip()
            except Exception:
                continue
            if current_name == material_name:
                return name_el.locator("xpath=ancestor::div[contains(@class,'material-item')]").first
        try:
            before = material_wrapper.evaluate("el => el.scrollTop")
            material_wrapper.evaluate("""el => {
                el.scrollTop = Math.min(el.scrollTop + el.clientHeight, el.scrollHeight);
            }""")
            after = material_wrapper.evaluate("el => el.scrollTop")
            if after == before:
                seen_rounds += 1
            else:
                seen_rounds = 0
        except Exception:
            seen_rounds += 1
        if seen_rounds >= 3:
            break
        wait_small(popup, W.MEDIUM)
    return None


def _open_material_dialog(popup, drama_name, cfg, logger, W, use_keyword):
    clear_btn = popup.locator("div.table-header:has(span:has-text('创意素材')) button:has-text('清空')").first
    safe_click(popup, clear_btn, desc="清空创意素材", logger=logger, W=W)
    wait_small(popup, W.NORMAL)
    edit3 = popup.locator("tfoot td button:has-text('编辑')").nth(2)
    safe_click(popup, edit3, desc="编辑按钮(素材)", logger=logger, W=W)
    wait_small(popup, W.EXTRA)
    batch_add_btn = popup.locator("button:has-text('批量添加素材')").first
    safe_click(popup, batch_add_btn, desc="批量添加素材", logger=logger, W=W)
    wait_small(popup, W.EXTRA)
    tab_media = popup.locator("#tab-account-material")
    safe_click(popup, tab_media, desc="素材账户tab", logger=logger, W=W)
    wait_small(popup, W.LONGER)
    material_dlg = popup.locator("div.el-dialog:visible").last
    pane = material_dlg.locator("#pane-account-material")
    if use_keyword:
        search_input = pane.locator("input[placeholder*='关键词查询']").first
        search_input.fill(drama_name)
        wait_small(popup, 600)
        logger.info(f"🔎 素材关键词已填入: {drama_name}")
    batch_icon = pane.locator("div.cl-input-area-trigger__icon[title='批量输入']").first
    batch_icon.click(force=True)
    wait_small(popup, W.LONG)
    id_input = popup.locator("input[placeholder*='请粘贴或输入账户ID']").last
    id_input.wait_for(state="visible", timeout=TIMEOUT)
    id_input.click(force=True)
    wait_small(popup, W.SHORT)
    id_input.fill(cfg["material_account_id"])
    wait_small(popup, W.MEDIUM)
    logger.info(f"🧾 固定素材账号已填入: {cfg['material_account_id']}")
    safe_click(popup, popup.locator("button.el-button--primary:visible").filter(has_text="搜索").last, desc="搜索素材", logger=logger, W=W)
    wait_small(popup, W.SEARCH)
    material_dlg = popup.locator("div.el-dialog:visible").last
    material_dlg.wait_for(state="visible", timeout=TIMEOUT)
    pane = material_dlg.locator("#pane-account-material")
    return material_dlg, pane


def _configure_material_filters(popup, pane, material_dlg, logger, W):
    material_wrapper = pane.locator("div.material-wrapper").first
    try:
        type_area = pane.locator("div.select-area:has(label:has-text('类型'))").first
        if type_area.count() > 0:
            type_input = type_area.locator("div.el-select input.el-input__inner").first
            if type_input.count() > 0:
                if safe_select_option(popup, type_input, "视频", desc="素材类型", logger=logger, W=W):
                    logger.info("🎥 已将素材类型切换为【视频】")
                    wait_loading_gone(popup, material_wrapper, timeout=15_000)
                    wait_small(popup, W.LONG)
        else:
            logger.warning('⚠️ 未找到"类型："选择区域，无法强制切为视频')
    except Exception as e:
        logger.warning(f"⚠️ 切换素材类型为视频时出错(忽略继续): {e}")
    try:
        size_input = pane.locator("span.el-pagination__sizes input.el-input__inner").first
        if size_input.count() > 0:
            size_input.click(force=True)
            wait_small(popup, W.NORMAL)
            option_100 = popup.locator("ul.el-select-dropdown__list:visible li.el-select-dropdown__item").filter(has_text="100").last
            if option_100.count() > 0:
                option_100.click(force=True)
                wait_loading_gone(popup, material_wrapper, timeout=30_000)
                wait_small(popup, W.LONG)
                logger.info("📄 已切换分页为100条/页")
            else:
                logger.warning("⚠️ 未找到100条/页选项")
    except Exception as e:
        logger.warning(f"⚠️ 分页切换失败(忽略继续): {e}")


def _collect_material_candidates(popup, material_dlg, pane, drama_name, logger, W):
    material_wrapper = pane.locator("div.material-wrapper").first
    material_wrapper.wait_for(state="visible", timeout=TIMEOUT)
    wait_loading_gone(popup, material_wrapper, timeout=30_000)
    load_start = time.time()
    load_deadline = load_start + 60
    stable_count = -1
    stable_since = 0.0
    STABLE_THRESHOLD = 3.0
    name_confirmed = False
    while time.time() < load_deadline:
        loading_mask = material_wrapper.locator(".el-loading-mask").first
        if loading_mask.count() > 0:
            try:
                style = loading_mask.get_attribute("style") or ""
            except Exception:
                style = ""
            if "display: none" not in style:
                stable_count = -1
                name_confirmed = False
                wait_small(popup, W.LONG)
                continue
        current = material_dlg.locator("div.material-name").count()
        if current > 0:
            if not name_confirmed:
                try:
                    first_name = material_dlg.locator("div.material-name").first.inner_text().strip()
                except Exception:
                    first_name = ""
                if first_name and drama_name in first_name:
                    name_confirmed = True
                    logger.info(f"🔎 首条素材名匹配剧名，确认搜索结果已加载: {first_name}")
            if name_confirmed:
                if current != stable_count:
                    stable_count = current
                    stable_since = time.time()
                    wait_small(popup, W.LONG)
                    continue
                if time.time() - stable_since >= STABLE_THRESHOLD:
                    logger.info(f"✅ 素材列表已加载: {drama_name} | {current} 个素材名 (耗时 {fmt_duration(time.time() - load_start)})")
                    break
            else:
                stable_count = -1
                wait_small(popup, W.LONG)
                continue
        wait_small(popup, W.EXTRA)
    else:
        if stable_count > 0:
            logger.info(f"⏳ 素材列表加载超时但已有 {stable_count} 个素材名，继续扫描: {drama_name}")
        elif name_confirmed:
            logger.info(f"⏳ 素材数量未完全稳定但已确认搜索结果，继续扫描: {drama_name}")
        else:
            logger.error(f"❌ 素材加载超时(60s): {drama_name}")
            return []
    wait_loading_gone(popup, material_wrapper)
    processed_names = set()
    available_candidates = []
    skipped_by_name = skipped_by_tag = skipped_by_future = 0
    scanned_pages = set()
    bad_keywords = ("低质", "巨量广告建议", "拒审")
    total_count = _get_material_total(pane)
    total_poll_rounds = 0
    logger.info(f"📄 素材分页总数: {total_count if total_count is not None else '未知'} 条 | 当前页: {_get_active_material_page(pane)}")
    consecutive_no_progress = 0
    while True:
        current_page = _get_active_material_page(pane)
        scanned_pages.add(current_page)
        no_new_rounds = 0
        page_start_count = len(processed_names)
        try:
            material_wrapper.evaluate("el => el.scrollTop = 0")
        except Exception:
            pass
        wait_small(popup, W.SHORT)
        for scroll_round in range(90):
            names = material_wrapper.locator("div.material-name")
            count = names.count()
            found_new_this_round = 0
            for i in range(count):
                name_el = names.nth(i)
                try:
                    material_name = name_el.inner_text().strip()
                except Exception:
                    continue
                if not material_name or material_name in processed_names:
                    continue
                processed_names.add(material_name)
                found_new_this_round += 1
                if not is_valid_material_name(drama_name, material_name):
                    skipped_by_name += 1
                    continue
                card = name_el.locator("xpath=ancestor::div[contains(@class,'material-item')]").first
                try:
                    tag_text = card.locator("div.tag-wrapper").inner_text().strip()
                except Exception:
                    tag_text = ""
                if any(keyword in tag_text for keyword in bad_keywords):
                    skipped_by_tag += 1
                    logger.info(f"⏭️ 跳过素材: {material_name} | 标签: {tag_text}")
                    continue
                mmdd = extract_mmdd(material_name)
                if mmdd and int(mmdd) > int(TODAY_STR):
                    skipped_by_future += 1
                    continue
                available_candidates.append({"name": material_name, "page": current_page, "date": mmdd})
            if found_new_this_round > 0:
                no_new_rounds = 0
            else:
                no_new_rounds += 1
            if no_new_rounds >= 6:
                break
            try:
                has_more_before_scroll = material_wrapper.evaluate("el => el.scrollTop + el.clientHeight < el.scrollHeight - 2")
                if total_count is not None and len(processed_names) < total_count and not has_more_before_scroll:
                    total_poll_rounds += 1
                    if total_poll_rounds <= 30:
                        wait_small(popup, W.LONG)
                        material_wrapper.evaluate("el => el.scrollTop = 0")
                        wait_small(popup, W.SHORT)
                        continue
                    else:
                        break
                material_wrapper.evaluate("""el => {
                    el.scrollTop = Math.min(el.scrollTop + el.clientHeight, el.scrollHeight);
                }""")
                if not has_more_before_scroll:
                    material_wrapper.evaluate("el => el.scrollTop = el.scrollHeight")
            except Exception as e:
                logger.warning(f"⚠️ 素材列表滚动失败(忽略继续): {e}")
            wait_small(popup, W.LOAD if no_new_rounds > 0 else W.LONGER)
        page_new_count = len(processed_names) - page_start_count
        logger.info(f"📄 第{current_page}页扫描完成，本页新增 {page_new_count} 个素材名，累计 {len(processed_names)}/{total_count if total_count is not None else '未知'}")
        if page_new_count == 0:
            consecutive_no_progress += 1
        else:
            consecutive_no_progress = 0
        if consecutive_no_progress >= 2:
            logger.info(f"📄 连续 {consecutive_no_progress} 页无新素材，停止翻页")
            break
        if total_count is not None and len(processed_names) >= total_count:
            logger.info(f"✅ 已扫描素材数 {len(processed_names)}/{total_count}，达到分页总数")
            break
        if not _has_next_material_page(pane):
            logger.info(f"📄 已到素材最后一页，扫描素材数 {len(processed_names)}/{total_count if total_count is not None else '未知'}")
            break
        logger.info(f"➡️ 当前已扫描素材 {len(processed_names)}/{total_count if total_count is not None else '未知'}，继续翻到下一页")
        total_poll_rounds = 0
        old_page = _get_active_material_page(pane)
        if not _go_to_next_material_page(popup, pane, logger, W):
            logger.warning("⚠️ 素材下一页切换失败，停止继续翻页")
            break
        new_page = _get_active_material_page(pane)
        if new_page == old_page:
            logger.warning("⚠️ 翻页后页码未变化，停止继续翻页")
            break
        logger.info(f"📄 已翻到第{new_page}页，等待内容加载...")
        wait_small(popup, W.NORMAL)
        wait_loading_gone(popup, material_wrapper, timeout=10_000)
        logger.info(f"📄 第{new_page}页加载完成，开始扫描")
    logger.info(f"🔎 候选素材收集完成: {drama_name} | 共 {len(available_candidates)} 个可用 | 扫描素材名 {len(processed_names)} 个/{total_count if total_count is not None else '未知'} | 扫描页数 {len(scanned_pages)} | 跳过: 剧名不符 {skipped_by_name}, 质量标签 {skipped_by_tag}, 未来日期 {skipped_by_future}")
    return available_candidates


def _select_and_submit_materials(popup, material_dlg, pane, available_candidates, drama_name, logger, W):
    def _cancel_and_return():
        cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
        if cancel_btn.count() > 0:
            cancel_btn.click(force=True)
            wait_small(popup, W.NORMAL)
    dated_candidates = [x for x in available_candidates if x["date"]]
    if not dated_candidates:
        logger.warning(f"⚠️ 没有找到带日期的可用素材: {drama_name}")
        _cancel_and_return()
        return 0
    date_groups = {}
    for item in dated_candidates:
        date_groups.setdefault(item["date"], []).append(item)
    valid_dates = sorted(date_groups.keys(), reverse=True)
    logger.info(f"📅 今天日期: {TODAY_STR} | 可用历史日期(近到远): {', '.join(valid_dates)}")
    selected_candidates = []
    for d in valid_dates:
        group = date_groups[d]
        remain = 30 - len(selected_candidates)
        if remain <= 0:
            break
        selected_candidates.extend(group[:remain])
        logger.info(f"📅 日期 {d} 可用 {len(group)} 个，本轮选中 {len(group[:remain])} 个，当前累计 {len(selected_candidates)} 个")
        if len(selected_candidates) >= 30:
            break
    if not selected_candidates:
        logger.warning(f"⚠️ 没有可提交素材: {drama_name}")
        _cancel_and_return()
        return 0
    logger.info(f"✅ 共找到带日期可用素材: {len(dated_candidates)} 个，实际准备选择: {len(selected_candidates)} 个")
    selected_candidates.sort(key=lambda x: (x.get("page", 1), x.get("date", ""), x.get("name", "")))
    picked_count = 0
    pick_start = time.time()
    current_pick_page = _get_active_material_page(pane)
    if current_pick_page != 1:
        if _go_to_material_page(popup, pane, 1, logger, W):
            current_pick_page = 1
    for item in selected_candidates:
        material_name = item["name"]
        target_page = item.get("page", 1)
        try:
            if current_pick_page != target_page:
                if not _go_to_material_page(popup, pane, target_page, logger, W):
                    logger.warning(f"⚠️ 跳转素材第{target_page}页失败: {material_name}")
                    continue
                current_pick_page = target_page
            card = _find_material_card_on_current_page(popup, material_dlg, pane, material_name, W)
            if card is None or card.count() == 0:
                logger.warning(f"⚠️ 当前页未重新定位到素材: {material_name}")
                continue
            safe_click(popup, card, desc=f"素材卡片:{material_name[:15]}", logger=logger, W=W)
            wait_small(popup, W.SHORT)
            picked_count += 1
            logger.info(f"✅ 已选择素材 ({picked_count}/{len(selected_candidates)}): {material_name}")
        except Exception as e:
            logger.warning(f"⚠️ 选择素材失败: {material_name} | {e}")
    logger.info(f"🎞️ 素材选择完成: 成功 {picked_count}/{len(selected_candidates)} 个，耗时 {fmt_duration(time.time() - pick_start)}")
    if picked_count == 0:
        logger.warning(f"⚠️ 没有成功选中任何素材: {drama_name}")
        _cancel_and_return()
        return 0
    submit_btn = material_dlg.locator("button.submit-button:visible").last
    if submit_btn.count() > 0:
        safe_click(popup, submit_btn, desc="素材提交按钮", logger=logger, W=W)
    else:
        safe_click(popup, material_dlg.locator("button:visible").filter(has_text="提交").last, desc="素材提交(备选)", logger=logger, W=W)
    click_optional_confirm(popup, desc="素材提交确认按钮", timeout=8_000, logger=logger, W=W)
    try:
        material_dlg.wait_for(state="hidden", timeout=20_000)
        logger.info("✅ 素材弹窗已关闭，已回到批量新建页面")
    except Exception:
        logger.warning("⚠️ 素材弹窗关闭等待超时，继续尝试后续提交")
    wait_idle(popup, mask_timeout=5_000)
    wait_small(popup, W.LOAD)
    return picked_count


def _pick_materials_by_keyword(popup, drama_name, cfg, logger, W):
    material_dlg, pane = _open_material_dialog(popup, drama_name, cfg, logger, W, use_keyword=True)
    _configure_material_filters(popup, pane, material_dlg, logger, W)
    available_candidates = _collect_material_candidates(popup, material_dlg, pane, drama_name, logger, W)
    if not available_candidates:
        logger.warning(f"⚠️ 未找到可用素材: {drama_name}")
        cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
        if cancel_btn.count() > 0:
            cancel_btn.click(force=True)
            wait_small(popup, W.NORMAL)
        return 0
    return _select_and_submit_materials(popup, material_dlg, pane, available_candidates, drama_name, logger, W)


def _pick_materials_by_ids(popup, drama_name, material_ids, cfg, logger, W):
    material_dlg, pane = _open_material_dialog(popup, drama_name, cfg, logger, W, use_keyword=False)
    try:
        search_select = pane.locator("div.cl-search-input div.el-select, div.cl-input-area div.el-select").first
        if search_select.count() > 0:
            select_input = search_select.locator("input.el-input__inner").first
            safe_click(popup, select_input, desc="搜索类型下拉", logger=logger, W=W)
            wait_small(popup, W.NORMAL)
            dropdown = popup.locator("ul.el-select-dropdown__list:visible").last
            dropdown.wait_for(state="visible", timeout=TIMEOUT)
            id_option = dropdown.locator("li.el-select-dropdown__item").filter(has_text="素材ID").first
            if id_option.count() > 0:
                safe_click(popup, id_option, desc="选择素材ID", logger=logger, W=W)
                wait_small(popup, W.MEDIUM)
    except Exception as e:
        logger.warning(f"⚠️ 切换搜索类型失败: {e}")
    batch_search_icon = pane.locator("div.cl-search-input__suffix-icon[title='批量搜索']").first
    if batch_search_icon.count() == 0:
        batch_search_icon = pane.locator("[title='批量搜索']").first
    safe_click(popup, batch_search_icon, desc="批量搜索图标", logger=logger, W=W)
    wait_small(popup, W.LONG)
    batch_textarea = popup.locator("textarea:visible").last
    batch_input_el = popup.locator("input[placeholder*='请粘贴'], input[placeholder*='请输入'], input[placeholder*='素材']").last
    if batch_textarea.count() > 0:
        target_el = batch_textarea
    elif batch_input_el.count() > 0:
        target_el = batch_input_el
    else:
        target_el = None
    if target_el:
        target_el.wait_for(state="visible", timeout=TIMEOUT)
        target_el.click(force=True)
        wait_small(popup, W.SHORT)
        for mid in material_ids:
            popup.keyboard.type(str(mid))
            popup.keyboard.press("Enter")
        wait_small(popup, W.MEDIUM)
        logger.info(f"✅ 已逐行输入 {len(material_ids)} 个素材ID")
    old_material_count = pane.locator("div.material-item").count()
    safe_click(popup, popup.locator("button.el-button--primary:visible").filter(has_text="搜索").last, desc="搜索素材ID", logger=logger, W=W)
    wait_small(popup, W.SEARCH)
    material_dlg = popup.locator("div.el-dialog:visible").last
    material_dlg.wait_for(state="visible", timeout=TIMEOUT)
    pane = material_dlg.locator("#pane-account-material")
    material_wrapper = pane.locator("div.material-wrapper").first
    material_wrapper.wait_for(state="visible", timeout=TIMEOUT)
    expected_count = len(material_ids)
    load_deadline = time.time() + 60
    refreshed = False
    stable_count = -1
    stable_since = 0
    while time.time() < load_deadline:
        loading_mask = material_wrapper.locator(".el-loading-mask").first
        if loading_mask.count() > 0:
            try:
                style = loading_mask.get_attribute("style") or ""
            except Exception:
                style = ""
            if "display: none" not in style:
                stable_count = -1
                wait_small(popup, W.LONG)
                continue
        current_count = pane.locator("div.material-item").count()
        if current_count == expected_count and current_count > 0:
            refreshed = True
            break
        if current_count != stable_count:
            stable_count = current_count
            stable_since = time.time()
            wait_small(popup, W.LONG)
            continue
        if current_count == stable_count and (time.time() - stable_since) > 3:
            if current_count != old_material_count and current_count > 0:
                refreshed = True
                break
        wait_small(popup, W.LONG)
    if not refreshed:
        logger.error("❌ 素材列表刷新超时(60s)")
        cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
        if cancel_btn.count() > 0:
            cancel_btn.click(force=True)
        return 0
    wait_loading_gone(popup, material_wrapper)
    wait_small(popup, W.NORMAL)
    cards = pane.locator("div.material-item")
    total = cards.count()
    picked = 0
    for i in range(total):
        try:
            safe_click(popup, cards.nth(i), desc=f"素材卡片{i+1}/{total}", logger=logger, W=W)
            wait_small(popup, W.TINY)
            picked += 1
        except Exception:
            pass
    if picked == 0:
        cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
        if cancel_btn.count() > 0:
            cancel_btn.click(force=True)
        return 0
    logger.info(f"✅ 已全选 {picked}/{total} 个素材")
    submit_btn = material_dlg.locator("button.submit-button:visible").last
    if submit_btn.count() > 0:
        safe_click(popup, submit_btn, desc="素材提交按钮", logger=logger, W=W)
    else:
        safe_click(popup, material_dlg.locator("button:visible").filter(has_text="提交").last, desc="素材提交(备选)", logger=logger, W=W)
    click_optional_confirm(popup, desc="素材提交确认按钮", timeout=8_000, logger=logger, W=W)
    try:
        material_dlg.wait_for(state="hidden", timeout=20_000)
        logger.info("✅ 素材弹窗已关闭，已回到批量新建页面")
    except Exception:
        logger.warning("⚠️ 素材弹窗关闭等待超时，继续尝试后续提交")
    wait_idle(popup, mask_timeout=5_000)
    wait_small(popup, W.LOAD)
    return picked


def step_pick_media_materials(popup, drama_name, material_ids, cfg, logger, W):
    if material_ids:
        logger.info(f"🎯 检测到配置内素材ID {len(material_ids)} 个，使用自定义素材ID逻辑")
        picked = _pick_materials_by_ids(popup, drama_name, material_ids, cfg, logger, W)
    else:
        logger.info("🎯 未检测到素材ID，使用普通按剧名筛素材逻辑")
        picked = _pick_materials_by_keyword(popup, drama_name, cfg, logger, W)
    if picked <= 0:
        raise Exception(f"❌ 未成功选择素材: {drama_name}")
    logger.info(f"✅ 素材已提交到创意素材区: {picked} 个")


def wait_return_to_main_after_material(popup, logger, W):
    logger.info("⏳ 等待返回 main 主界面")
    deadline = time.time() + 60
    selectors = [
        "main button:has-text('生成广告预览')",
        "#main button:has-text('生成广告预览')",
        ".main button:has-text('生成广告预览')",
        "div.main button:has-text('生成广告预览')",
        "button:has-text('生成广告预览')",
    ]
    while time.time() < deadline:
        wait_idle(popup, mask_timeout=2_000)
        for sel in selectors:
            btn = popup.locator(sel).first
            if _locator_count(btn) > 0:
                try:
                    btn.wait_for(state="visible", timeout=1_000)
                    logger.info(f"✅ 已回到 main 主界面，可生成广告预览：{sel}")
                    return btn
                except Exception:
                    pass
        wait_small(popup, W.NORMAL)
    raise Exception("素材提交后未回到 main 主界面，未看到“生成广告预览”按钮")


def step_submit_and_close(popup, page, logger, W):
    logger.info("➡️ 最终提交1/5：生成广告预览")
    preview_btn = wait_return_to_main_after_material(popup, logger, W)
    safe_click(popup, preview_btn, desc="生成广告预览", logger=logger, W=W)
    logger.info("⏳ 已点击生成广告预览，等待预览生成完成")
    try:
        popup.locator(".el-loading-mask").first.wait_for(state="hidden", timeout=60_000)
    except Exception:
        pass
    wait_idle(popup, mask_timeout=5_000)

    logger.info("➡️ 最终提交2/5：等待提交审核按钮")
    sub = popup.locator("button:has-text('全部提交审核')").first
    try:
        sub.wait_for(state="visible", timeout=120_000)
        logger.info("✅ 已出现“全部提交审核”按钮")
    except Exception:
        logger.warning("⚠️ 未找到“全部提交审核”，尝试“提交审核”")
        sub = popup.locator("button:has-text('提交审核')").first
        sub.wait_for(state="visible", timeout=30_000)

    logger.info("➡️ 最终提交3/5：点击提交审核")
    safe_click(popup, sub, desc="全部提交审核", logger=logger, W=W)
    wait_idle(popup, mask_timeout=8_000)

    logger.info("➡️ 最终提交4/5：确认提交弹窗")
    final_dlg = popup.locator("div.el-dialog__wrapper:visible, div.el-dialog:visible, div[role='dialog']:visible").last
    final_dlg.wait_for(state="visible", timeout=TIMEOUT)
    ok_btn = final_dlg.locator("button.el-button--primary:visible").filter(has_text="确 定").first
    if ok_btn.count() == 0:
        ok_btn = final_dlg.locator("button.el-button--primary:visible").filter(has_text="确定").first
    if ok_btn.count() == 0:
        ok_btn = final_dlg.locator("button:visible").filter(has_text="确定").first
    safe_click(popup, ok_btn, desc="最终确定按钮", logger=logger, W=W)
    wait_small(popup, 3000)

    logger.info("➡️ 最终提交5/5：转为后台提交并关闭页面")
    bg_btn = popup.locator("button:has-text('转为后台提交')").first
    try:
        bg_btn.wait_for(state="visible", timeout=60_000)
    except Exception:
        logger.warning("⚠️ 未看到“转为后台提交”，尝试直接关闭批量新建页面")
        bg_btn = None
    if bg_btn is not None:
        safe_click(popup, bg_btn, desc="转为后台提交", logger=logger, W=W)
    wait_small(popup, W.LOAD)
    popup.close()
    try:
        popup.wait_for_event("close", timeout=5_000)
    except Exception:
        pass
    wait_idle(page, mask_timeout=3_000, network=True, network_timeout=3_000)


# ═══════════════════════════════════════════════════════════════
#  主运行函数（参数化）
# ═══════════════════════════════════════════════════════════════
def run_build(profile_key: str, log_callback=None, stop_event=None):
    """
    执行搭建流程。
    profile_key: PROFILES 中的 key
    log_callback: 可选，接收 (str) 的回调，用于 GUI 日志显示
    stop_event: threading.Event，外部可设置以中止
    """
    app_cfg = load_config()
    cfg = build_runtime_profile_config(profile_key, app_cfg)
    cdp_endpoint = (app_cfg.get("common") or {}).get("cdp_endpoint") or "http://localhost:9222"

    W = WaitTimes(cfg["wait_scale"])
    logger = setup_logger(cfg["log_dir"])

    # 包装 logger 使其同时回调 GUI
    if log_callback:
        class _GUIHandler(logging.Handler):
            def emit(self, record):
                try: log_callback(self.format(record))
                except: pass
        gh = _GUIHandler()
        gh.setLevel(logging.INFO)
        gh.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S"))
        logger.addHandler(gh)

    logger.info(f"🚀 开始搭建: {profile_key}")
    logger.info(
        "⚙️ 本次运行变量："
        f"策略={cfg['strategy']} | "
        f"素材账号ID={cfg['material_account_id']} | "
        f"受众关键词={cfg['audience_keyword']} | "
        f"监控按钮={cfg['monitor_btn_text']} | "
        f"命名前缀={cfg['name_prefix']} | "
        f"等待倍率={cfg['wait_scale']}"
    )
    t0 = time.time()
    failed_dramas = []
    completed_dramas = []
    success_account_ids = set()

    groups = profile_groups_from_config(app_cfg, profile_key)
    if groups:
        logger.info(f"📦 数据来源：内置配置 config.json（{len(groups)} 组）")
    else:
        logger.info("📄 内置配置为空，回退读取 ids.txt")
        groups = read_data(cfg["ids_file"], logger)
    if not groups:
        logger.error("❌ 没有读取到任何数据（请打开「⚙ 设置」录入账号 ID / 链接 / 素材 ID）"); return

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.connect_over_cdp(cdp_endpoint)
            if not browser.contexts:
                raise RuntimeError("已连接浏览器，但没有可用的浏览器上下文")
            context = browser.contexts[0]
            page = select_build_page(context, logger)
            logger.info("✅ 已连接浏览器")

            for g_idx, (ids, dramas) in enumerate(groups, 1):
                check_stop(stop_event)
                logger.info(f"\n{'='*50}\n📦 第 {g_idx}/{len(groups)} 组 | 账号数: {len(ids)} | 剧数: {len(dramas)}\n{'='*50}")

                for d_idx, drama in enumerate(dramas, 1):
                    check_stop(stop_event)
                    drama_name = drama["name"]
                    logger.info(f"\n--- 第 {d_idx}/{len(dramas)} 部剧: {drama_name} ---")

                    popup = None
                    try:
                        check_stop(stop_event)
                        if page.is_closed():
                            page = select_build_page(context, logger)
                        try:
                            page.bring_to_front()
                        except Exception:
                            pass
                        logger.info(f"📍 当前操作页面：{_safe_page_title(page) or '无标题'} | {_safe_page_url(page) or '无URL'}")
                        batch_btns = page.locator("button:has-text('批量新建')")
                        batch_count = 0
                        try:
                            batch_count = batch_btns.count()
                        except Exception:
                            pass
                        logger.info(f"🔘 页面匹配到“批量新建”按钮：{batch_count} 个")
                        if batch_count == 0:
                            page = select_build_page(context, logger)
                            batch_btns = page.locator("button:has-text('批量新建')")
                            try:
                                batch_count = batch_btns.count()
                            except Exception:
                                batch_count = 0
                            logger.info(f"🔁 重新选择页面后匹配到“批量新建”按钮：{batch_count} 个")
                        if batch_count == 0:
                            raise RuntimeError("当前连接的页面没有找到“批量新建”按钮，请把浏览器切到媒体账户-巨量广告-广告管理页后重试")
                        batch_btn = batch_btns.first
                        batch_btn.wait_for(state="visible", timeout=TIMEOUT)
                        batch_btn.scroll_into_view_if_needed()
                        try:
                            expect(batch_btn).to_be_enabled(timeout=5_000)
                        except Exception:
                            pass
                        wait_idle(page, mask_timeout=3_000)
                        logger.info("🖱️ 准备点击“批量新建”")
                        with page.expect_popup() as popup_info:
                            batch_btn.click(force=True)
                        logger.info("✅ 已点击“批量新建”，等待批量新建页面打开")
                        popup = popup_info.value
                        popup.set_default_timeout(15_000)
                        try:
                            popup.bring_to_front()
                        except Exception:
                            pass
                        logger.info(f"🪟 批量新建页面已打开：{_safe_page_title(popup) or '无标题'} | {_safe_page_url(popup) or '无URL'}")
                        try:
                            popup.wait_for_load_state("networkidle", timeout=60_000)
                            logger.info("✅ 批量新建页面网络已空闲")
                        except PlaywrightTimeout:
                            logger.warning("⚠️ 批量新建页面等待 networkidle 超时，按参考流程继续尝试")

                        check_stop(stop_event)
                        logger.info("➡️ 步骤1/8：选择策略")
                        step_select_strategy(popup, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤2/8：选择媒体账户")
                        step_select_media_accounts(popup, ids, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤3/8：关联产品")
                        step_link_product(popup, drama_name, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤4/8：填写监测链接")
                        step_fill_monitor_links(popup, drama, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤5/8：选择定向包")
                        step_select_audience_package(popup, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤6/8：填写项目名称")
                        step_fill_project_name(popup, drama_name, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤7/8：填写广告名称")
                        step_fill_ad_name(popup, drama_name, cfg, logger, W)
                        check_stop(stop_event)
                        logger.info("➡️ 步骤8/8：选择素材并提交")
                        step_pick_media_materials(popup, drama_name, drama.get("material_ids", []), cfg, logger, W)
                        check_stop(stop_event)
                        step_submit_and_close(popup, page, logger, W)
                        completed_dramas.append(drama_name)
                        success_account_ids.update(ids)
                        logger.info(f"✅ {drama_name} 搭建完成")

                    except AccountsMissingError as e:
                        failed_dramas.append(drama_name)
                        logger.error(f"❌ 媒体账户缺失: {e}")
                        logger.error(f"⏭️ 第 {g_idx} 组检测到缺少账户，跳过本组剩余剧，直接进入下一组")
                        try:
                            if popup:
                                popup.close()
                        except Exception:
                            pass
                        break
                    except StopRequested:
                        logger.info("⏹ 用户中止")
                        try:
                            if popup:
                                popup.close()
                        except Exception:
                            pass
                        raise
                    except Exception as e:
                        failed_dramas.append(drama_name)
                        logger.error(f"❌ {drama_name} 搭建失败: {e}")
                        try:
                            if popup:
                                popup.close()
                        except Exception:
                            pass
                        continue
    except StopRequested:
        logger.info("⏹ 已停止")
        return

    elapsed = time.time() - t0
    logger.info(f"\n📊 搭建结果：成功 {len(completed_dramas)} 个，失败/异常/未完成 {len(failed_dramas)} 个")
    if failed_dramas:
        logger.error("\n❌ 未搭建完成剧名汇总：")
        for name in failed_dramas:
            logger.error(f"  {name}")
    else:
        logger.info("✅ 本次没有未搭建完成的剧")
    logger.info(f"\n🎉 全部完成! 总耗时: {fmt_duration(elapsed)}")

    if completed_dramas:
        record_build_success(len(success_account_ids), len(completed_dramas))
        logger.info(f"📝 基建记录已更新：账户 {len(success_account_ids)} 个，项目 {len(completed_dramas)} 个")
        logger.info(f"📝 本次账户ID: {', '.join(sorted(success_account_ids))}")


# ─────────────── 激励搭建：关联产品（空搜选第一个） ───────────────
def step_link_product_incentive(popup, cfg, logger, W):
    try:
        popup.wait_for_load_state("networkidle", timeout=5_000)
    except PlaywrightTimeout:
        pass
    wait_small(popup, W.LONG)
    edit_btn = popup.locator("tfoot td button:has-text('编辑')").nth(0)
    edit_btn.wait_for(state="visible", timeout=TIMEOUT)
    edit_btn.scroll_into_view_if_needed()
    wait_small(popup, W.MEDIUM)
    edit_btn.click()
    wait_small(popup, W.LOAD)

    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "link-product", W)
    safe_click(popup, popup.locator("#link-product button:has-text('选择产品')"), desc="选择产品按钮", logger=logger, W=W)
    prod_dlg = popup.locator("div.el-dialog__wrapper:visible").last
    prod_dlg.wait_for(state="visible", timeout=TIMEOUT)
    safe_click(popup, prod_dlg.locator("button.el-button--text.el-button--mini:has-text('清空')").first, desc="清空产品", logger=logger, W=W)
    wait_small(popup, W.LONG)
    safe_click(popup, prod_dlg.locator("button:has-text('查询')"), desc="查询全部产品", logger=logger, W=W)
    wait_small(popup, W.HEAVY)
    all_p = prod_dlg.locator("div.el-checkbox-group label.el-checkbox")
    total = all_p.count()
    if total > 0:
        safe_click(popup, all_p.first, desc="选择第一个产品", logger=logger, W=W)
        logger.info(f"🎯 激励模式：空搜选中第 1/{total} 个产品")
    else:
        logger.warning("⚠️ 未找到任何产品")
    click_top_confirm(popup, prod_dlg, desc="产品确认", wait_close=True, logger=logger, W=W)
    wait_small(popup, W.LONG)


# ─────────────── 激励搭建：监测链接（从meta取URL） ───────────────
def step_fill_monitor_links_incentive(popup, meta, cfg, logger, W):
    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "goal", W)

    safe_click(popup,
        popup.locator(f"#select-track-group button:has-text('{cfg['monitor_btn_text']}')"),
        desc=cfg['monitor_btn_text'], logger=logger, W=W)

    popup.locator("text=手动输入监测链接").first.wait_for(state="visible", timeout=TIMEOUT)
    monitor_drawer, monitor_wrap = get_visible_drawer(popup)

    def switch_on(label):
        item = monitor_drawer.locator(f"div.el-form-item:has(label.el-form-item__label:text-is('{label}'))").first
        if item.count() > 0:
            item.locator("label.el-radio-button:has-text('开启')").first.click(force=True)
            wait_small(popup, W.MEDIUM)

    def fill_link(label, value):
        item = monitor_drawer.locator(f"div.el-form-item:has(label.el-form-item__label:has-text('{label}'))").first
        input_box = item.locator("input.el-input__inner:visible").first
        input_box.wait_for(state="visible", timeout=TIMEOUT)
        input_box.click(force=True)
        wait_small(popup, W.TINY)
        input_box.fill(sanitize_link_text(value))
        wait_small(popup, W.SHORT)
        try:
            actual = input_box.input_value().strip()
        except Exception:
            actual = ""
        if not actual:
            raise Exception(f"{label}未填写成功")

    switch_on('展示链接'); switch_on('有效触点链接'); switch_on('视频有效播放链接')
    fill_link('请输入展示链接', meta.get("show_url", ""))
    logger.info("  展示链接☑️")
    fill_link('请输入有效触点链接', meta.get("click_url", ""))
    logger.info("  监测链接☑️")
    fill_link('请输入视频有效播放链接', meta.get("play_url", ""))
    logger.info("  播放链接☑️")
    logger.info("  ✅ 三个链接已填写完成")

    scroll_wrap_to_bottom(popup, monitor_wrap, W)
    click_top_confirm(popup, logger=logger, W=W)
    wait_small(popup, W.LONG)


# ─────────────── 激励搭建：项目名称（用组名） ───────────────
def step_fill_project_name_incentive(popup, group_name, cfg, logger, W):
    drawer, wrap = get_visible_drawer(popup)
    scroll_to_module(popup, wrap, "project-name", W)
    project_name = f"{cfg['name_prefix']}-lzp-<日期>-{group_name}"
    name_input = popup.locator("#project-name input.el-input__inner").first
    name_input.wait_for(state="visible", timeout=TIMEOUT)
    name_input.click(force=True); name_input.fill("")
    wait_small(popup, W.TINY)
    name_input.fill(project_name)
    wait_small(popup, W.SHORT)
    try: actual_value = name_input.input_value().strip()
    except: actual_value = ""
    if actual_value != project_name:
        name_input.click(force=True); name_input.fill("")
        wait_small(popup, W.TINY); name_input.fill(project_name)
        wait_small(popup, W.SHORT)
    logger.info(f"📝 项目名称: {project_name}")
    name_input.press("Tab"); wait_small(popup, W.SHORT)
    scroll_wrap_to_bottom(popup, wrap, W)
    drawer_confirm = drawer.locator("button.el-button--primary:not(.is-disabled):visible").filter(has_text=RE_CONFIRM).last
    if _locator_count(drawer_confirm) == 0:
        drawer_confirm = popup.locator("div.drawer-content:visible").last.locator("button:not(.is-disabled):visible").filter(has_text=RE_CONFIRM).last
    if _locator_count(drawer_confirm) > 0:
        safe_click(popup, drawer_confirm, desc="项目名称确定按钮", logger=logger, W=W)
    else:
        click_top_confirm(popup, desc="项目名称确定按钮", logger=logger, W=W)
    wait_idle(popup, mask_timeout=W.LOAD)


# ─────────────── 激励搭建：广告名称（用组名） ───────────────
def step_fill_ad_name_incentive(popup, group_name, cfg, logger, W):
    promo_block = popup.locator("div.module-container#promotion-name")
    edit_btn = popup.locator("tfoot td button:has-text('编辑')").nth(1)
    edit_btn.wait_for(state="visible", timeout=TIMEOUT)
    edit_btn.scroll_into_view_if_needed()
    last_err = None
    for _attempt in range(5):
        try: popup.keyboard.press("Escape")
        except: pass
        popup.wait_for_timeout(W.SHORT)
        try: edit_btn.click(force=True)
        except Exception as e:
            last_err = e; popup.wait_for_timeout(W.LONG); continue
        try: promo_block.first.wait_for(state="visible", timeout=8_000); break
        except Exception as e:
            last_err = e; popup.wait_for_timeout(W.LONG)
    else:
        raise Exception(f"点击'广告编辑'后始终未出现广告名称模块 (最后错误: {last_err})")
    wait_small(popup, W.NORMAL)
    ad_name = f"{cfg['name_prefix']}-lzp-<日期>-<动态标号>-{group_name}"
    popup.locator("div.module-container#promotion-name").locator(
        "div.el-form-item:has(label:has-text('广告名称')) input.el-input__inner"
    ).first.fill(ad_name)
    logger.info(f"📝 广告名称: {ad_name}")
    drawer, wrap = get_visible_drawer(popup)
    scroll_wrap_to_bottom(popup, wrap, W)
    click_top_confirm(popup, logger=logger, W=W)
    wait_small(popup, W.LONGER)


# ─────────────── 激励搭建：随机选取素材 ───────────────
def step_pick_materials_by_page(popup, pages_count, cfg, logger, W,
                                pick_min=30, pick_max=50, resume_position=None):
    logger.info("📦 进入素材编辑区域…")
    clear_btn = popup.locator("div.table-header:has(span:has-text('创意素材')) button:has-text('清空')").first
    if clear_btn.count() > 0:
        safe_click(popup, clear_btn, desc="清空创意素材", logger=logger, W=W)
        wait_small(popup, W.NORMAL)

    edit3 = popup.locator("tfoot td button:has-text('编辑')").nth(2)
    safe_click(popup, edit3, desc="编辑按钮(素材)", logger=logger, W=W)
    wait_small(popup, W.EXTRA)

    batch_add_btn = popup.locator("button:has-text('批量添加素材')").first
    safe_click(popup, batch_add_btn, desc="批量添加素材", logger=logger, W=W)
    wait_small(popup, W.EXTRA)

    tab_media = popup.locator("#tab-account-material")
    safe_click(popup, tab_media, desc="素材账户tab", logger=logger, W=W)
    wait_small(popup, W.LONGER)

    material_dlg = popup.locator("div.el-dialog:visible").last
    pane = material_dlg.locator("#pane-account-material")

    batch_icon = pane.locator("div.cl-input-area-trigger__icon[title='批量输入']").first
    batch_icon.click(force=True)
    wait_small(popup, W.LONG)
    id_input = popup.locator("input[placeholder*='请粘贴或输入账户ID']").last
    id_input.wait_for(state="visible", timeout=TIMEOUT)
    id_input.click(force=True)
    wait_small(popup, W.SHORT)
    id_input.fill(cfg["material_account_id"])
    wait_small(popup, W.MEDIUM)
    logger.info(f"🧾 固定素材账号已填入: {cfg['material_account_id']}")

    safe_click(popup, popup.locator("button.el-button--primary:visible").filter(has_text="搜索").last, desc="搜索素材", logger=logger, W=W)
    wait_small(popup, W.SEARCH)

    material_dlg = popup.locator("div.el-dialog:visible").last
    material_dlg.wait_for(state="visible", timeout=TIMEOUT)
    pane = material_dlg.locator("#pane-account-material")
    logger.info("📦 素材选择弹窗已打开")

    _configure_material_filters(popup, pane, material_dlg, logger, W)

    material_wrapper = pane.locator("div.material-wrapper").first
    material_wrapper.wait_for(state="visible", timeout=TIMEOUT)
    wait_loading_gone(popup, material_wrapper, timeout=30_000)

    load_start = time.time()
    load_deadline = load_start + 60
    loaded = False
    stable_count = -1
    stable_since = 0.0
    STABLE_THRESHOLD = 3.0
    while time.time() < load_deadline:
        loading_mask = material_wrapper.locator(".el-loading-mask").first
        if loading_mask.count() > 0:
            try:
                style = loading_mask.get_attribute("style") or ""
            except Exception:
                style = ""
            if "display: none" not in style:
                stable_count = -1
                wait_small(popup, W.LONG)
                continue
        current = material_wrapper.locator("div.material-name").count()
        if current > 0:
            if current != stable_count:
                stable_count = current
                stable_since = time.time()
                wait_small(popup, W.LONG)
                continue
            if time.time() - stable_since >= STABLE_THRESHOLD:
                logger.info(f"✅ 素材列表已加载 | {current} 个素材名 (耗时 {fmt_duration(time.time() - load_start)})")
                loaded = True
                break
        wait_small(popup, W.EXTRA)
    if not loaded:
        if stable_count > 0:
            logger.info(f"⏳ 素材列表加载超时但已有 {stable_count} 个素材名，继续选取")
            loaded = True
        else:
            logger.error("❌ 素材加载超时(60s)")
            cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
            if cancel_btn.count() > 0:
                cancel_btn.click(force=True)
            return resume_position
    wait_loading_gone(popup, material_wrapper)

    used_names = get_used_material_names()
    if used_names:
        logger.info(f"📋 已排除历史素材 {len(used_names)} 条")

    pick_count = random.randint(pick_min, pick_max)
    logger.info(f"🎯 本次目标选取: {pick_count} 条素材")

    start_page = (resume_position or {}).get("page", 0)
    start_offset = (resume_position or {}).get("offset", 0)

    if start_page > 0:
        _go_to_material_page(popup, pane, start_page + 1, logger, W)
        wait_loading_gone(popup, material_wrapper, timeout=10_000)
        logger.info(f"📄 从上次结束位置继续: 第 {start_page + 1} 页, 偏移 {start_offset}")

    picked_count = 0
    chosen_names = []
    current_page = start_page
    global_offset = start_offset

    while picked_count < pick_count:
        try:
            material_wrapper.evaluate("el => el.scrollTop = 0")
        except Exception:
            pass
        wait_small(popup, W.SHORT)

        seen_names_on_page = set()
        ordered_items = []
        no_new_rounds = 0

        for _ in range(90):
            names = material_wrapper.locator("div.material-name")
            count = names.count()
            found_new = 0
            for i in range(count):
                try:
                    name_text = names.nth(i).inner_text(timeout=3_000).strip()
                except Exception:
                    continue
                if not name_text or name_text in seen_names_on_page:
                    continue
                seen_names_on_page.add(name_text)
                found_new += 1
                ordered_items.append(name_text)
            if found_new > 0:
                no_new_rounds = 0
            else:
                no_new_rounds += 1
            if no_new_rounds >= 4:
                break
            try:
                at_bottom = material_wrapper.evaluate(
                    "el => el.scrollTop + el.clientHeight >= el.scrollHeight - 2")
                if at_bottom:
                    break
                material_wrapper.evaluate(
                    "el => { el.scrollTop = Math.min(el.scrollTop + el.clientHeight, el.scrollHeight); }")
            except Exception:
                break
            wait_small(popup, W.MEDIUM)

        logger.info(f"📄 第 {current_page + 1} 页识别到 {len(ordered_items)} 个素材")

        items_to_process = ordered_items[global_offset:] if current_page == start_page else ordered_items
        skip_on_page = 0
        pick_on_page = 0

        for item_name in items_to_process:
            if picked_count >= pick_count:
                break
            if item_name in used_names:
                skip_on_page += 1
                global_offset += 1
                continue
            card = _find_material_card_on_current_page(popup, material_dlg, pane, item_name, W)
            if card and card.count() > 0:
                try:
                    card.click(force=True)
                    wait_small(popup, W.TINY)
                    picked_count += 1
                    pick_on_page += 1
                    chosen_names.append(item_name)
                except Exception:
                    pass
            global_offset += 1

        if skip_on_page > 0:
            logger.info(f"⏭️ 第 {current_page + 1} 页跳过已用素材 {skip_on_page} 条")
        if pick_on_page > 0:
            logger.info(f"✅ 第 {current_page + 1} 页选取 {pick_on_page} 条 (累计 {picked_count}/{pick_count})")

        if picked_count >= pick_count:
            break

        if not _has_next_material_page(pane):
            logger.info("📄 已到最后一页，无更多素材")
            break
        if not _go_to_next_material_page(popup, pane, logger, W):
            break
        wait_loading_gone(popup, material_wrapper, timeout=10_000)
        current_page += 1
        global_offset = 0

    new_resume = {"page": current_page, "offset": global_offset}
    logger.info(f"📦 素材选取完成: 已选 {picked_count}/{pick_count} 条 | 结束位置: 第{current_page+1}页 偏移{global_offset}")

    if picked_count == 0:
        logger.warning("⚠️ 没有选到任何素材")
        cancel_btn = material_dlg.locator("button:visible").filter(has_text="取消").last
        if cancel_btn.count() > 0:
            cancel_btn.click(force=True)
            wait_small(popup, W.NORMAL)
        return new_resume

    submit_btn = material_dlg.locator("button.submit-button:visible").last
    if submit_btn.count() > 0:
        safe_click(popup, submit_btn, desc="素材提交按钮", logger=logger, W=W)
    else:
        safe_click(popup, material_dlg.locator("button:visible").filter(has_text="提交").last, desc="素材提交(备选)", logger=logger, W=W)
    click_optional_confirm(popup, desc="素材提交确认按钮", timeout=8_000, logger=logger, W=W)
    try:
        material_dlg.wait_for(state="hidden", timeout=20_000)
        logger.info("✅ 素材弹窗已关闭，已回到批量新建页面")
    except Exception:
        logger.warning("⚠️ 素材弹窗关闭等待超时，继续尝试后续提交")
    wait_idle(popup, mask_timeout=5_000)
    wait_small(popup, W.LOAD)

    add_material_history(chosen_names)
    logger.info(f"✅ 素材选择完成，已记录 {len(chosen_names)} 条素材到历史")
    return new_resume


# ─────────────── 激励搭建主流程 ───────────────
def run_build_incentive(profile_key: str, log_callback=None, stop_event=None):
    app_cfg = load_config()
    cfg = build_runtime_profile_config(profile_key, app_cfg)
    cdp_endpoint = (app_cfg.get("common") or {}).get("cdp_endpoint") or "http://localhost:9222"
    pages_per_round = cfg.get("pages_per_round", 3)

    W = WaitTimes(cfg["wait_scale"])
    logger = setup_logger(cfg["log_dir"])

    if log_callback:
        class _GUIHandler(logging.Handler):
            def emit(self, record):
                try: log_callback(self.format(record))
                except: pass
        gh = _GUIHandler()
        gh.setLevel(logging.INFO)
        gh.setFormatter(logging.Formatter("[%(asctime)s] %(message)s", datefmt="%H:%M:%S"))
        logger.addHandler(gh)

    logger.info(f"🚀 开始激励搭建: {profile_key}")
    t0 = time.time()
    completed_groups = []
    failed_groups = []
    success_account_ids = set()

    groups = profile_groups_from_config(app_cfg, profile_key)
    if not groups:
        logger.error("❌ 没有读取到任何数据"); return

    logger.info(f"📦 共 {len(groups)} 组")

    resume_position = None

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.connect_over_cdp(cdp_endpoint)
            if not browser.contexts:
                raise RuntimeError("已连接浏览器，但没有可用的浏览器上下文")
            context = browser.contexts[0]
            page = select_build_page(context, logger)
            logger.info("✅ 已连接浏览器")

            for g_idx, group_data in enumerate(groups, 1):
                check_stop(stop_event)
                ids = group_data[0]
                meta = group_data[2] if len(group_data) > 2 else {}
                group_name = meta.get("group_name", f"组{g_idx}")
                logger.info(f"\n{'='*50}\n📦 第 {g_idx}/{len(groups)} 组: {group_name} | 账号数: {len(ids)}\n{'='*50}")

                popup = None
                try:
                    check_stop(stop_event)
                    if page.is_closed():
                        page = select_build_page(context, logger)
                    try: page.bring_to_front()
                    except: pass

                    batch_btns = page.locator("button:has-text('批量新建')")
                    batch_count = 0
                    try: batch_count = batch_btns.count()
                    except: pass
                    if batch_count == 0:
                        page = select_build_page(context, logger)
                        batch_btns = page.locator("button:has-text('批量新建')")
                        try: batch_count = batch_btns.count()
                        except: batch_count = 0
                    if batch_count == 0:
                        raise RuntimeError("当前页面没有找到【批量新建】按钮")
                    batch_btn = batch_btns.first
                    batch_btn.wait_for(state="visible", timeout=TIMEOUT)
                    batch_btn.scroll_into_view_if_needed()
                    try: expect(batch_btn).to_be_enabled(timeout=5_000)
                    except: pass
                    wait_idle(page, mask_timeout=3_000)

                    with page.expect_popup() as popup_info:
                        batch_btn.click(force=True)
                    popup = popup_info.value
                    popup.set_default_timeout(15_000)
                    try: popup.bring_to_front()
                    except: pass
                    try: popup.wait_for_load_state("networkidle", timeout=60_000)
                    except PlaywrightTimeout: pass

                    check_stop(stop_event)
                    logger.info("➡️ 步骤1/8：选择策略")
                    step_select_strategy(popup, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤2/8：选择媒体账户")
                    step_select_media_accounts(popup, ids, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤3/8：关联产品（激励-空搜）")
                    step_link_product_incentive(popup, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤4/8：填写监测链接")
                    step_fill_monitor_links_incentive(popup, meta, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤5/8：选择定向包")
                    step_select_audience_package(popup, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤6/8：填写项目名称")
                    step_fill_project_name_incentive(popup, group_name, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤7/8：填写广告名称")
                    step_fill_ad_name_incentive(popup, group_name, cfg, logger, W)
                    check_stop(stop_event)
                    logger.info("➡️ 步骤8/8：顺序选取素材（30-50条）")
                    resume_position = step_pick_materials_by_page(popup, pages_per_round, cfg, logger, W, resume_position=resume_position)
                    check_stop(stop_event)
                    step_submit_and_close(popup, page, logger, W)
                    completed_groups.append(group_name)
                    success_account_ids.update(ids)
                    logger.info(f"✅ {group_name} 搭建完成")

                except AccountsMissingError as e:
                    failed_groups.append(group_name)
                    logger.error(f"❌ 媒体账户缺失: {e}")
                    try:
                        if popup: popup.close()
                    except: pass
                    continue
                except StopRequested:
                    logger.info("⏹ 用户中止")
                    try:
                        if popup: popup.close()
                    except: pass
                    raise
                except Exception as e:
                    failed_groups.append(group_name)
                    logger.error(f"❌ {group_name} 搭建失败: {e}")
                    try:
                        if popup: popup.close()
                    except: pass
                    continue
    except StopRequested:
        logger.info("⏹ 已停止"); return

    elapsed = time.time() - t0
    logger.info(f"\n📊 搭建结果：成功 {len(completed_groups)} 组，失败 {len(failed_groups)} 组")
    if failed_groups:
        logger.error("\n❌ 未搭建完成组汇总：")
        for name in failed_groups:
            logger.error(f"  {name}")
    logger.info(f"\n🎉 全部完成! 总耗时: {fmt_duration(elapsed)}")

    if completed_groups:
        record_build_success(len(success_account_ids), len(completed_groups))
        logger.info(f"📝 基建记录已更新：账户 {len(success_account_ids)} 个，项目 {len(completed_groups)} 个")
        logger.info(f"📝 本次账户ID: {', '.join(sorted(success_account_ids))}")


# ─────────────── 激励推广链生成（sync） ───────────────
INCENTIVE_PROMO_DRAWER_SEL = ".arco-drawer.promotion_form_wrapper"
INCENTIVE_PROMO_NAME_INPUT = "#promotion_name_input"
INCENTIVE_PROMO_CONFIRM_BTN = ".arco-drawer:visible button:has-text('确定')"
INCENTIVE_PROMO_TIMEOUT = 10_000
INCENTIVE_PROMO_DRAWER_TIMEOUT = 15_000


def _incentive_promo_run_once(page, index: int, date_str: str, suffix: str, log_func) -> bool:
    promotion_name = f"{date_str}-组{index}-{suffix}"
    log_func(f"\n[{index}] 开始创建推广链：{promotion_name}\n")
    try:
        activity_item = page.locator(
            "h3:has-text('短剧激励活动')"
        ).locator("xpath=ancestor::div[contains(@class, 'AmDBjA8pHU8hRXib')]")
        activity_item.wait_for(state="visible", timeout=INCENTIVE_PROMO_TIMEOUT)

        get_link_btn = activity_item.locator("text=获取推广链接")
        get_link_btn.wait_for(state="visible", timeout=INCENTIVE_PROMO_TIMEOUT)
        get_link_btn.click()
        log_func(f"[{index}] 已点击「获取推广链接」\n")

        drawer = page.locator(INCENTIVE_PROMO_DRAWER_SEL)
        drawer.wait_for(state="visible", timeout=INCENTIVE_PROMO_DRAWER_TIMEOUT)

        name_input = page.locator(INCENTIVE_PROMO_NAME_INPUT)
        name_input.wait_for(state="visible", timeout=INCENTIVE_PROMO_TIMEOUT)
        name_input.click(click_count=3)
        name_input.fill("")
        name_input.fill(promotion_name)
        log_func(f"[{index}] 名称已填入：{promotion_name}\n")

        confirm_btn = page.locator(INCENTIVE_PROMO_CONFIRM_BTN)
        confirm_btn.wait_for(state="visible", timeout=INCENTIVE_PROMO_TIMEOUT)
        confirm_btn.click()

        drawer.wait_for(state="hidden", timeout=INCENTIVE_PROMO_DRAWER_TIMEOUT)
        log_func(f"[{index}] ✅ 成功：{promotion_name}\n")

        page.wait_for_timeout(500)
        page.mouse.click(10, 10)
        page.wait_for_timeout(500)
        return True
    except Exception as e:
        log_func(f"[{index}] ❌ 失败：{e}\n")
        try:
            close_icon = page.locator(".arco-drawer:visible .arco-drawer-close-icon")
            if close_icon.count() > 0:
                close_icon.click()
                page.wait_for_timeout(500)
        except Exception:
            pass
        return False


def run_incentive_promo_chain(count: int, suffix: str, log_func, stop_event: threading.Event):
    from playwright.sync_api import sync_playwright
    cdp_url = "http://127.0.0.1:9222"
    date_str = datetime.now().strftime("%Y-%m-%d")
    log_func(f"日期：{date_str}，计划执行 {count} 次，后缀：{suffix}\n")
    log_func("=" * 50 + "\n")

    success_count = 0
    fail_count = 0
    failed_names = []

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(cdp_url)
        ctx = browser.contexts[0]
        page = None
        for pg in ctx.pages:
            if "changdupingtai" in pg.url or "hongguo" in pg.url or "promote" in pg.url or "buyin" in pg.url:
                page = pg
                break
        if not page:
            page = ctx.pages[0]
        page.bring_to_front()
        page.set_default_timeout(INCENTIVE_PROMO_TIMEOUT)

        for i in range(1, count + 1):
            if stop_event.is_set():
                log_func("\n已停止\n")
                break
            if _incentive_promo_run_once(page, i, date_str, suffix, log_func):
                success_count += 1
            else:
                fail_count += 1
                failed_names.append(f"{date_str}-组{i}-{suffix}")

    log_func(f"\n{'='*40}\n")
    log_func(f"🎉 完成！成功：{success_count}，失败：{fail_count}\n")
    if failed_names:
        log_func("失败列表：\n")
        for name in failed_names:
            log_func(f"  {name}\n")


# ─────────────── 激励素材推送（sync） ───────────────
INCENTIVE_PUSH_TIMEOUT = 10_000
INCENTIVE_PUSH_WAIT_AFTER = 120_000
INCENTIVE_PUSH_BETWEEN = 2_000
INCENTIVE_PUSH_DIALOG_SEL = ".arco-drawer, .arco-modal"
INCENTIVE_PUSH_OPTION_SEL = ".arco-select-option"
INCENTIVE_PUSH_NEXT_BTN = ".arco-pagination-item-next"


def _incentive_push_read_pages(page):
    items = page.locator(
        ".arco-pagination-item"
        ":not(.arco-pagination-item-prev)"
        ":not(.arco-pagination-item-next)"
        ":not(.arco-pagination-item-jumper)"
    )
    count = items.count()
    if count == 0:
        return 1
    for idx in range(count - 1, -1, -1):
        text = items.nth(idx).inner_text().strip()
        if text.isdigit():
            return int(text)
    return 1


def run_incentive_push(account_id: str, log_func, stop_event: threading.Event):
    from playwright.sync_api import sync_playwright
    cdp_url = "http://127.0.0.1:9222"

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(cdp_url)
        ctx = browser.contexts[0]
        page = next((pg for pg in ctx.pages if "material" in pg.url), ctx.pages[0])
        page.bring_to_front()
        page.set_default_timeout(INCENTIVE_PUSH_TIMEOUT)

        total_pages = _incentive_push_read_pages(page)
        log_func(f"📄 共 {total_pages} 页\n")
        log_func(f"🚀 将从第 1 页逐页推送到第 {total_pages} 页\n\n")

        success_count = 0
        fail_count = 0

        for i in range(total_pages):
            if stop_event.is_set():
                log_func("\n已停止\n")
                break
            page_num = i + 1
            log_func(f"\n🔄 正在推送第 {page_num}/{total_pages} 页...\n")
            try:
                checkbox = page.locator("table thead .arco-checkbox-mask")
                checkbox.wait_for(state="visible")
                checkbox.click()

                batch_btn = page.get_by_role("button", name="批量操作")
                batch_btn.wait_for(state="visible")
                batch_btn.click()

                push_btn = page.get_by_text("批量推送", exact=True)
                push_btn.wait_for(state="visible")
                push_btn.click()

                dialog = page.locator(INCENTIVE_PUSH_DIALOG_SEL).filter(has_text="批量素材推送")
                dialog.wait_for(state="visible")

                dialog.locator(".arco-select", has=page.get_by_placeholder("请选择媒体渠道")).click()
                page.locator(INCENTIVE_PUSH_OPTION_SEL, has_text="巨量引擎").wait_for(state="visible")
                page.locator(INCENTIVE_PUSH_OPTION_SEL, has_text="巨量引擎").click()

                dialog.locator(".arco-select", has=page.get_by_placeholder("请选择投放产品")).click()
                page.locator(INCENTIVE_PUSH_OPTION_SEL, has_text="红果免费短剧").wait_for(state="visible")
                page.locator(INCENTIVE_PUSH_OPTION_SEL, has_text="红果免费短剧").click()

                page.keyboard.press("Escape")

                account_input = dialog.locator("#ad_account_ids_input")
                account_input.wait_for(state="visible")
                account_input.fill(account_id)

                confirm_btn = dialog.get_by_role("button", name="确定")
                confirm_btn.wait_for(state="visible")
                confirm_btn.click()

                dialog.wait_for(state="hidden", timeout=INCENTIVE_PUSH_WAIT_AFTER)
                log_func(f"  ✅ 第 {page_num} 页推送完成\n")
                success_count += 1
                page.wait_for_timeout(INCENTIVE_PUSH_BETWEEN)

            except Exception as e:
                log_func(f"  ❌ 第 {page_num} 页推送失败：{e}\n")
                fail_count += 1
                page.keyboard.press("Escape")
                page.wait_for_timeout(INCENTIVE_PUSH_BETWEEN)
                try:
                    page.locator(INCENTIVE_PUSH_DIALOG_SEL).filter(has_text="批量素材推送").wait_for(
                        state="hidden", timeout=INCENTIVE_PUSH_TIMEOUT)
                except Exception:
                    pass

            if page_num < total_pages:
                try:
                    next_btn = page.locator(INCENTIVE_PUSH_NEXT_BTN)
                    next_btn.wait_for(state="visible")
                    next_btn.click()
                    page.wait_for_timeout(INCENTIVE_PUSH_BETWEEN)
                except Exception as e:
                    log_func(f"  ⚠️ 翻页失败：{e}，停止执行\n")
                    break

    log_func(f"\n{'='*40}\n")
    log_func(f"🎉 全部完成！成功：{success_count} 页，失败：{fail_count} 页\n")


# ═══════════════════════════════════════════════════════════════
#  CustomTkinter 现代 GUI
# ═══════════════════════════════════════════════════════════════
import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk

# 主题
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# 系统字体（Windows 标准中英文字体）
F_UI       = "Microsoft YaHei UI"
F_MONO     = "JetBrains Mono"  # 优先使用 JetBrains Mono，回退 Consolas
F_MONO_FB  = "Consolas"

# ── 设计 Token（参考 Linear / Vercel / Notion 浅色规范）──
# 中性色阶（Slate）
C_BG        = "#f7f8fa"   # 应用背景：极浅冷灰
C_CARD      = "#ffffff"   # 卡片：纯白
C_SURFACE   = "#f1f3f7"   # 二级面：浅灰（嵌套区域）
C_BORDER    = "#e4e7ec"   # 主边框：浅灰
C_BORDER_S  = "#eef0f3"   # 次级分割线：更浅
C_HOVER     = "#f4f5f7"   # hover 背景

# 文本
C_TEXT      = "#0f172a"   # 主文本：近黑
C_TEXT_2    = "#475569"   # 次文本：中灰蓝
C_DIM       = "#94a3b8"   # 弱文本：浅灰

# 品牌 / 强调
C_ACCENT    = "#ef4444"   # 红果红（主品牌）
C_ACCENT_H  = "#dc2626"   # 红 hover
C_PRIMARY   = "#4f46e5"   # 主操作色（靛蓝）
C_PRIMARY_H = "#4338ca"

# 语义色
C_GREEN     = "#10b981"   # 成功
C_GREEN_H   = "#059669"
C_RED       = "#ef4444"   # 错误
C_RED_H     = "#dc2626"
C_ORANGE    = "#f59e0b"   # 警告
C_BLUE      = "#3b82f6"   # 信息

# 日志
C_LOG_BG    = "#0f172a"   # 日志深色面板（高对比终端风）
C_LOG_FG    = "#e2e8f0"   # 日志默认文本：浅灰白
C_LOG_DIM   = "#64748b"   # 日志时间戳/分隔
C_LOG_SEL   = "#1e293b"   # 选中背景

# 圆角 / 间距 token
R_LG        = 14
R_MD        = 10
R_SM        = 8

# 兼容旧引用
C_ACCENT2   = C_PRIMARY


# ════════════════════════════════════════════════════════════════════
#  🔗 剧名链接整理工具（内嵌窗口）
# ════════════════════════════════════════════════════════════════════

PROMOTION_CHAIN_CDP = "http://127.0.0.1:9222"
PROMOTION_CHAIN_TIMEOUT = 20_000
PROMOTION_CHAIN_NAV_TIMEOUT = 15_000
PROMOTION_CHAIN_ELEMENT_TIMEOUT = 8_000
PROMOTION_CHAIN_SEARCH_DELAY = 800
PROMOTION_CHAIN_CLICK_DELAY = 200
PROMOTION_CHAIN_LIST_URL_PATTERN = re.compile(r"/short-play/list(?!/detail)")
PROMOTION_CHAIN_DETAIL_URL_PATTERN = re.compile(r"/short-play/list/detail")
PROMOTION_CHAIN_LIST_FRAG = "/short-play/list"
PROMOTION_CHAIN_DETAIL_FRAG = "/short-play/list/detail"
PROMOTION_CHAIN_MENU_SEL = "a[href='/sale/short-play/list']"
PROMOTION_CHAIN_QUERY_SEL = "#query_input"
PROMOTION_CHAIN_SEARCH_BTN = "button:has-text('搜索')"
PROMOTION_CHAIN_ROW_SEL = "tr.arco-table-tr.e2e-promotion-table-row:visible"
PROMOTION_CHAIN_BOOK_NAME_SEL = ".book_name_content"
PROMOTION_CHAIN_VIEW_DETAIL_SEL = "text=查看详情"
PROMOTION_CHAIN_GET_LINK_BTN = "button:has-text('获取短剧推广链')"
PROMOTION_CHAIN_PROMO_INPUT_SEL = "#promotion_name_input"
PROMOTION_CHAIN_IOS_RADIO_SEL = ".arco-radio"
PROMOTION_CHAIN_CONFIRM_SEL = (
    ".arco-modal:visible button:has-text('确定'), "
    ".arco-drawer:visible button:has-text('确定')"
)
PROMOTION_CHAIN_CONFIRM_FALLBACK = "button:has-text('确定')"
_PC_CLEAN_NAME_RE = re.compile(r"[^\u4e00-\u9fa5A-Za-z0-9]")
_PC_CHINESE_ONLY_RE = re.compile(r"[^\u4e00-\u9fa5]")
PROMOTION_CHAIN_TASKS = [
    ("安卓每留", "每留", False),
    ("安卓七留", "七留", False),
    ("iOS每留", "每留", True),
    ("iOS七留", "七留", True),
]
PROMOTION_SPLIT_DEFAULT_DIR = Path(r"C:\Users\Administrator\Downloads")
PROMOTION_SPLIT_KEEP_COLS = [1, 7, 8, 9, 10, 11]
PROMOTION_SPLIT_GROUP_ORDER = ["IOS-每留", "IOS-七留", "Android-每留", "Android-七留"]
PROMOTION_SPLIT_DISPLAY_ORDER = [("Android-每留", "安卓每留"), ("Android-七留", "安卓七留"), ("IOS-每留", "iOS每留"), ("IOS-七留", "iOS七留")]
PROMOTION_SPLIT_GAP_ROWS = 3

INCENTIVE_SPLIT_DEFAULT_DIR = Path(r"C:\Users\Administrator\Downloads")
INCENTIVE_SPLIT_KEEP_COLS = [1, 7, 8, 9, 10, 11]
INCENTIVE_SPLIT_GROUP_ORDER = ["激励-每留", "激励-七留"]
INCENTIVE_SPLIT_DISPLAY_ORDER = [("激励-每留", "激励每留"), ("激励-七留", "激励七留")]
INCENTIVE_SPLIT_GAP_ROWS = 3
PROMOTION_TASK_OPTIONS = [
    ("安卓每留", 0),
    ("安卓七留", 1),
    ("iOS每留", 2),
    ("iOS七留", 3),
]

def _normalize_title(title):
    """标题归一化：小写、去空格、去标点，仅保留中文+字母+数字。"""
    if not title:
        return ""
    import unicodedata
    t = title.lower().strip()
    t = re.sub(r'\s+', '', t)
    t = re.sub(r'[^\u4e00-\u9fa5a-z0-9]', '', t)
    return t


def _fuzzy_find(name, mapping):
    """模糊匹配：精确 → 包含关系。"""
    key = _normalize_title(name)
    if not key:
        return None
    if key in mapping:
        return mapping[key]
    for k in mapping:
        if key in k or k in key:
            return mapping[k]
    return None


def _parse_judan_map(text):
    """解析剧单文本 → {normalized_title: original_title}。"""
    m = {}
    if not text:
        return m
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        key = _normalize_title(line)
        if key and key not in m:
            m[key] = line
    return m


def _parse_material_map(text):
    """解析素材文本 → {normalized_title: {"rawTitle": str, "ids": [str]}}。"""
    m = {}
    if not text:
        return m
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(r'^(.*?)(\s+\d{10,}(?:\s+\d{10,})*)$', line)
        if match:
            raw_title = match.group(1).strip()
            ids = match.group(2).strip().split()
            key = _normalize_title(raw_title)
            if key and ids:
                m[key] = {"rawTitle": raw_title, "ids": ids}
    return m


def _parse_drama_blocks(text):
    """解析短剧数据文本 → [{"name": str, "links": [str]}]。"""
    if not text:
        return []
    # 先尝试整理好的格式
    dramas = _parse_clean_format(text)
    if dramas:
        return dramas
    return _parse_raw_format(text)


def _parse_clean_format(text):
    """解析整理好的格式：剧名 + 链接，空行分隔。"""
    blocks = re.split(r'\n{2,}', text)
    dramas = []
    current = None
    for block in blocks:
        trimmed = block.strip()
        if not trimmed or is_separator_line(trimmed):
            continue
        if re.match(r'^https?://', trimmed, re.I):
            if current:
                link = sanitize_link_text(trimmed)
                if link:
                    current["links"].append(link)
        elif '://' not in trimmed and not re.match(r'^\d+$', trimmed):
            if current and current["links"]:
                dramas.append(current)
            current = {"name": trimmed, "links": []}
    if current and current["links"]:
        dramas.append(current)
    return dramas


def _parse_raw_format(text):
    """解析原始数据格式（token 解析，支持 '短剧名-xxx' 格式）。"""
    tokens = text.split()
    groups = []
    current = None
    for token in tokens:
        if not token or is_separator_line(token):
            continue
        if re.match(r'^https?://', token, re.I):
            link = sanitize_link_text(token)
            if current and link and link not in current["links"]:
                current["links"].append(link)
        elif '-' in token:
            parts = token.split('-')
            current = {"name": parts[-1].strip(), "links": []}
            groups.append(current)
        else:
            if current and not current["links"] and not re.match(r'^\d+$', token):
                current["name"] += token
    return [g for g in groups if g["links"]]


class JuMingToolFrame(ctk.CTkFrame):
    """剧名链接整理工具 — 可嵌入 Frame，移植自 HTML 版本。"""

    def __init__(self, master=None, profile_keys=None, on_add_to_profile=None):
        super().__init__(master, fg_color=C_BG)
        self.cfg = load_config()
        self._profile_keys = profile_keys or list(ALL_PROFILES.keys())
        self._on_add_to_profile = on_add_to_profile  # callback(key, groups_data)
        self._current_tab = "batch"
        self._build_ui()

    def _build_ui(self):
        tab_bar = ctk.CTkFrame(self, fg_color=C_BG, height=40)
        tab_bar.pack(fill="x", padx=16, pady=(12, 0))
        self._tab_btns = {}
        for label, key in [("批量分配", "batch"), ("剧单管理", "titles")]:
            btn = ctk.CTkButton(
                tab_bar, text=label, height=32, width=120,
                font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                corner_radius=R_SM,
                command=lambda k=key: self._switch_tab(k))
            btn.pack(side="left", padx=4)
            self._tab_btns[key] = btn

        container = ctk.CTkFrame(self, fg_color=C_BG)
        container.pack(fill="both", expand=True, padx=16, pady=8)
        self._tab_frames = {
            "batch": self._build_batch_tab(container),
            "titles": self._build_titles_tab(container),
        }
        self._switch_tab("batch")

    def _switch_tab(self, tab):
        self._current_tab = tab
        for key, frame in self._tab_frames.items():
            if key == tab:
                frame.pack(fill="both", expand=True)
            else:
                frame.pack_forget()
        for key, btn in self._tab_btns.items():
            if key == tab:
                btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")
            else:
                btn.configure(fg_color=C_SURFACE, text_color=C_TEXT_2)
        if tab == "titles":
            self._refresh_titles_view()

    # ── 批量分配模式 ──
    def _build_batch_tab(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=C_BG)
        scroll = ctk.CTkScrollableFrame(frame, fg_color=C_BG)
        scroll.pack(fill="both", expand=True)

        # 账户 ID
        ctk.CTkLabel(scroll, text="账户 ID（每行一个纯数字）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._b_ids = ctk.CTkTextbox(scroll, height=80,
                                     font=ctk.CTkFont(family=F_MONO, size=11),
                                     fg_color="#ffffff", border_color=C_BORDER,
                                     border_width=1, corner_radius=R_SM)
        self._b_ids.pack(fill="x", pady=(0, 6))

        # 短剧数据
        ctk.CTkLabel(scroll, text="短剧数据（剧名+链接，空行分隔 或 原始格式）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._b_dramas = ctk.CTkTextbox(scroll, height=120,
                                        font=ctk.CTkFont(family=F_MONO, size=11),
                                        fg_color="#ffffff", border_color=C_BORDER,
                                        border_width=1, corner_radius=R_SM)
        self._b_dramas.pack(fill="x", pady=(0, 6))

        # 素材 ID
        ctk.CTkLabel(scroll, text="素材 ID 数据（可选，格式：剧名 素材ID1 素材ID2 ...）",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(4, 2))
        self._b_material = ctk.CTkTextbox(scroll, height=60,
                                          font=ctk.CTkFont(family=F_MONO, size=11),
                                          fg_color="#ffffff", border_color=C_BORDER,
                                          border_width=1, corner_radius=R_SM)
        self._b_material.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(scroll, text="剧单请在旁边「剧单管理」里追加维护；批量分配会自动读取剧单库匹配完整剧名",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(0, 6))

        # 参数行
        param_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        param_row.pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(param_row, text="每组ID数：",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(side="left")
        self._b_ids_per = ctk.CTkEntry(param_row, width=50, height=28,
                                       font=ctk.CTkFont(family=F_UI, size=11),
                                       fg_color="#ffffff", border_color=C_BORDER,
                                       border_width=1, corner_radius=R_SM)
        self._b_ids_per.insert(0, "6")
        self._b_ids_per.pack(side="left", padx=(2, 12))

        ctk.CTkLabel(param_row, text="每组剧数：",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(side="left")
        self._b_dramas_per = ctk.CTkEntry(param_row, width=50, height=28,
                                          font=ctk.CTkFont(family=F_UI, size=11),
                                          fg_color="#ffffff", border_color=C_BORDER,
                                          border_width=1, corner_radius=R_SM)
        self._b_dramas_per.insert(0, "3")
        self._b_dramas_per.pack(side="left", padx=(2, 12))

        ctk.CTkLabel(param_row, text="行间距：",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(side="left")
        self._b_spacing = ctk.CTkOptionMenu(
            param_row, values=["0", "1", "2", "3"], width=60, height=28,
            font=ctk.CTkFont(family=F_UI, size=11))
        self._b_spacing.set("1")
        self._b_spacing.pack(side="left", padx=2)

        # 按钮行
        btn_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(4, 6))
        ctk.CTkButton(btn_row, text="🚀 批量分配", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                      fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                      text_color="#ffffff", corner_radius=R_SM,
                      command=self._batch_process).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🗑 清空", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_RED, border_width=1, border_color="#fecaca",
                      corner_radius=R_SM,
                      command=self._batch_clear).pack(side="left")

        # 状态
        self._b_stats = ctk.CTkLabel(scroll, text="",
                                     font=ctk.CTkFont(family=F_UI, size=11),
                                     text_color=C_PRIMARY)
        self._b_stats.pack(anchor="w", pady=(2, 2))

        # 输出
        ctk.CTkLabel(scroll, text="分配结果",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._b_output = ctk.CTkTextbox(scroll, height=250,
                                        font=ctk.CTkFont(family=F_MONO, size=11),
                                        fg_color="#ffffff", border_color=C_BORDER,
                                        border_width=1, corner_radius=R_SM,
                                        state="disabled")
        self._b_output.pack(fill="x", pady=(0, 8))

        # 「添加到配置」按钮行
        add_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        add_row.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(add_row, text="将分配结果添加到：",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left")
        self._add_btns = {}
        for pk in self._profile_keys:
            b = ctk.CTkButton(
                add_row, text=f"➕ 到{pk}", height=28, width=112,
                font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                fg_color=C_SURFACE, hover_color=C_HOVER,
                text_color=C_PRIMARY, border_width=1, border_color=C_PRIMARY,
                corner_radius=R_SM,
                command=lambda k=pk: self._add_to_profile(k))
            b.pack(side="left", padx=3)
            self._add_btns[pk] = b
        self._add_status = ctk.CTkLabel(add_row, text="",
                                        font=ctk.CTkFont(family=F_UI, size=10),
                                        text_color=C_PRIMARY)
        self._add_status.pack(side="left", padx=8)

        return frame

    # ── 剧单管理 ──
    def _build_titles_tab(self, parent):
        frame = ctk.CTkFrame(parent, fg_color=C_BG)
        scroll = ctk.CTkScrollableFrame(frame, fg_color=C_BG)
        scroll.pack(fill="both", expand=True)

        ctk.CTkLabel(scroll, text="新增剧单（每行一个完整剧名；每次点击追加，不覆盖已有剧单）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._titles_input = ctk.CTkTextbox(scroll, height=150,
                                            font=ctk.CTkFont(family=F_MONO, size=11),
                                            fg_color="#ffffff", border_color=C_BORDER,
                                            border_width=1, corner_radius=R_SM)
        self._titles_input.pack(fill="x", pady=(0, 6))

        btn_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(4, 6))
        ctk.CTkButton(btn_row, text="➕ 追加到剧单", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                      fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                      text_color="#ffffff", corner_radius=R_SM,
                      command=self._append_titles).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🗑 清空输入", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_RED, border_width=1, border_color="#fecaca",
                      corner_radius=R_SM,
                      command=lambda: self._titles_input.delete("1.0", "end")).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🔄 刷新剧单", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM,
                      command=self._refresh_titles_view).pack(side="left")

        self._titles_stats = ctk.CTkLabel(scroll, text="",
                                          font=ctk.CTkFont(family=F_UI, size=11),
                                          text_color=C_PRIMARY)
        self._titles_stats.pack(anchor="w", pady=(2, 2))

        ctk.CTkLabel(scroll, text="当前剧单库（批量分配会自动用这里匹配完整剧名）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(8, 2))
        self._titles_output = ctk.CTkTextbox(scroll, height=320,
                                             font=ctk.CTkFont(family=F_MONO, size=11),
                                             fg_color="#ffffff", border_color=C_BORDER,
                                             border_width=1, corner_radius=R_SM,
                                             state="disabled")
        self._titles_output.pack(fill="x", pady=(0, 8))
        return frame

    def _get_titles(self):
        titles = (self.cfg.get("common") or {}).get("drama_titles") or []
        return [str(x).strip() for x in titles if str(x).strip()]

    def _set_titles(self, titles):
        seen = set()
        normalized = []
        for title in titles:
            title = str(title).strip()
            key = _normalize_title(title)
            if title and key and key not in seen:
                seen.add(key)
                normalized.append(title)
        self.cfg.setdefault("common", {})["drama_titles"] = normalized
        save_config(self.cfg)
        return normalized

    def _refresh_titles_view(self):
        if not hasattr(self, "_titles_output"):
            return
        self.cfg = load_config()
        titles = self._get_titles()
        self._titles_output.configure(state="normal")
        self._titles_output.delete("1.0", "end")
        self._titles_output.insert("1.0", "\n".join(titles))
        self._titles_output.configure(state="disabled")
        self._titles_stats.configure(text=f"当前共 {len(titles)} 个剧名", text_color=C_PRIMARY)

    def _append_titles(self):
        text = self._titles_input.get("1.0", "end").strip()
        if not text:
            self._titles_stats.configure(text="请输入要追加的剧名", text_color=C_RED)
            return
        old_titles = self._get_titles()
        old_keys = {_normalize_title(x) for x in old_titles if _normalize_title(x)}
        incoming = [x.strip() for x in text.splitlines() if x.strip()]
        added = []
        skipped = 0
        for title in incoming:
            key = _normalize_title(title)
            if not key or key in old_keys:
                skipped += 1
                continue
            old_keys.add(key)
            added.append(title)
        titles = self._set_titles(old_titles + added)
        self._titles_input.delete("1.0", "end")
        self._refresh_titles_view()
        self._titles_stats.configure(
            text=f"已追加 {len(added)} 个，跳过重复 {skipped} 个，当前共 {len(titles)} 个剧名",
            text_color=C_PRIMARY)

    # ── 批量分配逻辑 ──
    def _batch_process(self):
        ids_text = self._b_ids.get("1.0", "end").strip()
        dramas_text = self._b_dramas.get("1.0", "end").strip()
        material_text = self._b_material.get("1.0", "end").strip()

        if not ids_text:
            self._b_stats.configure(text="请输入账户ID", text_color=C_RED)
            return
        if not dramas_text:
            self._b_stats.configure(text="请输入短剧数据", text_color=C_RED)
            return

        try:
            ids_per = int(self._b_ids_per.get() or "6")
        except ValueError:
            ids_per = 6
        try:
            dramas_per = int(self._b_dramas_per.get() or "3")
        except ValueError:
            dramas_per = 3
        spacing = int(self._b_spacing.get() or "1")
        gap = "\n" * (spacing + 1)

        # 解析 ID
        all_ids = [l.strip() for l in ids_text.splitlines() if re.match(r'^\d+$', l.strip())]
        if not all_ids:
            self._b_stats.configure(text="未识别到有效的账户ID（纯数字）", text_color=C_RED)
            return

        # 解析短剧
        all_dramas = _parse_drama_blocks(dramas_text)
        if not all_dramas:
            self._b_stats.configure(text="未识别到有效的短剧数据", text_color=C_RED)
            return

        material_map = _parse_material_map(material_text)
        has_material = len(material_map) > 0

        # 剧单修正
        self.cfg = load_config()
        judan_map = _parse_judan_map("\n".join(self._get_titles()))
        matched_count = 0
        if judan_map:
            for d in all_dramas:
                matched = _fuzzy_find(d["name"], judan_map)
                if matched and matched != d["name"]:
                    d["name"] = matched
                    matched_count += 1

        # 分组
        id_groups = [all_ids[i:i + ids_per] for i in range(0, len(all_ids), ids_per)]
        drama_groups = [all_dramas[i:i + dramas_per] for i in range(0, len(all_dramas), dramas_per)]
        pair_count = min(len(id_groups), len(drama_groups))

        if pair_count == 0:
            self._b_stats.configure(text="无法配对，请检查数据", text_color=C_RED)
            return

        all_contents = []
        for i in range(pair_count):
            ids = id_groups[i]
            dramas = drama_groups[i]
            content = "\n".join(ids) + "\n\n"
            drama_parts = []
            for drama in dramas:
                part = drama["name"] + gap + gap.join(drama["links"])
                if has_material:
                    mi = _fuzzy_find(drama["name"], material_map)
                    if mi and mi["ids"]:
                        part += gap + "\n".join(mi["ids"])
                drama_parts.append(part)
            content += "\n\n\n".join(drama_parts)
            drama_names = "、".join(d["name"] for d in dramas)
            header = f"═══ 第 {i+1} 组（{len(ids)} 个ID，{len(dramas)} 部剧：{drama_names}）═══"
            all_contents.append(header + "\n" + content)

        result = ("\n" + "=" * 50 + "\n").join(all_contents)
        self._b_output.configure(state="normal")
        self._b_output.delete("1.0", "end")
        self._b_output.insert("1.0", result)
        self._b_output.configure(state="disabled")
        self._b_stats.configure(
            text=f"成功分配 {pair_count} 组（共 {len(all_ids)} 个ID，{len(all_dramas)} 部剧，剧单匹配 {matched_count} 个）",
            text_color=C_PRIMARY)

    def _batch_clear(self):
        self._b_ids.delete("1.0", "end")
        self._b_dramas.delete("1.0", "end")
        self._b_material.delete("1.0", "end")
        self._b_output.configure(state="normal")
        self._b_output.delete("1.0", "end")
        self._b_output.configure(state="disabled")
        self._b_stats.configure(text="已清空", text_color=C_DIM)

    def _add_to_profile(self, profile_key):
        """将批量分配结果解析为 groups 并添加到指定 profile。"""
        self._b_output.configure(state="normal")
        text = self._b_output.get("1.0", "end").strip()
        self._b_output.configure(state="disabled")
        if not text:
            self._add_status.configure(text="没有分配结果", text_color=C_RED)
            return
        # 按 ═══ 分隔符拆分各组
        blocks = re.split(r'(?:\n\s*(?:═{3,}[^\n]*═{3,}|={3,})\s*\n?)', text)
        blocks = [b.strip() for b in blocks if b.strip()]
        if not blocks:
            self._add_status.configure(text="无法解析结果", text_color=C_RED)
            return
        # 解析每个 block 为 group dict
        _silent = logging.getLogger("_silent_add")
        _silent.addHandler(logging.NullHandler())
        _silent.propagate = False
        new_groups = []
        cleaned_count = 0
        for block in blocks:
            ids, dramas = _parse_single_group(block, _silent)
            if ids or dramas:
                group, changed = sanitize_config_groups([{
                    "account_ids": [str(x) for x in ids],
                    "dramas": [
                        {
                            "name": d.get("name", ""),
                            "click": d.get("click", ""),
                            "show": d.get("show", ""),
                            "video": d.get("video", ""),
                            "material_ids": list(d.get("material_ids", [])),
                        }
                        for d in dramas
                    ],
                }])
                if changed:
                    cleaned_count += 1
                new_groups.extend(group)
        if not new_groups:
            self._add_status.configure(text="解析失败，请检查结果格式", text_color=C_RED)
            return
        # 回调给 SettingsFrame 写入
        if self._on_add_to_profile:
            try:
                self._on_add_to_profile(profile_key, new_groups)
            except Exception as e:
                self._add_status.configure(text=f"添加失败：{e}", text_color=C_RED)
                return
            self._add_status.configure(
                text=f"已添加 {len(new_groups)} 组到「{profile_key}」，可返回主界面直接执行",
                text_color=C_PRIMARY)


def _pc_clean_name(name: str) -> str:
    return _PC_CLEAN_NAME_RE.sub("", name)


def _pc_extract_chinese(text: str) -> str:
    return _PC_CHINESE_ONLY_RE.sub("", text)


def _pc_build_promotion_name(name: str, prefix: str) -> str:
    pure = _pc_clean_name(name)
    return f"{datetime.now().strftime('%Y-%m-%d')}-{prefix}-{pure}"


def _pc_dismiss_overlay(page) -> None:
    page.mouse.click(10, 10)
    page.wait_for_timeout(PROMOTION_CHAIN_CLICK_DELAY)


def _pc_goto_list(page) -> None:
    if PROMOTION_CHAIN_LIST_FRAG in page.url and PROMOTION_CHAIN_DETAIL_FRAG not in page.url:
        return
    _pc_dismiss_overlay(page)
    menu = page.locator(PROMOTION_CHAIN_MENU_SEL).first
    if menu.is_visible():
        menu.click()
    try:
        page.wait_for_url(PROMOTION_CHAIN_LIST_URL_PATTERN, timeout=PROMOTION_CHAIN_NAV_TIMEOUT)
    except Exception:
        pass


def _pc_safe_goto_list(page) -> None:
    try:
        _pc_goto_list(page)
    except Exception:
        pass


def _pc_search_and_find_row(page, name: str, log):
    inp = page.locator(PROMOTION_CHAIN_QUERY_SEL)
    inp.click(click_count=3)
    inp.fill(name)
    page.locator(PROMOTION_CHAIN_SEARCH_BTN).click()
    page.wait_for_timeout(PROMOTION_CHAIN_SEARCH_DELAY)

    rows = page.locator(PROMOTION_CHAIN_ROW_SEL)
    try:
        rows.first.wait_for(state="visible", timeout=PROMOTION_CHAIN_NAV_TIMEOUT)
    except Exception:
        return None, "搜索无任何结果"

    count = rows.count()
    target_chinese = _pc_extract_chinese(name)
    candidates = []
    for i in range(count):
        row = rows.nth(i)
        row_name = row.locator(PROMOTION_CHAIN_BOOK_NAME_SEL).inner_text().strip()
        candidates.append(row_name)
        log(f"  第 {i + 1} 行：{row_name}\n")
        if _pc_extract_chinese(row_name) == target_chinese:
            log(f"  匹配第 {i + 1} 行（汉字一致）\n")
            return row, ""

    clist = "\n".join(f"        {j + 1}. {c}" for j, c in enumerate(candidates))
    return None, f"第一页共 {count} 条结果，汉字均不匹配\n      搜索到的结果：\n{clist}"


def _pc_fill_promotion_and_confirm(page, name: str, prefix: str, is_ios: bool, log):
    btn = page.locator(PROMOTION_CHAIN_GET_LINK_BTN).first
    btn.wait_for(state="visible")
    btn.click()

    promo_input = page.locator(PROMOTION_CHAIN_PROMO_INPUT_SEL).first
    promo_input.wait_for(state="visible", timeout=PROMOTION_CHAIN_ELEMENT_TIMEOUT)

    promo_name = _pc_build_promotion_name(name, prefix)
    promo_input.click(click_count=3)
    promo_input.fill("")
    promo_input.fill(promo_name)
    log(f"  推广链名称：{promo_name}\n")

    if is_ios:
        ios_radio = page.locator(PROMOTION_CHAIN_IOS_RADIO_SEL, has_text="IOS").first
        ios_radio.wait_for(state="visible", timeout=PROMOTION_CHAIN_ELEMENT_TIMEOUT)
        ios_radio.click()

    confirm = page.locator(PROMOTION_CHAIN_CONFIRM_SEL).first
    try:
        confirm.wait_for(state="visible", timeout=PROMOTION_CHAIN_ELEMENT_TIMEOUT)
    except Exception:
        confirm = page.locator(PROMOTION_CHAIN_CONFIRM_FALLBACK).first
        confirm.wait_for(state="visible", timeout=PROMOTION_CHAIN_ELEMENT_TIMEOUT)
    confirm.click()
    confirm.wait_for(state="hidden", timeout=PROMOTION_CHAIN_ELEMENT_TIMEOUT)


def _pc_process_drama(page, task_name: str, name: str, prefix: str, is_ios: bool, log):
    log(f"\n[{task_name}] 搜索：{name}\n")
    try:
        _pc_goto_list(page)
        matched, reason = _pc_search_and_find_row(page, name, log)
        if matched is None:
            _pc_safe_goto_list(page)
            return False, reason
        matched.locator(PROMOTION_CHAIN_VIEW_DETAIL_SEL).click()
        page.wait_for_url(PROMOTION_CHAIN_DETAIL_URL_PATTERN)
        _pc_fill_promotion_and_confirm(page, name, prefix, is_ios, log)
        _pc_dismiss_overlay(page)
        _pc_goto_list(page)
        return True, ""
    except Exception as e:
        reason = f"程序异常：{type(e).__name__}: {e}"
        log(f"  处理失败：{reason}\n")
        _pc_safe_goto_list(page)
        return False, reason


def run_promotion_chain(dramas: list[str], task_indices: list[int], log_func, stop_event: threading.Event):
    from playwright.sync_api import sync_playwright

    tasks = [PROMOTION_CHAIN_TASKS[i] for i in task_indices if 0 <= i < len(PROMOTION_CHAIN_TASKS)]
    if not tasks:
        log_func("未选择有效的执行方向\n")
        return

    log_func("本次输入剧名：\n")
    for d in dramas:
        log_func(f"  {d}\n")
    log_func("\n本次选择执行：\n")
    for tn, _, _ in tasks:
        log_func(f"  {tn}\n")

    failed = []
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(PROMOTION_CHAIN_CDP)
        ctx = browser.contexts[0]
        pages = ctx.pages
        page = next((pg for pg in pages if PROMOTION_CHAIN_LIST_FRAG in pg.url), pages[0])
        page.bring_to_front()
        page.set_default_timeout(PROMOTION_CHAIN_TIMEOUT)

        for task_name, prefix, is_ios in tasks:
            if stop_event.is_set():
                log_func("\n已停止\n")
                return
            log_func(f"\n========== 开始执行：{task_name} ==========\n")
            for idx, name in enumerate(dramas, 1):
                if stop_event.is_set():
                    log_func("\n已停止\n")
                    return
                log_func(f"\n------ {task_name} [{idx}/{len(dramas)}] ------\n")
                ok, reason = _pc_process_drama(page, task_name, name, prefix, is_ios, log_func)
                if not ok:
                    log_func(f"  失败：{name}，{reason}\n")
                    failed.append((task_name, name, reason))
            log_func(f"\n========== 执行结束：{task_name} ==========\n")

    if failed:
        log_func("\n失败剧名：\n")
        grouped: dict[str, list[str]] = {}
        for tn, n, _ in failed:
            grouped.setdefault(tn, []).append(n)
        for tn, _, _ in tasks:
            if tn not in grouped:
                continue
            log_func(f"{tn}\n")
            seen = set()
            for n in grouped[tn]:
                if n in seen:
                    continue
                seen.add(n)
                log_func(f"  {n}\n")
    else:
        log_func("\n选择的推广链已全部执行完成 ✅\n")


class PromotionChainToolFrame(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._stop_event = threading.Event()
        self._queue = queue.Queue()
        self._running = False
        self._task_vars = {}
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="推广链生成",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="输入剧名并选择方向后，会调用推广链汇总脚本生成对应推广链。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))

        ctk.CTkLabel(wrap, text="剧名（每行一个）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        self._drama_box = ctk.CTkTextbox(wrap, height=150,
                                         font=ctk.CTkFont(family=F_MONO, size=12),
                                         fg_color="#ffffff", text_color=C_TEXT,
                                         border_color=C_BORDER, border_width=1,
                                         corner_radius=R_SM)
        self._drama_box.pack(fill="x", pady=(0, 10))

        task_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        task_row.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(task_row, text="执行方向：",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left", padx=(0, 8))
        for label, value in PROMOTION_TASK_OPTIONS:
            var = tk.BooleanVar(value=True)
            self._task_vars[value] = var
            ctk.CTkCheckBox(task_row, text=label, variable=var,
                            font=ctk.CTkFont(family=F_UI, size=12),
                            text_color=C_TEXT_2, fg_color=C_PRIMARY,
                            hover_color=C_PRIMARY_H, border_color=C_BORDER).pack(side="left", padx=(0, 12))

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始生成", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        self._stop_btn = ctk.CTkButton(btn_row, text="停止", height=34, width=72,
                                       font=ctk.CTkFont(family=F_UI, size=12),
                                       fg_color=C_SURFACE, hover_color=C_HOVER,
                                       text_color=C_RED, border_width=1, border_color="#fecaca",
                                       corner_radius=R_SM, command=self._stop)
        self._stop_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))
        self._log_box = ctk.CTkTextbox(wrap, height=360,
                                       font=ctk.CTkFont(family=F_MONO, size=18),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        dramas = [x.strip() for x in self._drama_box.get("1.0", "end").splitlines() if x.strip()]
        if not dramas:
            self._status.configure(text="请输入剧名", text_color=C_RED)
            return
        selected = [value for value, var in self._task_vars.items() if var.get()]
        if not selected:
            self._status.configure(text="请选择至少一个执行方向", text_color=C_RED)
            return
        self._clear_log()
        task_names = [PROMOTION_CHAIN_TASKS[i][0] for i in selected]
        self._append_log(f"剧名数量：{len(dramas)}，方向：{'、'.join(task_names)}\n\n")
        self._running = True
        self._stop_event.clear()
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")

        def _worker():
            try:
                run_promotion_chain(dramas, selected, self._queue.put, self._stop_event)
                self._queue.put(("__DONE__", 0))
            except Exception as e:
                self._queue.put(f"执行失败：{e}\n")
                self._queue.put(("__DONE__", -1))

        threading.Thread(target=_worker, daemon=True).start()

    def _stop(self):
        if self._running:
            self._stop_event.set()
            self._append_log("\n已请求停止推广链生成\n")

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._stop_event.clear()
                    self._start_btn.configure(state="normal")
                    code = item[1]
                    if code == 0:
                        self._status.configure(text="执行完成", text_color=C_GREEN)
                    else:
                        self._status.configure(text=f"执行结束，退出码 {code}", text_color=C_ORANGE)
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


class PromotionSplitToolFrame(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._queue = queue.Queue()
        self._running = False
        self._result_boxes = {}
        self._result_counts = {}
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="推广链拆分",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="内置处理下载目录里的推广链统计表，按 IOS/Android 每留/七留拆分整理。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))
        ctk.CTkLabel(wrap, text=f"处理目录：{PROMOTION_SPLIT_DEFAULT_DIR}",
                     font=ctk.CTkFont(family=F_MONO, size=11),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(0, 10))

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始拆分", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))

        result_grid = ctk.CTkFrame(wrap, fg_color=C_BG)
        result_grid.pack(fill="both", expand=True, pady=(0, 8))
        for index, (group_key, title) in enumerate(PROMOTION_SPLIT_DISPLAY_ORDER):
            card = ctk.CTkFrame(result_grid, fg_color=C_SURFACE, border_color=C_BORDER, border_width=1, corner_radius=R_MD)
            card.grid(row=index // 2, column=index % 2, sticky="nsew", padx=(0, 8) if index % 2 == 0 else (8, 0), pady=(0, 10))
            result_grid.grid_columnconfigure(index % 2, weight=1)
            result_grid.grid_rowconfigure(index // 2, weight=1)
            header = ctk.CTkFrame(card, fg_color=C_SURFACE)
            header.pack(fill="x", padx=10, pady=(8, 4))
            ctk.CTkLabel(header, text=title,
                         font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                         text_color=C_TEXT).pack(side="left")
            ctk.CTkButton(header, text="一键复制", height=28, width=76,
                          font=ctk.CTkFont(family=F_UI, size=11),
                          fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                          text_color="#ffffff", corner_radius=R_SM,
                          command=lambda k=group_key: self._copy_result(k)).pack(side="right")
            box = ctk.CTkTextbox(card, height=150,
                                 font=ctk.CTkFont(family=F_MONO, size=11),
                                 fg_color="#ffffff", text_color=C_TEXT,
                                 border_color=C_BORDER, border_width=1,
                                 corner_radius=R_SM, state="disabled")
            box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self._result_boxes[group_key] = box

        self._log_box = ctk.CTkTextbox(wrap, height=130,
                                       font=ctk.CTkFont(family=F_MONO, size=11),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _set_result_text(self, group_key, text):
        box = self._result_boxes.get(group_key)
        if not box:
            return
        box.configure(state="normal")
        box.delete("1.0", "end")
        box.insert("1.0", text)
        box.configure(state="disabled")

    def _clear_results(self):
        self._result_counts = {}
        for group_key in self._result_boxes:
            self._set_result_text(group_key, "")

    def _copy_result(self, group_key):
        box = self._result_boxes.get(group_key)
        if not box:
            return
        text = box.get("1.0", "end").strip()
        if not text:
            self._status.configure(text="当前板块没有可复制内容", text_color=C_ORANGE)
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self._status.configure(text="已复制", text_color=C_GREEN)

    def _format_rows_for_copy(self, rows):
        lines = []
        for row in rows:
            cells = []
            for idx in PROMOTION_SPLIT_KEEP_COLS:
                value = row[idx] if len(row) > idx else ""
                cells.append("" if value is None else str(value))
            lines.append("\t".join(cells))
        return "\n".join(lines)

    def _ascii_text(self, value):
        return str(value).encode("unicode_escape").decode("ascii")

    def _ascii_counts(self, counts):
        return "{" + ", ".join(f"{self._ascii_text(k)}: {v}" for k, v in counts.items()) + "}"

    def _classify_row(self, row):
        page = str(row[3]) if len(row) > 3 and row[3] else ""
        if "激励" in page:
            return None
        name = str(row[1]) if len(row) > 1 and row[1] else ""
        os_v = str(row[2]) if len(row) > 2 and row[2] else ""
        if "每留" in name:
            return os_v + "-每留"
        if "七留" in name:
            return os_v + "-七留"
        return None

    def _process_file(self, src_path, dst_path):
        wb_src = openpyxl.load_workbook(src_path)
        ws_src = wb_src.active
        groups = {k: [] for k in PROMOTION_SPLIT_GROUP_ORDER}
        for row in ws_src.iter_rows(values_only=True):
            key = self._classify_row(row)
            if key in groups:
                groups[key].append(row)
        total_classified = sum(len(v) for v in groups.values())
        if total_classified == 0:
            wb_src.close()
            raise ValueError("未找到单本推广链数据（文件中可能全部为激励数据）")
        wb = Workbook()
        ws = wb.active
        ws.title = "推广链统计"
        for i, key in enumerate(PROMOTION_SPLIT_GROUP_ORDER):
            ws.append([key])
            for row in groups[key]:
                ws.append([row[j] if len(row) > j else None for j in PROMOTION_SPLIT_KEEP_COLS])
            if i < len(PROMOTION_SPLIT_GROUP_ORDER) - 1:
                for _ in range(PROMOTION_SPLIT_GAP_ROWS):
                    ws.append([])
        try:
            wb.save(dst_path)
        except PermissionError:
            alt_path = self._next_output_path(dst_path)
            try:
                wb.save(alt_path)
                dst_path = alt_path
            except PermissionError:
                ts = datetime.now().strftime('%H%M%S')
                final_path = os.path.join(os.path.dirname(dst_path),
                    os.path.splitext(os.path.basename(dst_path))[0] + f"_{ts}.xlsx")
                wb.save(final_path)
                dst_path = final_path
        counts = {k: len(v) for k, v in groups.items()}
        texts = {k: self._format_rows_for_copy(v) for k, v in groups.items()}
        return counts, texts, dst_path

    def _next_output_path(self, dst_path):
        base, ext = os.path.splitext(dst_path)
        for index in range(1, 100):
            candidate = f"{base}_{index}{ext}"
            if not os.path.exists(candidate):
                return candidate
            try:
                with open(candidate, "a"):
                    pass
                return candidate
            except (PermissionError, OSError):
                continue
        return f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        if openpyxl is None or Workbook is None:
            self._status.configure(text="缺少 openpyxl，无法处理 xlsx", text_color=C_RED)
            return
        self._clear_log()
        self._clear_results()
        self._append_log(f"处理目录：{PROMOTION_SPLIT_DEFAULT_DIR}\n")
        self._append_log("匹配文件：推广链统计_*.xlsx\n\n")
        self._running = True
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")

        def _worker():
            code = 0
            try:
                target_dir = str(PROMOTION_SPLIT_DEFAULT_DIR)
                pattern = os.path.join(target_dir, "推广链统计_*.xlsx")
                files = [f for f in glob.glob(pattern) if "_处理后" not in f and "_processed" not in f and "_样本" not in f and "_gui_test" not in f and "_激励拆分" not in f and "_拆分" not in f]
                if not files:
                    self._queue.put(f"No matching files found: {pattern}\n")
                else:
                    self._queue.put(f"Directory: {target_dir}\n")
                    self._queue.put(f"Found {len(files)} file(s)\n\n")
                    for src in files:
                        base = os.path.splitext(os.path.basename(src))[0]
                        output_dir = os.path.dirname(src)
                        dst = os.path.join(output_dir, base + "_拆分.xlsx")
                        try:
                            counts, texts, dst = self._process_file(src, dst)
                            self._queue.put(("__RESULTS__", texts, counts))
                            self._queue.put(f"[OK] {self._ascii_text(os.path.basename(src))}\n")
                            self._queue.put(f"     -> {self._ascii_text(os.path.basename(dst))}\n")
                            self._queue.put(f"     Groups: {self._ascii_counts(counts)}\n\n")
                        except Exception as e:
                            code = 1
                            import traceback
                            err_msg = traceback.format_exc()
                            self._queue.put(f"[FAIL] {self._ascii_text(os.path.basename(src))}:\n{err_msg}\n")
                            try:
                                err_log = os.path.join(os.path.dirname(__file__), "split_error.log")
                                with open(err_log, "w", encoding="utf-8") as ef:
                                    ef.write(f"src: {src}\ndst: {dst}\n\n{err_msg}")
                            except Exception:
                                pass
            except Exception as e:
                code = 1
                import traceback
                err_msg = traceback.format_exc()
                self._queue.put(f"处理失败：\n{err_msg}\n")
                try:
                    err_log = os.path.join(os.path.dirname(__file__), "split_error.log")
                    with open(err_log, "w", encoding="utf-8") as ef:
                        ef.write(err_msg)
                except Exception:
                    pass
            self._queue.put(("__DONE__", code))

        threading.Thread(target=_worker, daemon=True).start()

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._start_btn.configure(state="normal")
                    code = item[1]
                    if code == 0:
                        self._status.configure(text="执行完成，可直接复制四个板块", text_color=C_GREEN)
                    else:
                        self._status.configure(text=f"执行结束，退出码 {code}", text_color=C_ORANGE)
                elif isinstance(item, tuple) and item[0] == "__RESULTS__":
                    texts = item[1]
                    counts = item[2]
                    self._result_counts = counts
                    for group_key, _title in PROMOTION_SPLIT_DISPLAY_ORDER:
                        self._set_result_text(group_key, texts.get(group_key, ""))
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


class IncentivePromotionSplitToolFrame(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._queue = queue.Queue()
        self._running = False
        self._result_boxes = {}
        self._result_counts = {}
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="激励推广链拆分",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="处理下载目录里的推广链统计表，按激励每留/七留拆分整理。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))
        ctk.CTkLabel(wrap, text=f"处理目录：{INCENTIVE_SPLIT_DEFAULT_DIR}",
                     font=ctk.CTkFont(family=F_MONO, size=11),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(0, 10))

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始拆分", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))

        result_grid = ctk.CTkFrame(wrap, fg_color=C_BG)
        result_grid.pack(fill="both", expand=True, pady=(0, 8))
        for index, (group_key, title) in enumerate(INCENTIVE_SPLIT_DISPLAY_ORDER):
            card = ctk.CTkFrame(result_grid, fg_color=C_SURFACE, border_color=C_BORDER, border_width=1, corner_radius=R_MD)
            card.grid(row=0, column=index, sticky="nsew", padx=(0, 8) if index == 0 else (8, 0), pady=(0, 10))
            result_grid.grid_columnconfigure(index, weight=1)
            result_grid.grid_rowconfigure(0, weight=1)
            header = ctk.CTkFrame(card, fg_color=C_SURFACE)
            header.pack(fill="x", padx=10, pady=(8, 4))
            ctk.CTkLabel(header, text=title,
                         font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                         text_color=C_TEXT).pack(side="left")
            ctk.CTkButton(header, text="一键复制", height=28, width=76,
                          font=ctk.CTkFont(family=F_UI, size=11),
                          fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                          text_color="#ffffff", corner_radius=R_SM,
                          command=lambda k=group_key: self._copy_result(k)).pack(side="right")
            box = ctk.CTkTextbox(card, height=200,
                                 font=ctk.CTkFont(family=F_MONO, size=11),
                                 fg_color="#ffffff", text_color=C_TEXT,
                                 border_color=C_BORDER, border_width=1,
                                 corner_radius=R_SM, state="disabled")
            box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self._result_boxes[group_key] = box

        self._log_box = ctk.CTkTextbox(wrap, height=130,
                                       font=ctk.CTkFont(family=F_MONO, size=11),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _set_result_text(self, group_key, text):
        box = self._result_boxes.get(group_key)
        if not box:
            return
        box.configure(state="normal")
        box.delete("1.0", "end")
        box.insert("1.0", text)
        box.configure(state="disabled")

    def _clear_results(self):
        self._result_counts = {}
        for group_key in self._result_boxes:
            self._set_result_text(group_key, "")

    def _copy_result(self, group_key):
        box = self._result_boxes.get(group_key)
        if not box:
            return
        text = box.get("1.0", "end").strip()
        if not text:
            self._status.configure(text="当前板块没有可复制内容", text_color=C_ORANGE)
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        self._status.configure(text="已复制", text_color=C_GREEN)

    def _format_rows_for_copy(self, rows):
        lines = []
        for row in rows:
            cells = []
            for idx in INCENTIVE_SPLIT_KEEP_COLS:
                value = row[idx] if len(row) > idx else ""
                cells.append("" if value is None else str(value))
            lines.append("\t".join(cells))
        return "\n".join(lines)

    def _classify_row(self, row):
        page = str(row[3]) if len(row) > 3 and row[3] else ""
        if "激励" not in page:
            return None
        name = str(row[1]) if len(row) > 1 and row[1] else ""
        if "每留" in name:
            return "激励-每留"
        if "七留" in name:
            return "激励-七留"
        return None

    def _process_file(self, src_path, dst_path):
        import shutil, tempfile
        wb_src = openpyxl.load_workbook(src_path, data_only=True)
        ws_src = wb_src.active
        groups = {k: [] for k in INCENTIVE_SPLIT_GROUP_ORDER}
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            key = self._classify_row(row)
            if key in groups:
                groups[key].append(list(row))
        wb_src.close()
        wb = Workbook()
        ws = wb.active
        ws.title = "激励推广链统计"
        for i, key in enumerate(INCENTIVE_SPLIT_GROUP_ORDER):
            ws.append([key])
            for row in groups[key]:
                ws.append([row[j] if len(row) > j else None for j in INCENTIVE_SPLIT_KEEP_COLS])
            if i < len(INCENTIVE_SPLIT_GROUP_ORDER) - 1:
                for _ in range(INCENTIVE_SPLIT_GAP_ROWS):
                    ws.append([])
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        wb.save(tmp_path)
        final_path = self._move_to_target(tmp_path, dst_path)
        counts = {k: len(v) for k, v in groups.items()}
        texts = {k: self._format_rows_for_copy(v) for k, v in groups.items()}
        return counts, texts, final_path

    def _move_to_target(self, tmp_path, dst_path):
        import shutil
        base, ext = os.path.splitext(dst_path)
        candidates = [dst_path]
        for i in range(1, 10):
            candidates.append(f"{base}_{i}{ext}")
        app_out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        os.makedirs(app_out, exist_ok=True)
        fname = os.path.basename(dst_path)
        candidates.append(os.path.join(app_out, fname))
        ts = datetime.now().strftime('%H%M%S')
        candidates.append(os.path.join(app_out, os.path.splitext(fname)[0] + f"_{ts}{ext}"))
        for path in candidates:
            try:
                shutil.copy2(tmp_path, path)
                os.remove(tmp_path)
                return path
            except Exception:
                continue
        return tmp_path

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        if openpyxl is None or Workbook is None:
            self._status.configure(text="缺少 openpyxl，无法处理 xlsx", text_color=C_RED)
            return
        self._clear_log()
        self._clear_results()
        self._append_log(f"处理目录：{INCENTIVE_SPLIT_DEFAULT_DIR}\n")
        self._append_log("匹配文件：推广链统计_*.xlsx\n\n")
        self._running = True
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")

        def _worker():
            code = 0
            try:
                target_dir = str(INCENTIVE_SPLIT_DEFAULT_DIR)
                pattern = os.path.join(target_dir, "推广链统计_*.xlsx")
                files = [f for f in glob.glob(pattern) if "_处理后" not in f and "_processed" not in f and "_样本" not in f and "_gui_test" not in f and "_激励拆分" not in f and "_拆分" not in f]
                if not files:
                    self._queue.put(f"No matching files found: {pattern}\n")
                else:
                    self._queue.put(f"Directory: {target_dir}\n")
                    self._queue.put(f"Found {len(files)} file(s)\n\n")
                    for src in files:
                        base = os.path.splitext(os.path.basename(src))[0]
                        output_dir = os.path.dirname(src)
                        dst = os.path.join(output_dir, base + "_激励拆分.xlsx")
                        try:
                            counts, texts, dst = self._process_file(src, dst)
                            self._queue.put(("__RESULTS__", texts, counts))
                            self._queue.put(f"[OK] {os.path.basename(src)}\n")
                            self._queue.put(f"     -> {os.path.basename(dst)}\n")
                            cnt_str = ", ".join(f"{k}: {v}" for k, v in counts.items())
                            self._queue.put(f"     Groups: {{{cnt_str}}}\n\n")
                        except Exception as e:
                            code = 1
                            self._queue.put(f"[FAIL] {os.path.basename(src)}: {e}\n\n")
            except Exception as e:
                code = 1
                self._queue.put(f"处理失败：{e}\n")
            self._queue.put(("__DONE__", code))

        threading.Thread(target=_worker, daemon=True).start()

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._start_btn.configure(state="normal")
                    code = item[1]
                    if code == 0:
                        self._status.configure(text="执行完成，可直接复制板块内容", text_color=C_GREEN)
                    else:
                        self._status.configure(text=f"执行结束，退出码 {code}", text_color=C_ORANGE)
                elif isinstance(item, tuple) and item[0] == "__RESULTS__":
                    texts = item[1]
                    counts = item[2]
                    self._result_counts = counts
                    for group_key, _title in INCENTIVE_SPLIT_DISPLAY_ORDER:
                        self._set_result_text(group_key, texts.get(group_key, ""))
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


class SearchDramaMaterialPushToolFrame(ctk.CTkFrame):
    CDP_URL = "http://127.0.0.1:9222"
    ANDROID_ACCOUNT_ID = "1855367294448011"
    IOS_ACCOUNT_ID = "1859509275615367"
    AD_ACCOUNT_ID = ANDROID_ACCOUNT_ID
    DEFAULT_TIMEOUT = 10_000
    WAIT_AFTER_PUSH = 120_000
    WAIT_AFTER_SEARCH = 5_000
    WAIT_BETWEEN_ROUNDS = 2_000
    MATERIAL_PAGE_URL = "https://www.changdupingtai.com/sale/short-play/manage/material?material_type=2&audit_status=3&page_index=1&page_size=100"
    SEL_DIALOG = ".arco-drawer, .arco-modal"
    SEL_OPTION = ".arco-select-option"
    SEL_SEARCH_INPUT = 'input[placeholder="请输入素材名称"]'
    SEL_SEARCH_BTN = "button.distribution_search:has-text('搜索')"
    SEL_TABLE_ROW = "tbody tr.arco-table-tr"

    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._queue = queue.Queue()
        self._running = False
        self._stop_event = threading.Event()
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="搜索剧名素材推送",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="输入剧名后，工具会直接搜索素材并批量推送，不再调用外部脚本路径。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))

        ctk.CTkLabel(wrap, text="推送方向",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        self._platform_var = ctk.StringVar(value="安卓")
        self._platform_seg = ctk.CTkSegmentedButton(
            wrap, values=["安卓", "iOS"], variable=self._platform_var,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            corner_radius=R_SM, height=34,
            command=self._on_platform_change)
        self._platform_seg.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(wrap, text="广告账户 ID",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        self._account_entry = ctk.CTkEntry(wrap, height=34,
                                           font=ctk.CTkFont(family=F_MONO, size=12),
                                           fg_color="#ffffff", text_color=C_TEXT,
                                           border_color=C_BORDER, border_width=1,
                                           corner_radius=R_SM)
        self._account_entry.insert(0, self.AD_ACCOUNT_ID)
        self._account_entry.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(wrap, text="剧名（每行一个）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        self._drama_box = ctk.CTkTextbox(wrap, height=180,
                                         font=ctk.CTkFont(family=F_MONO, size=12),
                                         fg_color="#ffffff", text_color=C_TEXT,
                                         border_color=C_BORDER, border_width=1,
                                         corner_radius=R_SM)
        self._drama_box.pack(fill="x", pady=(0, 10))

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始推送", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="停止", height=34, width=72,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_RED, border_width=1, border_color="#fecaca",
                      corner_radius=R_SM, command=self._stop).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))
        self._log_box = ctk.CTkTextbox(wrap, height=360,
                                       font=ctk.CTkFont(family=F_MONO, size=18),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _default_account_for_platform(self, platform):
        if platform == "iOS":
            return self.IOS_ACCOUNT_ID
        return self.ANDROID_ACCOUNT_ID

    def _on_platform_change(self, platform):
        self._account_entry.delete(0, "end")
        self._account_entry.insert(0, self._default_account_for_platform(platform))

    def _log(self, text):
        self._queue.put(str(text) + "\n")

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        account_id = self._account_entry.get().strip()
        platform = self._platform_var.get().strip() or "安卓"
        if not account_id:
            self._status.configure(text="请输入广告账户 ID", text_color=C_RED)
            return
        dramas = [x.strip() for x in self._drama_box.get("1.0", "end").splitlines() if x.strip()]
        if not dramas:
            self._status.configure(text="请输入剧名", text_color=C_RED)
            return
        self._clear_log()
        self._append_log("开始内置执行搜索剧名素材推送\n")
        self._append_log(f"推送方向：{platform}\n")
        self._append_log(f"当前推送账户：{account_id}\n")
        self._append_log(f"剧名数量：{len(dramas)}\n\n")
        self._running = True
        self._stop_event.clear()
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")
        threading.Thread(target=self._run_push, args=(dramas, account_id, platform), daemon=True).start()

    def _stop(self):
        if self._running:
            self._stop_event.set()
            self._append_log("\n已请求停止，当前步骤结束后会中断\n")

    def _close_blocking_dialogs(self, page):
        try:
            blocking = page.locator(".arco-drawer-mask, .arco-modal-mask, .arco-drawer, .arco-modal")
            if blocking.count() == 0:
                return
        except Exception:
            return
        self._log("  检测到残留弹窗/遮罩，尝试关闭")
        for _ in range(3):
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(500)
            except Exception:
                pass
        close_selectors = [
            ".arco-drawer .arco-drawer-close-icon",
            ".arco-modal .arco-modal-close-icon",
            ".arco-drawer button[aria-label='Close']",
            ".arco-modal button[aria-label='Close']",
            ".arco-drawer .arco-icon-close",
            ".arco-modal .arco-icon-close",
        ]
        for selector in close_selectors:
            try:
                loc = page.locator(selector)
                for i in range(min(loc.count(), 3)):
                    try:
                        loc.nth(i).click(timeout=1_500, force=True)
                        page.wait_for_timeout(500)
                    except Exception:
                        continue
            except Exception:
                continue
        try:
            page.locator(".arco-drawer-mask, .arco-modal-mask").wait_for(state="hidden", timeout=3_000)
        except Exception:
            try:
                page.evaluate(
                    """
                    () => {
                        document.querySelectorAll('.arco-drawer-wrapper, .arco-modal-wrapper, .arco-drawer-mask, .arco-modal-mask').forEach(el => el.remove());
                        document.body.style.overflow = '';
                    }
                    """
                )
                page.wait_for_timeout(500)
                self._log("  已强制清理残留遮罩")
            except Exception:
                pass

    def _search_drama(self, page, name, platform="安卓"):
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        search_input = page.locator(self._search_input_selector_for_platform(platform)).first
        try:
            search_input.wait_for(state="visible", timeout=self.DEFAULT_TIMEOUT)
        except Exception:
            self._log(f"  当前页面 URL：{page.url}")
            self._log(f"  当前页面标题：{page.title()}")
            raise
        search_input.click()
        search_input.fill("")
        search_input.fill(name)
        self._log(f"  ✔ 已输入剧名：{name}")
        search_btn = page.locator(self.SEL_SEARCH_BTN).first
        search_btn.wait_for(state="visible")
        button_text = search_btn.inner_text(timeout=self.DEFAULT_TIMEOUT).strip()
        if "搜索" not in button_text:
            raise RuntimeError(f"搜索按钮定位异常，当前命中按钮文本：{button_text}")
        search_btn.click()
        self._log("  ✔ 已点击搜索按钮")
        page.wait_for_timeout(self.WAIT_AFTER_SEARCH)

    def _verify_search_result(self, page, name):
        rows = page.locator(self.SEL_TABLE_ROW)
        count = rows.count()
        if count == 0:
            self._log(f"  ⚠️ 搜索结果为空，跳过「{name}」")
            return False
        try:
            first_row = rows.nth(0)
            row_text = first_row.inner_text().strip()
            snippet = row_text.replace("\n", " ")[:60]
            if name in row_text:
                self._log(f"  ✔ 搜索结果验证通过（首行匹配：{snippet}...）")
                return True
            self._log(f"  ⚠️ 首行不包含「{name}」，首行内容：{snippet}...，跳过")
            return False
        except Exception as e:
            self._log(f"  ⚠️ 验证搜索结果出错：{e}，跳过")
            return False

    def _push_all(self, page, name, account_id):
        # ── 切换分页为 100 条/页 ──
        try:
            page_size_sel = page.locator(".arco-pagination-option .arco-select-view")
            if page_size_sel.count() > 0 and "100" not in page_size_sel.first.inner_text(timeout=3000):
                page_size_sel.first.click()
                page.wait_for_timeout(500)
                opt_100 = page.locator(".arco-select-option:visible", has_text="100 条/页")
                if opt_100.count() > 0:
                    opt_100.first.click()
                    page.wait_for_timeout(2000)
                    self._log("  ✔ 已切换分页为 100 条/页")
                else:
                    page.keyboard.press("Escape")
                    self._log("  ⚠️ 未找到 100 条/页选项，保持当前分页")
            else:
                self._log("  ✔ 分页已是 100 条/页，无需切换")
        except Exception as e:
            self._log(f"  ⚠️ 切换分页失败：{e}，继续执行")
        # ── 全选素材 ──
        checkbox = page.locator("table thead .arco-checkbox-mask")
        checkbox.wait_for(state="visible")
        checkbox.click()
        self._log("  ✔ 已点击全选框")
        batch_btn = page.get_by_role("button", name="批量操作")
        batch_btn.wait_for(state="visible")
        batch_btn.click()
        self._log("  ✔ 已点击批量操作")
        push_btn = page.get_by_text("批量推送", exact=True)
        push_btn.wait_for(state="visible")
        push_btn.click()
        self._log("  ✔ 已点击批量推送")
        dialog = page.locator(self.SEL_DIALOG).filter(has_text="批量素材推送")
        dialog.wait_for(state="visible")
        self._log("  ✔ 弹窗已出现")
        dialog.locator(".arco-select", has=page.get_by_placeholder("请选择媒体渠道")).click()
        opt_engine = page.locator(self.SEL_OPTION, has_text="巨量引擎")
        opt_engine.wait_for(state="visible")
        opt_engine.click()
        self._log("  ✔ 已选择媒体渠道：巨量引擎")
        dialog.locator(".arco-select", has=page.get_by_placeholder("请选择投放产品")).click()
        opt_product = page.locator(self.SEL_OPTION, has_text="红果免费短剧")
        opt_product.wait_for(state="visible")
        opt_product.click()
        self._log("  ✔ 已选择投放产品：红果免费短剧")
        page.keyboard.press("Escape")
        account_input = dialog.locator("#ad_account_ids_input")
        account_input.wait_for(state="visible")
        account_input.fill(account_id)
        self._log(f"  ✔ 已输入广告账户 ID：{account_id}")
        confirm_btn = dialog.get_by_role("button", name="确定")
        confirm_btn.wait_for(state="visible")
        confirm_btn.click()
        self._log("  ✔ 已点击确定，等待推送完成...")
        dialog.wait_for(state="hidden", timeout=self.WAIT_AFTER_PUSH)
        self._log(f"  ✅ 「{name}」推送完成（弹窗已自动关闭）")

    def _is_browser_system_page(self, page):
        url = (page.url or "").lower()
        return (
            not url
            or url.startswith("chrome://")
            or url.startswith("devtools://")
            or url.startswith("edge://")
            or url.startswith("about:")
        )

    def _search_input_selector_for_platform(self, platform):
        if platform == "iOS":
            return "#search_value_material_input"
        return self.SEL_SEARCH_INPUT

    def _page_has_search_input(self, page, platform="安卓", timeout=2500):
        try:
            page.locator(self._search_input_selector_for_platform(platform)).first.wait_for(state="visible", timeout=timeout)
            return True
        except Exception:
            return False

    def _click_text_by_js(self, page, text):
        try:
            return page.evaluate(
                """
                targetText => {
                    const tags = ['a', 'button', 'span', 'div', 'li'];
                    const nodes = Array.from(document.querySelectorAll(tags.join(',')));
                    const visibleNodes = nodes.filter(el => {
                        const content = (el.innerText || el.textContent || '').trim();
                        const rect = el.getBoundingClientRect();
                        return content && rect.width > 0 && rect.height > 0;
                    });
                    const exact = visibleNodes.find(el => (el.innerText || el.textContent || '').trim() === targetText);
                    const partial = visibleNodes.find(el => (el.innerText || el.textContent || '').includes(targetText));
                    let target = exact || partial;
                    if (!target) return false;
                    const clickableSelector = 'a,button,[role="button"],[role="menuitem"],.arco-menu-inline-header,.arco-menu-item,.menu_item_title,[class*="menu"],[class*="Menu"]';
                    const clickable = target.closest(clickableSelector);
                    if (clickable) target = clickable;
                    for (let i = 0; i < 4 && target.parentElement; i += 1) {
                        const parent = target.parentElement;
                        const cls = parent.className ? String(parent.className) : '';
                        const role = parent.getAttribute('role') || '';
                        if (role.includes('menu') || cls.includes('menu') || cls.includes('Menu') || cls.includes('arco-menu')) {
                            target = parent;
                            break;
                        }
                        if (target.tagName.toLowerCase() === 'div' && parent.tagName.toLowerCase() === 'div') {
                            target = parent;
                        }
                    }
                    target.scrollIntoView({block: 'center', inline: 'center'});
                    target.dispatchEvent(new MouseEvent('mouseover', {bubbles: true, cancelable: true, view: window}));
                    target.dispatchEvent(new MouseEvent('mousedown', {bubbles: true, cancelable: true, view: window}));
                    target.dispatchEvent(new MouseEvent('mouseup', {bubbles: true, cancelable: true, view: window}));
                    target.click();
                    return true;
                }
                """,
                text,
            )
        except Exception:
            return False

    def _log_page_text_sample(self, page):
        try:
            text = page.locator("body").inner_text(timeout=3_000).strip()
            text = re.sub(r"\s+", " ", text)
            if text:
                self._log(f"当前页面文本片段：{text[:300]}")
        except Exception as e:
            self._log(f"读取页面文本失败：{e}")

    def _click_text_candidate(self, page, text, timeout=3_000):
        if self._click_text_by_js(page, text):
            page.wait_for_timeout(1_000)
            self._log(f"已通过页面文本点击：{text}")
            return True
        candidates = [
            page.get_by_text(text, exact=True),
            page.get_by_role("link", name=text),
            page.get_by_role("button", name=text),
            page.locator("a", has_text=text),
            page.locator("button", has_text=text),
            page.locator("span", has_text=text),
            page.locator("div", has_text=text),
            page.locator("li", has_text=text),
            page.locator(".menu_item_title", has_text=text),
            page.locator(".arco-menu-inline-header", has_text=text),
            page.locator(".arco-menu-item", has_text=text),
        ]
        last_error = None
        for locator in candidates:
            try:
                item = locator.first
                item.scroll_into_view_if_needed(timeout=timeout)
                item.click(timeout=timeout, force=True)
                page.wait_for_timeout(1_000)
                self._log(f"已点击：{text}")
                return True
            except Exception as e:
                last_error = e
        if last_error:
            self._log(f"点击{text}失败：{last_error}")
        return False

    def _click_menu_text_by_js(self, page, text):
        try:
            return page.evaluate(
                """
                targetText => {
                    const selectors = [
                        '.arco-menu-inline-header',
                        '.arco-menu-item',
                        '.menu_item_title',
                        '[role="menuitem"]',
                        'aside a',
                        'aside div',
                        'nav a',
                        'nav div',
                        '[class*="side"] a',
                        '[class*="side"] div',
                        '[class*="menu"] a',
                        '[class*="menu"] div',
                        '[class*="Menu"] a',
                        '[class*="Menu"] div'
                    ];
                    const nodes = Array.from(document.querySelectorAll(selectors.join(',')));
                    const visibleNodes = nodes.filter(el => {
                        const content = (el.innerText || el.textContent || '').trim();
                        const rect = el.getBoundingClientRect();
                        return content && rect.width > 0 && rect.height > 0 && rect.x < Math.min(520, window.innerWidth * 0.55);
                    });
                    let target = visibleNodes.find(el => (el.innerText || el.textContent || '').trim() === targetText);
                    if (!target) {
                        target = visibleNodes.find(el => {
                            const content = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                            return content === targetText || content.split(' ').includes(targetText);
                        });
                    }
                    if (!target) return false;
                    const clickable = target.closest('a,button,[role="button"],[role="menuitem"],.arco-menu-inline-header,.arco-menu-item,.menu_item_title,[class*="menu"],[class*="Menu"]') || target;
                    clickable.scrollIntoView({block: 'center', inline: 'center'});
                    clickable.dispatchEvent(new MouseEvent('mouseover', {bubbles: true, cancelable: true, view: window}));
                    clickable.dispatchEvent(new MouseEvent('mousedown', {bubbles: true, cancelable: true, view: window}));
                    clickable.dispatchEvent(new MouseEvent('mouseup', {bubbles: true, cancelable: true, view: window}));
                    clickable.click();
                    return true;
                }
                """,
                text,
            )
        except Exception:
            return False

    def _click_menu_candidate(self, page, text, timeout=3_000):
        selectors = [
            f'.arco-menu-inline-header:has-text("{text}")',
            f'.arco-menu-item:has-text("{text}")',
            f'.menu_item_title:has-text("{text}")',
            f'[role="menuitem"]:has-text("{text}")',
            f'aside :text-is("{text}")',
            f'nav :text-is("{text}")',
            f'[class*="menu"] :text-is("{text}")',
            f'[class*="Menu"] :text-is("{text}")',
        ]
        for selector in selectors:
            try:
                item = page.locator(selector).first
                item.wait_for(state="visible", timeout=timeout)
                box = item.bounding_box(timeout=timeout)
                if box and box.get("x", 9999) > 520:
                    continue
                item.scroll_into_view_if_needed(timeout=timeout)
                item.click(timeout=timeout, force=True)
                page.wait_for_timeout(1_000)
                self._log(f"已点击左侧菜单：{text}")
                return True
            except Exception:
                continue
        self._log(f"未点到左侧菜单：{text}")
        return False

    def _goto_material_page(self, page, platform="安卓", wait_ms=2_000):
        page.goto(self.MATERIAL_PAGE_URL, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(wait_ms)
        return self._page_has_search_input(page, platform, timeout=8_000)

    def _select_supervisor_account(self, page):
        try:
            current_text = ""
            try:
                current_text = page.locator(".layout-menus-cascader-value").first.inner_text(timeout=2_000)
            except Exception:
                pass
            if "主管账号" in current_text:
                self._log("当前已是主管账号")
                return True
            selectors = [
                ".layout-menus-cascader",
                ".arco-cascader.layout-menus-cascader",
                ".layout-menus-cascader-value",
                ".arco-cascader-view",
            ]
            clicked = False
            for selector in selectors:
                try:
                    item = page.locator(selector).first
                    item.scroll_into_view_if_needed(timeout=3_000)
                    item.click(timeout=3_000, force=True)
                    page.wait_for_timeout(1_000)
                    clicked = True
                    self._log("已打开账号类型选择器")
                    break
                except Exception:
                    continue
            if not clicked:
                self._log("未找到账号类型选择器，尝试直接点击 APP分销 文本")
                self._click_text_candidate(page, "APP分销")
                page.wait_for_timeout(1_000)
            option_selectors = [
                '.arco-cascader-list-item-label:has-text("主管账号")',
                '.arco-cascader-popup .arco-cascader-list-item:has-text("主管账号")',
                '.arco-trigger-popup .arco-cascader-list-item-label:has-text("主管账号")',
                '.arco-trigger-popup div:has-text("主管账号")',
            ]
            for selector in option_selectors:
                try:
                    option = page.locator(selector).first
                    option.wait_for(state="visible", timeout=5_000)
                    option.scroll_into_view_if_needed(timeout=3_000)
                    option.click(timeout=3_000, force=True)
                    page.wait_for_timeout(1_500)
                    current_text = page.locator(".layout-menus-cascader-value").first.inner_text(timeout=3_000)
                    if "主管账号" in current_text:
                        self._log("已切换到主管账号")
                        return True
                except Exception:
                    continue
            self._log("未能切换到主管账号")
            self._log_page_text_sample(page)
            return False
        except Exception as e:
            self._log(f"切换主管账号失败：{e}")
            return False

    def _ensure_material_page_ready(self, page, platform="安卓"):
        try:
            if self._page_has_search_input(page, platform, timeout=4_000):
                return True
        except Exception:
            pass
        try:
            self._log(f"尝试进入素材管理页：{self.MATERIAL_PAGE_URL}")
            if self._goto_material_page(page, platform):
                return True
            self._log(f"直接进入后仍未看到素材搜索框，当前页面：{page.url}")
            self._log_page_text_sample(page)
        except Exception as e:
            self._log(f"直接进入素材页失败，尝试菜单导航：{e}")
            self._log_page_text_sample(page)
        self._select_supervisor_account(page)
        for text in ("推广中心", "短剧列表", "素材管理"):
            self._click_menu_candidate(page, text)
            if self._page_has_search_input(page, platform, timeout=3_000):
                return True
            self._log(f"点击{text}后当前页面：{page.url}")
        try:
            if page.locator('a[href*="/sale/short-play/list"]').count():
                page.locator('a[href*="/sale/short-play/list"]').first.click(timeout=5_000)
                page.wait_for_timeout(1_500)
                if self._page_has_search_input(page, platform, timeout=3_000):
                    return True
        except Exception as e:
            self._log(f"点击短剧列表链接失败：{e}")
        try:
            if page.locator('a[href*="/sale/short-play/manage/material"]').count():
                page.locator('a[href*="/sale/short-play/manage/material"]').first.click(timeout=5_000)
                page.wait_for_timeout(2_000)
                if self._page_has_search_input(page, platform, timeout=5_000):
                    return True
        except Exception as e:
            self._log(f"点击素材管理链接失败：{e}")
        try:
            self._log("菜单导航后再次尝试进入素材管理页")
            if self._goto_material_page(page, platform):
                return True
            self._log(f"仍未找到素材搜索框，当前页面：{page.url}")
        except Exception as e:
            self._log(f"菜单导航后进入素材页失败：{e}")
        return self._page_has_search_input(page, platform, timeout=10_000)

    def _find_material_page(self, context, platform="安卓"):
        pages = [pg for pg in context.pages if not self._is_browser_system_page(pg)]
        def _score(pg):
            url = (pg.url or "").lower()
            if "changdupingtai.com/sale/short-play/manage/material" in url:
                return 0
            if "manage/material" in url:
                return 1
            if "material" in url:
                return 2
            if "changdupingtai.com" in url:
                return 3
            if "oceanengine" in url:
                return 4
            return 5
        pages.sort(key=_score)
        for pg in pages:
            try:
                url = (pg.url or "").lower()
                pg.bring_to_front()
                pg.set_default_timeout(self.DEFAULT_TIMEOUT)
                if "changdupingtai.com" in url:
                    self._log(f"✔ 使用已打开的长读页面：{pg.url}")
                    if self._ensure_material_page_ready(pg, platform):
                        self._log(f"✔ 素材管理页已就绪：{pg.url}")
                        return pg
                    self._log(f"当前长读页面未能进入素材管理页：{pg.url}")
                    continue
                if self._page_has_search_input(pg, platform, timeout=6_000):
                    self._log(f"✔ 已找到包含素材搜索框的页面：{pg.url}")
                    return pg
            except Exception as e:
                self._log(f"检查页面失败：{e}")
                continue
        candidate_urls = [self.MATERIAL_PAGE_URL]
        for url in candidate_urls:
            try:
                self._log(f"未找到可用的长读页面，尝试打开：{url}")
                pg = context.new_page()
                pg.set_default_timeout(self.DEFAULT_TIMEOUT)
                pg.goto(url, wait_until="domcontentloaded", timeout=30_000)
                pg.bring_to_front()
                if self._ensure_material_page_ready(pg, platform):
                    self._log(f"✔ 已打开并进入素材管理页：{pg.url}")
                    return pg
            except Exception as e:
                self._log(f"打开素材管理页失败：{e}")
        opened = [pg.url for pg in context.pages if not self._is_browser_system_page(pg)]
        if opened:
            self._log("当前可用页面：")
            for url in opened:
                self._log(f"  - {url}")
        raise RuntimeError("没有找到素材管理页面，请先在浏览器打开素材管理页后再开始推送")

    def _run_push(self, drama_names, account_id, platform="安卓"):
        code = 0
        try:
            self._log(f"共收到 {len(drama_names)} 个剧名，开始处理...")
            with sync_playwright() as p:
                browser = p.chromium.connect_over_cdp(self.CDP_URL)
                context = browser.contexts[0]
                page = self._find_material_page(context, platform)
                page.bring_to_front()
                page.set_default_timeout(self.DEFAULT_TIMEOUT)
                self._log(f"✔ 使用页面：{page.url}")
                success_count = 0
                skip_count = 0
                fail_count = 0
                failed_names = []
                for idx, name in enumerate(drama_names, 1):
                    if self._stop_event.is_set():
                        self._log("已停止后续处理")
                        break
                    self._log("─" * 50)
                    self._log(f"🔄 [{idx}/{len(drama_names)}] 处理剧名：{name}")
                    try:
                        self._search_drama(page, name, platform)
                        if not self._verify_search_result(page, name):
                            skip_count += 1
                            failed_names.append(name)
                            self._close_blocking_dialogs(page)
                            page.wait_for_timeout(self.WAIT_BETWEEN_ROUNDS)
                            continue
                        self._push_all(page, name, account_id)
                        success_count += 1
                        page.wait_for_timeout(self.WAIT_BETWEEN_ROUNDS)
                    except Exception as e:
                        self._log(f"  ❌ 处理「{name}」失败：{e}")
                        fail_count += 1
                        failed_names.append(name)
                        try:
                            self._close_blocking_dialogs(page)
                            page.keyboard.press("Escape")
                            page.wait_for_timeout(self.WAIT_BETWEEN_ROUNDS)
                            page.locator(self.SEL_DIALOG).filter(has_text="批量素材推送").wait_for(
                                state="hidden", timeout=self.DEFAULT_TIMEOUT)
                        except Exception:
                            pass
                self._log("=" * 50)
                self._log(f"🎉 全部完成！共处理 {len(drama_names)} 个剧名")
                self._log(f"   ✅ 成功推送：{success_count} 个")
                self._log(f"   ⚠️ 跳过（搜索无结果或不匹配）：{skip_count} 个")
                self._log(f"   ❌ 失败：{fail_count} 个")
                self._log("=" * 50)
                if failed_names:
                    self._log("📋 失败/跳过的剧名清单：")
                    for n in failed_names:
                        self._log(n)
        except Exception as e:
            code = -1
            self._log(f"执行失败：{e}")
        self._queue.put(("__DONE__", code))

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._start_btn.configure(state="normal")
                    code = item[1]
                    if code == 0:
                        self._status.configure(text="执行完成", text_color=C_GREEN)
                    else:
                        self._status.configure(text=f"执行结束，退出码 {code}", text_color=C_ORANGE)
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


# ════════════════════════════════════════════════════════════════════
#  🚀 激励推广链生成工具面板
# ════════════════════════════════════════════════════════════════════
class IncentivePromoChainToolFrame(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._stop_event = threading.Event()
        self._queue = queue.Queue()
        self._running = False
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="激励推广链生成",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="连接浏览器后，为「短剧激励活动」批量创建推广链接（格式：日期-组N-每留/七留）。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))

        row1 = ctk.CTkFrame(wrap, fg_color=C_BG)
        row1.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(row1, text="执行次数：",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left", padx=(0, 4))
        self._count_var = tk.StringVar(value="10")
        ctk.CTkEntry(row1, width=80, height=32, textvariable=self._count_var,
                     font=ctk.CTkFont(family=F_MONO, size=12),
                     fg_color="#ffffff", text_color=C_TEXT,
                     border_color=C_BORDER, border_width=1,
                     corner_radius=R_SM).pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="方向后缀：",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left", padx=(0, 4))
        self._suffix_var = tk.StringVar(value="每留")
        ctk.CTkOptionMenu(row1, values=["每留", "七留"], variable=self._suffix_var,
                          width=100, height=32,
                          font=ctk.CTkFont(family=F_UI, size=12),
                          fg_color="#ffffff", text_color=C_TEXT,
                          button_color=C_PRIMARY, button_hover_color=C_PRIMARY_H,
                          corner_radius=R_SM).pack(side="left")

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始生成", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        self._stop_btn = ctk.CTkButton(btn_row, text="停止", height=34, width=72,
                                       font=ctk.CTkFont(family=F_UI, size=12),
                                       fg_color=C_SURFACE, hover_color=C_HOVER,
                                       text_color=C_RED, border_width=1, border_color="#fecaca",
                                       corner_radius=R_SM, command=self._stop)
        self._stop_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))
        self._log_box = ctk.CTkTextbox(wrap, height=360,
                                       font=ctk.CTkFont(family=F_MONO, size=18),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        try:
            count = int(self._count_var.get().strip())
            if count <= 0:
                raise ValueError
        except ValueError:
            self._status.configure(text="请输入有效的正整数", text_color=C_RED)
            return
        suffix = self._suffix_var.get()
        self._clear_log()
        self._running = True
        self._stop_event.clear()
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")

        def _worker():
            try:
                run_incentive_promo_chain(count, suffix, self._queue.put, self._stop_event)
                self._queue.put(("__DONE__", 0))
            except Exception as e:
                self._queue.put(f"执行失败：{e}\n")
                self._queue.put(("__DONE__", -1))

        threading.Thread(target=_worker, daemon=True).start()

    def _stop(self):
        if self._running:
            self._stop_event.set()
            self._append_log("\n已请求停止\n")

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._stop_event.clear()
                    self._start_btn.configure(state="normal")
                    self._status.configure(
                        text="执行完成" if item[1] == 0 else f"执行结束（退出码 {item[1]}）",
                        text_color=C_GREEN if item[1] == 0 else C_ORANGE)
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


# ════════════════════════════════════════════════════════════════════
#  📦 激励素材推送工具面板
# ════════════════════════════════════════════════════════════════════
class IncentivePushToolFrame(ctk.CTkFrame):
    def __init__(self, master=None):
        super().__init__(master, fg_color=C_BG)
        self._stop_event = threading.Event()
        self._queue = queue.Queue()
        self._running = False
        self._build_ui()
        self.after(120, self._poll_output)

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="激励素材推送",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="连接浏览器后，在素材管理页面逐页全选并批量推送到巨量引擎。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))

        row1 = ctk.CTkFrame(wrap, fg_color=C_BG)
        row1.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(row1, text="广告账户 ID：",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left", padx=(0, 4))
        self._account_var = tk.StringVar(value="1855641147536460")
        ctk.CTkEntry(row1, width=220, height=32, textvariable=self._account_var,
                     font=ctk.CTkFont(family=F_MONO, size=12),
                     fg_color="#ffffff", text_color=C_TEXT,
                     border_color=C_BORDER, border_width=1,
                     corner_radius=R_SM).pack(side="left")

        btn_row = ctk.CTkFrame(wrap, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(0, 8))
        self._start_btn = ctk.CTkButton(btn_row, text="🚀 开始推送", height=34,
                                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                                        fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                                        text_color="#ffffff", corner_radius=R_SM,
                                        command=self._start)
        self._start_btn.pack(side="left", padx=(0, 8))
        self._stop_btn = ctk.CTkButton(btn_row, text="停止", height=34, width=72,
                                       font=ctk.CTkFont(family=F_UI, size=12),
                                       fg_color=C_SURFACE, hover_color=C_HOVER,
                                       text_color=C_RED, border_width=1, border_color="#fecaca",
                                       corner_radius=R_SM, command=self._stop)
        self._stop_btn.pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="清空日志", height=34, width=88,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._clear_log).pack(side="left")

        self._status = ctk.CTkLabel(wrap, text="就绪",
                                    font=ctk.CTkFont(family=F_UI, size=11),
                                    text_color=C_DIM)
        self._status.pack(anchor="w", pady=(0, 4))
        self._log_box = ctk.CTkTextbox(wrap, height=360,
                                       font=ctk.CTkFont(family=F_MONO, size=18),
                                       fg_color=C_LOG_BG, text_color=C_LOG_FG,
                                       border_color=C_BORDER, border_width=1,
                                       corner_radius=R_SM, state="disabled")
        self._log_box.pack(fill="both", expand=True)

    def _append_log(self, text):
        self._log_box.configure(state="normal")
        self._log_box.insert("end", text)
        self._log_box.see("end")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box.configure(state="normal")
        self._log_box.delete("1.0", "end")
        self._log_box.configure(state="disabled")

    def _start(self):
        if self._running:
            self._status.configure(text="正在运行中", text_color=C_ORANGE)
            return
        account_id = self._account_var.get().strip()
        if not account_id:
            self._status.configure(text="请输入广告账户 ID", text_color=C_RED)
            return
        self._clear_log()
        self._running = True
        self._stop_event.clear()
        self._status.configure(text="运行中", text_color=C_PRIMARY)
        self._start_btn.configure(state="disabled")

        def _worker():
            try:
                run_incentive_push(account_id, self._queue.put, self._stop_event)
                self._queue.put(("__DONE__", 0))
            except Exception as e:
                self._queue.put(f"执行失败：{e}\n")
                self._queue.put(("__DONE__", -1))

        threading.Thread(target=_worker, daemon=True).start()

    def _stop(self):
        if self._running:
            self._stop_event.set()
            self._append_log("\n已请求停止\n")

    def _poll_output(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if isinstance(item, tuple) and item[0] == "__DONE__":
                    self._running = False
                    self._stop_event.clear()
                    self._start_btn.configure(state="normal")
                    self._status.configure(
                        text="执行完成" if item[1] == 0 else f"执行结束（退出码 {item[1]}）",
                        text_color=C_GREEN if item[1] == 0 else C_ORANGE)
                else:
                    self._append_log(str(item))
        except queue.Empty:
            pass
        self.after(120, self._poll_output)


# ════════════════════════════════════════════════════════════════════
#  🔗 激励链接分配工具面板
# ════════════════════════════════════════════════════════════════════
class IncentiveLinkToolFrame(ctk.CTkFrame):

    def __init__(self, master=None, on_add_to_profile=None):
        super().__init__(master, fg_color=C_BG)
        self._on_add_to_profile = on_add_to_profile
        self._build_ui()

    def _build_ui(self):
        wrap = ctk.CTkFrame(self, fg_color=C_BG)
        wrap.pack(fill="both", expand=True, padx=16, pady=12)
        ctk.CTkLabel(wrap, text="激励链接分配",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(0, 4))
        ctk.CTkLabel(wrap, text="将推广链数据 + 账户ID + 素材ID 按组整理分配，生成 ids.txt 格式并可直接写入配置。",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(wrap, fg_color=C_BG)
        scroll.pack(fill="both", expand=True)

        ctk.CTkLabel(scroll, text="原始数据（Tab分隔：日期-组X-标签 + 点击链接 + 展示链接 + 有效播放链接）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._raw_input = ctk.CTkTextbox(scroll, height=100,
                                         font=ctk.CTkFont(family=F_MONO, size=11),
                                         fg_color="#ffffff", border_color=C_BORDER,
                                         border_width=1, corner_radius=R_SM)
        self._raw_input.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(scroll, text="账户 ID（每行一个，支持 ID\\t\\t短剧组XX 格式）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._acc_input = ctk.CTkTextbox(scroll, height=100,
                                         font=ctk.CTkFont(family=F_MONO, size=11),
                                         fg_color="#ffffff", border_color=C_BORDER,
                                         border_width=1, corner_radius=R_SM)
        self._acc_input.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(scroll, text="素材 ID（空格 / 换行 / Tab 分隔，每组后重复分配）",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._mat_input = ctk.CTkTextbox(scroll, height=60,
                                         font=ctk.CTkFont(family=F_MONO, size=11),
                                         fg_color="#ffffff", border_color=C_BORDER,
                                         border_width=1, corner_radius=R_SM)
        self._mat_input.pack(fill="x", pady=(0, 6))

        param_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        param_row.pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(param_row, text="每组账户数：",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(side="left")
        self._ids_per = ctk.CTkEntry(param_row, width=50, height=28,
                                     font=ctk.CTkFont(family=F_UI, size=11),
                                     fg_color="#ffffff", border_color=C_BORDER,
                                     border_width=1, corner_radius=R_SM)
        self._ids_per.insert(0, "6")
        self._ids_per.pack(side="left", padx=(2, 12))
        ctk.CTkLabel(param_row, text="行间距：",
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_TEXT_2).pack(side="left")
        self._spacing = ctk.CTkOptionMenu(
            param_row, values=["0", "1", "2", "3"], width=60, height=28,
            font=ctk.CTkFont(family=F_UI, size=11))
        self._spacing.set("1")
        self._spacing.pack(side="left", padx=2)

        btn_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        btn_row.pack(fill="x", pady=(4, 6))
        ctk.CTkButton(btn_row, text="🚀 开始分配", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                      fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                      text_color="#ffffff", corner_radius=R_SM,
                      command=self._process).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="🗑 清空", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_RED, border_width=1, border_color="#fecaca",
                      corner_radius=R_SM,
                      command=self._clear_all).pack(side="left")

        self._stats = ctk.CTkLabel(scroll, text="",
                                   font=ctk.CTkFont(family=F_UI, size=11),
                                   text_color=C_PRIMARY)
        self._stats.pack(anchor="w", pady=(2, 2))

        ctk.CTkLabel(scroll, text="分配结果",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", pady=(4, 2))
        self._output = ctk.CTkTextbox(scroll, height=250,
                                      font=ctk.CTkFont(family=F_MONO, size=11),
                                      fg_color="#ffffff", border_color=C_BORDER,
                                      border_width=1, corner_radius=R_SM,
                                      state="disabled")
        self._output.pack(fill="x", pady=(0, 8))

        add_row = ctk.CTkFrame(scroll, fg_color=C_BG)
        add_row.pack(fill="x", pady=(0, 8))
        ctk.CTkLabel(add_row, text="添加到配置：",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left")
        for pk in list(INCENTIVE_PROFILES.keys()):
            ctk.CTkButton(
                add_row, text=f"➕ {pk}", height=28, width=120,
                font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                fg_color=C_SURFACE, hover_color=C_HOVER,
                text_color=C_PRIMARY, border_width=1, border_color=C_PRIMARY,
                corner_radius=R_SM,
                command=lambda k=pk: self._add_to_profile(k)).pack(side="left", padx=3)
        self._add_status = ctk.CTkLabel(add_row, text="",
                                        font=ctk.CTkFont(family=F_UI, size=10),
                                        text_color=C_PRIMARY)
        self._add_status.pack(side="left", padx=8)

    def _process(self):
        raw_text = self._raw_input.get("1.0", "end").strip()
        acc_text = self._acc_input.get("1.0", "end").strip()
        mat_text = self._mat_input.get("1.0", "end").strip()

        if not raw_text:
            self._stats.configure(text="请输入原始数据", text_color=C_RED)
            return
        if not acc_text:
            self._stats.configure(text="请输入账户ID", text_color=C_RED)
            return

        try:
            ids_per = int(self._ids_per.get() or "6")
            if ids_per <= 0:
                ids_per = 6
        except ValueError:
            ids_per = 6
        spacing = int(self._spacing.get() or "1")
        group_sep = "\n" * (spacing + 1)

        raw_lines = [l.strip() for l in raw_text.splitlines() if l.strip()]
        data_rows = []
        for line in raw_lines:
            parts = line.split("\t")
            if len(parts) < 2:
                continue
            col0 = parts[0].strip()
            links = [p.strip() for p in parts[1:] if p.strip()]
            if not links:
                continue
            m = re.search(r"组\d+", col0)
            title = m.group() if m else col0
            data_rows.append({"title": title, "links": links})

        if not data_rows:
            self._stats.configure(text="未识别到有效的原始数据行", text_color=C_RED)
            return

        acc_lines = [l.strip() for l in acc_text.splitlines() if l.strip()]
        account_ids = []
        drama_groups = []
        for line in acc_lines:
            parts = [p.strip() for p in line.split("\t") if p.strip()]
            account_ids.append(parts[0])
            dg = next((p for p in parts if "短剧组" in p), "")
            drama_groups.append(dg)

        if not account_ids:
            self._stats.configure(text="未识别到有效的账户ID", text_color=C_RED)
            return

        material_ids = re.split(r"\s+", mat_text.strip()) if mat_text.strip() else []

        id_groups = [account_ids[i:i + ids_per] for i in range(0, len(account_ids), ids_per)]
        dg_groups = []
        for i in range(0, len(drama_groups), ids_per):
            slc = drama_groups[i:i + ids_per]
            found = next((d for d in slc if d), "")
            dg_groups.append(found)

        total = min(len(data_rows), len(id_groups))
        if total == 0:
            self._stats.configure(text="无法配对，请检查数据", text_color=C_RED)
            return

        blocks = []
        for i in range(total):
            row = data_rows[i]
            grp = id_groups[i]
            title = row["title"]
            dg = dg_groups[i] if i < len(dg_groups) else ""
            id_lines = "\n".join(grp)
            link_lines = "\n".join(l + "\t" for l in row["links"])
            mat_line = " ".join(material_ids) if material_ids else ""
            block = title
            if dg:
                block += "\n" + dg
            block += "\n" + id_lines + "\n\t\t\n" + link_lines
            if mat_line:
                block += "\n" + mat_line
            blocks.append(block)

        result = group_sep.join(blocks)
        self._output.configure(state="normal")
        self._output.delete("1.0", "end")
        self._output.insert("1.0", result)
        self._output.configure(state="disabled")
        self._stats.configure(
            text=f"成功分配 {total} 组（共 {len(account_ids)} 个账户ID）",
            text_color=C_PRIMARY)

        self._parsed_blocks = blocks
        self._parsed_data_rows = data_rows[:total]
        self._parsed_id_groups = id_groups[:total]

    def _clear_all(self):
        self._raw_input.delete("1.0", "end")
        self._acc_input.delete("1.0", "end")
        self._mat_input.delete("1.0", "end")
        self._output.configure(state="normal")
        self._output.delete("1.0", "end")
        self._output.configure(state="disabled")
        self._stats.configure(text="已清空", text_color=C_DIM)
        self._parsed_blocks = []
        self._parsed_data_rows = []
        self._parsed_id_groups = []

    def _add_to_profile(self, profile_key):
        if not hasattr(self, "_parsed_data_rows") or not self._parsed_data_rows:
            self._add_status.configure(text="请先执行分配", text_color=C_RED)
            return
        groups = []
        for i, row in enumerate(self._parsed_data_rows):
            ids = self._parsed_id_groups[i] if i < len(self._parsed_id_groups) else []
            links = row["links"]
            click_url = links[0] if len(links) > 0 else ""
            show_url = links[1] if len(links) > 1 else ""
            play_url = links[2] if len(links) > 2 else ""
            groups.append({
                "account_ids": ids,
                "group_name": row["title"],
                "click_url": click_url,
                "show_url": show_url,
                "play_url": play_url,
                "dramas": [],
            })
        if self._on_add_to_profile:
            try:
                self._on_add_to_profile(profile_key, groups)
            except Exception as e:
                self._add_status.configure(text=f"添加失败：{e}", text_color=C_RED)
                return
        self._add_status.configure(
            text=f"✅ 已添加 {len(self._parsed_data_rows)} 组到 {profile_key}",
            text_color=C_GREEN)


# ════════════════════════════════════════════════════════════════════
#  📋 素材历史记录面板
# ════════════════════════════════════════════════════════════════════
class MaterialHistoryFrame(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_BG, **kwargs)
        self._build_ui()

    def _build_ui(self):
        header_row = ctk.CTkFrame(self, fg_color="transparent")
        header_row.pack(fill="x", pady=(0, 16))
        ctk.CTkLabel(header_row, text="📋 素材选择历史",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(side="left")

        self._clear_btn = ctk.CTkButton(
            header_row, text="🗑 清空记录", width=90, height=28,
            font=ctk.CTkFont(family=F_UI, size=11),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color="#ef4444", border_width=1, border_color=C_BORDER,
            corner_radius=R_SM, command=self._clear_history)
        self._clear_btn.pack(side="right", padx=(8, 0))

        self._refresh_btn = ctk.CTkButton(
            header_row, text="🔄 刷新", width=70, height=28,
            font=ctk.CTkFont(family=F_UI, size=11),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY, border_width=1, border_color=C_BORDER,
            corner_radius=R_SM, command=self.refresh)
        self._refresh_btn.pack(side="right")

        stats_row = ctk.CTkFrame(self, fg_color="transparent")
        stats_row.pack(fill="x", pady=(0, 16))
        stats_row.columnconfigure((0, 1), weight=1, uniform="mstat")

        self._stat_total = self._make_stat_card(stats_row, "已使用素材总数", "0", C_PRIMARY)
        self._stat_total.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self._stat_today = self._make_stat_card(stats_row, "今日新增", "0", C_GREEN)
        self._stat_today.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        list_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=R_LG,
                                  border_width=1, border_color=C_BORDER)
        list_frame.pack(fill="both", expand=True)

        list_header = ctk.CTkFrame(list_frame, fg_color="transparent")
        list_header.pack(fill="x", padx=16, pady=(12, 4))
        for col_text, col_w, anchor in [("日期", 80, "w"), ("素材名称", 400, "w")]:
            ctk.CTkLabel(list_header, text=col_text, width=col_w,
                         font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                         text_color=C_DIM, anchor=anchor).pack(side="left", padx=(0, 24))

        sep = ctk.CTkFrame(list_frame, fg_color=C_BORDER, height=1)
        sep.pack(fill="x", padx=16, pady=(0, 4))

        self._list_scroll = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent", corner_radius=0,
            scrollbar_button_color=C_BORDER, scrollbar_button_hover_color=C_DIM)
        self._list_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.refresh()

    def _make_stat_card(self, parent, label, value, color):
        card = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=R_LG,
                            border_width=1, border_color=C_BORDER, height=80)
        card.pack_propagate(False)
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.place(relx=0.5, rely=0.5, anchor="center")
        val_lbl = ctk.CTkLabel(inner, text=value,
                               font=ctk.CTkFont(family=F_UI, size=26, weight="bold"),
                               text_color=color)
        val_lbl.pack()
        ctk.CTkLabel(inner, text=label,
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_DIM).pack(pady=(2, 0))
        card._val_label = val_lbl
        return card

    def _update_stat(self, card, value):
        card._val_label.configure(text=str(value))

    def refresh(self):
        history = load_material_history()
        today_tag = datetime.now().strftime("%m%d")
        today_count = sum(1 for r in history if r.get("date") == today_tag)

        self._update_stat(self._stat_total, len(history))
        self._update_stat(self._stat_today, today_count)

        for widget in self._list_scroll.winfo_children():
            widget.destroy()

        if not history:
            ctk.CTkLabel(self._list_scroll, text="暂无素材历史记录",
                         font=ctk.CTkFont(family=F_UI, size=13),
                         text_color=C_DIM).pack(pady=40)
            return

        for i, record in enumerate(history):
            date_tag = record.get("date", "")
            name = record.get("name", "")
            display_name = f"{date_tag}-{name}" if date_tag else name
            is_today = (date_tag == today_tag)

            row = ctk.CTkFrame(self._list_scroll,
                               fg_color=C_SURFACE if i % 2 == 0 else "transparent",
                               corner_radius=R_SM, height=32)
            row.pack(fill="x", padx=4, pady=1)
            row.pack_propagate(False)

            ctk.CTkLabel(row, text=date_tag, width=80,
                         font=ctk.CTkFont(family=F_UI, size=12,
                                          weight="bold" if is_today else "normal"),
                         text_color=C_PRIMARY if is_today else C_TEXT_2,
                         anchor="w").pack(side="left", padx=(12, 0))

            ctk.CTkLabel(row, text=display_name,
                         font=ctk.CTkFont(family=F_UI, size=12),
                         text_color=C_TEXT if is_today else C_TEXT_2,
                         anchor="w").pack(side="left", padx=(24, 0), fill="x", expand=True)

    def _clear_history(self):
        save_material_history([])
        self.refresh()


# ════════════════════════════════════════════════════════════════════
#  📊 基建记录面板
# ════════════════════════════════════════════════════════════════════
class BuildRecordFrame(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=C_BG, **kwargs)
        self._build_ui()

    def _build_ui(self):
        header_row = ctk.CTkFrame(self, fg_color="transparent")
        header_row.pack(fill="x", pady=(0, 16))
        ctk.CTkLabel(header_row, text="📊 每日基建记录",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(side="left")
        self._refresh_btn = ctk.CTkButton(
            header_row, text="🔄 刷新", width=70, height=28,
            font=ctk.CTkFont(family=F_UI, size=11),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY, border_width=1, border_color=C_BORDER,
            corner_radius=R_SM, command=self.refresh)
        self._refresh_btn.pack(side="right")

        stats_row = ctk.CTkFrame(self, fg_color="transparent")
        stats_row.pack(fill="x", pady=(0, 16))
        stats_row.columnconfigure((0, 1, 2), weight=1, uniform="stat")

        self._stat_today_acc = self._make_stat_card(stats_row, "今日账户", "0", C_PRIMARY)
        self._stat_today_acc.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self._stat_today_proj = self._make_stat_card(stats_row, "今日项目", "0", C_GREEN)
        self._stat_today_proj.grid(row=0, column=1, sticky="nsew", padx=6)
        self._stat_total_proj = self._make_stat_card(stats_row, "累计项目", "0", C_ORANGE)
        self._stat_total_proj.grid(row=0, column=2, sticky="nsew", padx=(6, 0))

        list_frame = ctk.CTkFrame(self, fg_color=C_CARD, corner_radius=R_LG,
                                  border_width=1, border_color=C_BORDER)
        list_frame.pack(fill="both", expand=True)

        list_header = ctk.CTkFrame(list_frame, fg_color="transparent")
        list_header.pack(fill="x", padx=16, pady=(12, 4))
        for col_text, col_w, anchor in [("日期", 120, "w"), ("账户数", 80, "center"), ("项目数", 80, "center")]:
            ctk.CTkLabel(list_header, text=col_text, width=col_w,
                         font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                         text_color=C_DIM, anchor=anchor).pack(side="left", padx=(0, 24))

        sep = ctk.CTkFrame(list_frame, fg_color=C_BORDER, height=1)
        sep.pack(fill="x", padx=16, pady=(0, 4))

        self._list_scroll = ctk.CTkScrollableFrame(
            list_frame, fg_color="transparent", corner_radius=0,
            scrollbar_button_color=C_BORDER, scrollbar_button_hover_color=C_DIM)
        self._list_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.refresh()

    def _make_stat_card(self, parent, label, value, color):
        card = ctk.CTkFrame(parent, fg_color=C_CARD, corner_radius=R_LG,
                            border_width=1, border_color=C_BORDER, height=80)
        card.pack_propagate(False)
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.place(relx=0.5, rely=0.5, anchor="center")
        val_lbl = ctk.CTkLabel(inner, text=value,
                               font=ctk.CTkFont(family=F_UI, size=26, weight="bold"),
                               text_color=color)
        val_lbl.pack()
        ctk.CTkLabel(inner, text=label,
                     font=ctk.CTkFont(family=F_UI, size=11),
                     text_color=C_DIM).pack(pady=(2, 0))
        card._val_label = val_lbl
        return card

    def _update_stat(self, card, value):
        card._val_label.configure(text=str(value))

    def refresh(self):
        records = load_build_records()
        today = datetime.now().strftime("%Y-%m-%d")
        today_data = records.get(today, {})
        today_acc = today_data.get("accounts", 0)
        today_proj = today_data.get("projects", 0)
        total_proj = sum(d.get("projects", 0) for d in records.values())

        self._update_stat(self._stat_today_acc, today_acc)
        self._update_stat(self._stat_today_proj, today_proj)
        self._update_stat(self._stat_total_proj, total_proj)

        for widget in self._list_scroll.winfo_children():
            widget.destroy()

        sorted_dates = sorted(records.keys(), reverse=True)
        if not sorted_dates:
            ctk.CTkLabel(self._list_scroll, text="暂无基建记录",
                         font=ctk.CTkFont(family=F_UI, size=13),
                         text_color=C_DIM).pack(pady=40)
            return

        for i, date_str in enumerate(sorted_dates):
            day = records[date_str]
            acc = day.get("accounts", 0)
            proj = day.get("projects", 0)
            is_today = (date_str == today)

            row = ctk.CTkFrame(self._list_scroll, fg_color=C_SURFACE if i % 2 == 0 else "transparent",
                               corner_radius=R_SM, height=36)
            row.pack(fill="x", padx=4, pady=1)
            row.pack_propagate(False)

            date_display = f"📅 {date_str}"
            if is_today:
                date_display = f"📅 {date_str}  (今天)"
            ctk.CTkLabel(row, text=date_display, width=160,
                         font=ctk.CTkFont(family=F_UI, size=12,
                                          weight="bold" if is_today else "normal"),
                         text_color=C_PRIMARY if is_today else C_TEXT,
                         anchor="w").pack(side="left", padx=(12, 0))

            ctk.CTkLabel(row, text=str(acc), width=80,
                         font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                         text_color=C_PRIMARY, anchor="center").pack(side="left", padx=(0, 24))

            ctk.CTkLabel(row, text=str(proj), width=80,
                         font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                         text_color=C_GREEN, anchor="center").pack(side="left")


# ════════════════════════════════════════════════════════════════════
#  ⚙ 设置窗口（一站式：基础参数 + 账号组 + 剧 + 链接 + 素材 ID）
# ════════════════════════════════════════════════════════════════════
class SettingsFrame(ctk.CTkFrame):
    BASE_FIELD_LABELS = {
        "strategy":            "投放策略",
        "material_account_id": "素材账号 ID",
        "audience_keyword":    "受众关键词",
        "monitor_btn_text":    "监控按钮文案",
        "name_prefix":         "命名前缀",
        "wait_scale":          "等待倍率",
    }

    def __init__(self, master, on_close=None, on_profile_added=None):
        super().__init__(master, fg_color=C_BG)
        self.on_close = on_close
        self.on_profile_added = on_profile_added

        self.cfg = load_config()
        self.profile_keys = list(ALL_PROFILES.keys())
        self.current_key = self.profile_keys[0]

        self._render_token = 0
        self._card1 = None
        self._card2 = None
        self._groups_holder = None
        self._groups_head_btn = None
        self._card1_grid = None
        self._raw_holder = None
        self._raw_textbox = None
        self._raw_mode = False
        self._toggle_btn = None
        self.base_vars = {}

        self._build_ui()
        self.after_idle(lambda: self._render_profile(self.current_key))

    def reload_config(self):
        self.cfg = load_config()
        self._render_profile(self.current_key)

    # ─────────────── UI 主框架 ───────────────
    def _build_ui(self):
        top = ctk.CTkFrame(self, fg_color=C_CARD, height=64,
                           corner_radius=0, border_width=0)
        top.pack(side="top", fill="x")
        top.pack_propagate(False)

        ctk.CTkButton(top, text="←  返回", width=72, height=28,
                      font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, command=self._on_close
                      ).pack(side="left", padx=(24, 12), pady=18)

        ctk.CTkLabel(top, text="⚙  设置中心",
                     font=ctk.CTkFont(family=F_UI, size=18, weight="bold"),
                     text_color=C_TEXT).pack(side="left", padx=(0, 10), pady=14)
        ctk.CTkLabel(top, text="账号 ID · 链接 · 素材 ID 一站式管理",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_DIM).pack(side="left", padx=(0, 20), pady=18)

        # CDP endpoint（公共）
        ctk.CTkLabel(top, text="CDP 端点：",
                     font=ctk.CTkFont(family=F_UI, size=12),
                     text_color=C_TEXT_2).pack(side="left", padx=(20, 4))
        self.var_cdp = tk.StringVar(value=(self.cfg.get("common") or {})
                                          .get("cdp_endpoint", "http://localhost:9222"))
        self.ent_cdp = ctk.CTkEntry(top, width=240, height=28,
                                    textvariable=self.var_cdp,
                                    font=ctk.CTkFont(family=F_MONO, size=11),
                                    fg_color="#ffffff", text_color=C_TEXT,
                                    border_color=C_BORDER, border_width=1,
                                    corner_radius=R_SM)
        self.ent_cdp.pack(side="left", padx=(0, 20))

        # 主体
        body = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        body.pack(side="top", fill="both", expand=True)

        # 左侧导航
        side = ctk.CTkFrame(body, fg_color=C_CARD, width=220,
                            corner_radius=0, border_width=0)
        side.pack(side="left", fill="y")
        side.pack_propagate(False)
        side_scroll = ctk.CTkScrollableFrame(side, fg_color=C_CARD, corner_radius=0)
        side_scroll.pack(fill="both", expand=True)

        ctk.CTkLabel(side_scroll, text="选择投放方向",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", padx=20, pady=(20, 8))

        self.nav_buttons = {}
        self._cat_frames = {}
        self._cat_collapsed = {}

        for cat_name, cat_keys in PROFILE_CATEGORIES:
            cat_header = ctk.CTkButton(
                side_scroll, text=f"▾  {cat_name}", height=30, anchor="w",
                font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                fg_color="transparent", hover_color=C_HOVER,
                text_color=C_TEXT, corner_radius=R_SM, border_width=0,
                command=lambda cn=cat_name: self._toggle_category(cn),
            )
            cat_header.pack(fill="x", padx=12, pady=(6, 2))
            self._cat_frames[cat_name] = {"header": cat_header, "container": None, "keys": cat_keys}
            self._cat_collapsed[cat_name] = False

            container = ctk.CTkFrame(side_scroll, fg_color="transparent")
            container.pack(fill="x", padx=(12, 0))
            self._cat_frames[cat_name]["container"] = container

            if cat_keys:
                for key in cat_keys:
                    btn = ctk.CTkButton(
                        container, text="  " + key, height=34, anchor="w",
                        font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                        fg_color=C_CARD, hover_color=C_HOVER,
                        text_color=C_TEXT_2, corner_radius=R_SM,
                        border_width=0,
                        command=lambda k=key: self._switch_profile(k),
                    )
                    btn.pack(fill="x", padx=8, pady=2)
                    self.nav_buttons[key] = btn
            else:
                ctk.CTkLabel(container, text="  暂无方向",
                             font=ctk.CTkFont(family=F_UI, size=11),
                             text_color=C_DIM).pack(anchor="w", padx=16, pady=4)

        # 底部操作按钮（在 side 下方）
        ctk.CTkFrame(side_scroll, fg_color=C_BORDER, height=1).pack(fill="x", padx=12, pady=(12, 6))

        ctk.CTkButton(side_scroll, text="💾  保存全部", height=36,
                      font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                      fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                      text_color="#ffffff", corner_radius=R_SM,
                      command=self._save).pack(fill="x", padx=12, pady=6)

        ctk.CTkButton(side_scroll, text="↺  恢复当前默认", height=34,
                      font=ctk.CTkFont(family=F_UI, size=12),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM,
                      command=self._restore_defaults).pack(fill="x", padx=12, pady=4)

        ctk.CTkFrame(side_scroll, fg_color=C_BORDER, height=1).pack(fill="x", padx=12, pady=(10, 8))
        ctk.CTkLabel(side_scroll, text="单本工具",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", padx=20, pady=(0, 6))

        self._tool_nav_btn = ctk.CTkButton(
            side_scroll, text="🔗  链接分配", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_tool_view)
        self._tool_nav_btn.pack(fill="x", padx=12, pady=4)

        self._promotion_nav_btn = ctk.CTkButton(
            side_scroll, text="🚀  推广链生成", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_promotion_tool_view)
        self._promotion_nav_btn.pack(fill="x", padx=12, pady=3)

        self._promotion_split_nav_btn = ctk.CTkButton(
            side_scroll, text="🧩  推广链拆分", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_promotion_split_tool_view)
        self._promotion_split_nav_btn.pack(fill="x", padx=12, pady=3)

        self._search_push_nav_btn = ctk.CTkButton(
            side_scroll, text="🔎  搜索剧名素材推送", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_search_push_tool_view)
        self._search_push_nav_btn.pack(fill="x", padx=12, pady=3)

        ctk.CTkFrame(side_scroll, fg_color=C_BORDER, height=1).pack(fill="x", padx=12, pady=(10, 8))
        ctk.CTkLabel(side_scroll, text="激励工具",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", padx=20, pady=(0, 6))

        self._inc_promo_nav_btn = ctk.CTkButton(
            side_scroll, text="🚀  激励推广链生成", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_inc_promo_view)
        self._inc_promo_nav_btn.pack(fill="x", padx=12, pady=3)

        self._inc_push_nav_btn = ctk.CTkButton(
            side_scroll, text="📦  激励素材推送", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_inc_push_view)
        self._inc_push_nav_btn.pack(fill="x", padx=12, pady=3)

        self._inc_link_nav_btn = ctk.CTkButton(
            side_scroll, text="🔗  激励链接分配", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_inc_link_view)
        self._inc_link_nav_btn.pack(fill="x", padx=12, pady=3)

        self._inc_split_nav_btn = ctk.CTkButton(
            side_scroll, text="🧩  激励推广链拆分", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_inc_split_view)
        self._inc_split_nav_btn.pack(fill="x", padx=12, pady=3)

        self._mat_history_nav_btn = ctk.CTkButton(
            side_scroll, text="📋  素材历史记录", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_mat_history_view)
        self._mat_history_nav_btn.pack(fill="x", padx=12, pady=3)

        ctk.CTkFrame(side_scroll, fg_color=C_BORDER, height=1).pack(fill="x", padx=12, pady=(10, 8))
        ctk.CTkLabel(side_scroll, text="数据统计",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", padx=20, pady=(0, 6))

        self._build_record_nav_btn = ctk.CTkButton(
            side_scroll, text="📊  基建记录", height=34,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_PRIMARY,
            border_width=1, border_color=C_PRIMARY,
            corner_radius=R_SM,
            command=self._show_build_record_view)
        self._build_record_nav_btn.pack(fill="x", padx=12, pady=3)

        self._right_wrap = ctk.CTkFrame(body, fg_color=C_BG, corner_radius=0)
        self._right_wrap.pack(side="left", fill="both", expand=True)

        # 视图 1：设置表单
        self._settings_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
        self._settings_view.pack(fill="both", expand=True)
        self.scroll = ctk.CTkScrollableFrame(self._settings_view, fg_color=C_BG,
                                             corner_radius=0)
        self.scroll.pack(fill="both", expand=True, padx=20, pady=16)

        # 视图 2：单本工具（延迟创建）
        self._tool_view = None
        self._tool_frame = None
        self._promotion_tool_view = None
        self._promotion_tool_frame = None
        self._promotion_split_tool_view = None
        self._promotion_split_tool_frame = None
        self._search_push_tool_view = None
        self._search_push_tool_frame = None
        self._inc_promo_view = None
        self._inc_promo_frame = None
        self._inc_push_view = None
        self._inc_push_frame = None
        self._inc_link_view = None
        self._inc_link_frame = None
        self._inc_split_view = None
        self._inc_split_frame = None
        self._mat_history_view = None
        self._mat_history_frame = None
        self._build_record_view = None
        self._build_record_frame = None
        self._current_view = "settings"

    # ─────────────── 分类折叠/展开 ───────────────
    def _toggle_category(self, cat_name):
        info = self._cat_frames.get(cat_name)
        if not info:
            return
        collapsed = self._cat_collapsed.get(cat_name, False)
        container = info["container"]
        if collapsed:
            container.pack(fill="x", padx=(12, 0), after=info["header"])
            info["header"].configure(text=f"▾  {cat_name}")
            self._cat_collapsed[cat_name] = False
        else:
            container.pack_forget()
            info["header"].configure(text=f"▸  {cat_name}")
            self._cat_collapsed[cat_name] = True

    # ─────────────── 渲染指定方向 ───────────────
    def _switch_profile(self, key):
        self._collect_current()  # 离开前先把当前表单值收回 self.cfg
        # 如果当前在工具视图，先切回设置视图
        if self._current_view != "settings":
            self._show_settings_view()
        self.current_key = key
        self._render_profile(key)

    def _highlight_nav(self, key):
        for k, btn in self.nav_buttons.items():
            if k == key:
                btn.configure(fg_color=C_SURFACE, text_color=C_TEXT)
            else:
                btn.configure(fg_color=C_CARD, text_color=C_TEXT_2)
        # 工具按钮取消高亮
        try:
            self._tool_nav_btn.configure(fg_color=C_SURFACE, text_color=C_PRIMARY)
            self._promotion_nav_btn.configure(fg_color=C_SURFACE, text_color=C_PRIMARY)
            self._promotion_split_nav_btn.configure(fg_color=C_SURFACE, text_color=C_PRIMARY)
            self._search_push_nav_btn.configure(fg_color=C_SURFACE, text_color=C_PRIMARY)
        except Exception:
            pass

    def _render_profile(self, key):
        self._highlight_nav(key)

        # 优化③：第一次渲染时构造 card1/card2 骨架；之后切换 PROFILE 仅清理内部内容
        prof = ((self.cfg.get("profiles") or {}).get(key)) or _profile_defaults(key)
        self.cfg.setdefault("profiles", {})[key] = prof

        first_time = self._card1 is None

        if first_time:
            # ── 卡片 1：基础参数 ──（仅创建一次）
            self._card1 = self._card(self.scroll, f"基础参数 · {key}")
            grid = ctk.CTkFrame(self._card1, fg_color=C_CARD)
            grid.pack(fill="x", padx=18, pady=(4, 14))
            self._card1_grid = grid

            # ── 卡片 2：账号组 + 剧（层级） ──（仅创建一次）
            self._card2 = self._card(self.scroll, "账号组 / 剧 / 链接 / 素材 ID（层级配置）")
            head = ctk.CTkFrame(self._card2, fg_color=C_CARD)
            head.pack(fill="x", padx=18, pady=(0, 8))
            ctk.CTkLabel(head, text="每个「账号组」可包含多个账号 ID 和多部「剧」；每部剧含 3 条链接 + 多个素材 ID。",
                         font=ctk.CTkFont(family=F_UI, size=11),
                         text_color=C_DIM).pack(side="left")
            self._groups_head_btn = ctk.CTkButton(
                head, text="＋ 添加账号组", height=30, width=120,
                font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                fg_color=C_PRIMARY, hover_color=C_PRIMARY_H,
                text_color="#ffffff", corner_radius=R_SM,
                command=self._add_group)
            self._groups_head_btn.pack(side="right")

            # 切换按钮：整齐视图 ↔ 原始 ids
            self._toggle_btn = ctk.CTkButton(
                head, text="📝 原始设置", height=30, width=100,
                font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                fg_color=C_SURFACE, hover_color=C_HOVER,
                text_color=C_TEXT_2, border_width=1, border_color=C_BORDER,
                corner_radius=R_SM,
                command=self._toggle_raw_mode)
            self._toggle_btn.pack(side="right", padx=(0, 8))

            self._groups_holder = ctk.CTkFrame(self._card2, fg_color=C_CARD)
            self._groups_holder.pack(fill="both", expand=True, padx=18, pady=(0, 14))

            # 原始 ids 文本编辑框（默认隐藏）
            self._raw_holder = ctk.CTkFrame(self._card2, fg_color=C_CARD)
            ctk.CTkLabel(self._raw_holder,
                         text="原始 ids.txt 格式（账号ID 每行一个 → 空行 → 剧名/链接/素材 → === 分隔多组）",
                         font=ctk.CTkFont(family=F_UI, size=11),
                         text_color=C_DIM).pack(anchor="w", pady=(0, 4))
            self._raw_textbox = ctk.CTkTextbox(
                self._raw_holder, height=400,
                font=ctk.CTkFont(family=F_MONO, size=12),
                fg_color="#ffffff", text_color=C_TEXT,
                border_color=C_BORDER, border_width=1,
                corner_radius=R_SM, wrap="none")
            self._raw_textbox.pack(fill="both", expand=True)
            # _raw_holder 不 pack，默认隐藏
        else:
            try:
                title_lbl = self._card1.winfo_children()[0]
                if isinstance(title_lbl, ctk.CTkLabel):
                    title_lbl.configure(text=f"基础参数 · {key}")
            except Exception:
                pass
            for w in self._groups_holder.winfo_children():
                w.destroy()
            if self._raw_mode:
                self._raw_mode = False
                self._raw_holder.pack_forget()
                self._groups_holder.pack(fill="both", expand=True, padx=18, pady=(0, 14))
                self._groups_head_btn.pack(side="right")
                self._toggle_btn.configure(text="📝 原始设置")

        if not self.base_vars:
            self.base_vars = {}
            grid = self._card1_grid
            for i, field in enumerate(PROFILE_EDITABLE_FIELDS):
                row, col = divmod(i, 2)
                cell = ctk.CTkFrame(grid, fg_color=C_CARD)
                cell.grid(row=row, column=col, sticky="ew", padx=8, pady=6)
                grid.grid_columnconfigure(col, weight=1)
                ctk.CTkLabel(cell, text=self.BASE_FIELD_LABELS.get(field, field),
                             font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                             text_color=C_TEXT_2).pack(anchor="w")
                var = tk.StringVar(value=str(prof.get(field, "")))
                ent = ctk.CTkEntry(cell, height=32, textvariable=var,
                                   font=ctk.CTkFont(family=F_UI, size=12),
                                   fg_color="#ffffff", text_color=C_TEXT,
                                   border_color=C_BORDER, border_width=1,
                                   corner_radius=R_SM)
                ent.pack(fill="x", pady=(4, 0))
                self.base_vars[field] = var
        else:
            for field in PROFILE_EDITABLE_FIELDS:
                self.base_vars[field].set(str(prof.get(field, "")))

        # 优化④：分批渲染 groups header（每帧渲染 5 个，避免一次性创建过多 widget 卡顿）
        self.group_widgets = []
        is_incentive = ALL_PROFILES.get(key, {}).get("build_mode") == "incentive"
        if is_incentive:
            self._groups_head_btn.pack_forget()
            self._toggle_btn.pack_forget()
            try:
                card2_title = self._card2.winfo_children()[0]
                if isinstance(card2_title, ctk.CTkLabel):
                    card2_title.configure(text="账号组 / 监测链接（激励配置）")
            except Exception:
                pass
        else:
            self._groups_head_btn.pack(side="right")
            self._toggle_btn.pack(side="right", padx=(0, 8))
            try:
                card2_title = self._card2.winfo_children()[0]
                if isinstance(card2_title, ctk.CTkLabel):
                    card2_title.configure(text="账号组 / 剧 / 链接 / 素材 ID（层级配置）")
            except Exception:
                pass
        if not prof.get("groups"):
            prof["groups"] = [_empty_group()]

        groups = prof["groups"]
        self._render_token += 1
        token = self._render_token
        BATCH = 5
        _render_fn = self._render_incentive_group if is_incentive else self._render_group

        first_batch = min(BATCH, len(groups))
        for i in range(first_batch):
            _render_fn(i, groups[i])

        def _render_rest(idx):
            if token != self._render_token:
                return
            if idx >= len(groups):
                return
            end = min(idx + BATCH, len(groups))
            try:
                for i in range(idx, end):
                    _render_fn(i, groups[i])
            except Exception:
                return
            self.after(1, lambda: _render_rest(end))

        if len(groups) > first_batch:
            self.after(1, lambda: _render_rest(first_batch))

    def _render_incentive_group(self, gi, group):
        gframe = ctk.CTkFrame(self._groups_holder, fg_color=C_SURFACE,
                              corner_radius=R_MD, border_width=1, border_color=C_BORDER)
        gframe.pack(fill="x", pady=8)

        ghead = ctk.CTkFrame(gframe, fg_color=C_SURFACE)
        ghead.pack(fill="x", padx=14, pady=(10, 4))

        acc_ids = group.get("account_ids") or []
        gname = group.get("group_name", f"组{gi+1}")
        summary = f"📦  {gname}  ·  {len(acc_ids)} 个账号"

        ctk.CTkLabel(ghead, text=summary,
                     font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                     text_color=C_TEXT).pack(side="left")

        content_frame = ctk.CTkFrame(gframe, fg_color=C_SURFACE)
        content_frame._expanded = False

        expand_btn = ctk.CTkButton(ghead, text="▶ 展开", height=26, width=72,
                      font=ctk.CTkFont(family=F_UI, size=11),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT_2, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM,
                      command=lambda cf=content_frame, g=group, gf=gframe: self._toggle_incentive_group(cf, gf, g))
        expand_btn.pack(side="right", padx=(0, 6))
        content_frame._expand_btn = expand_btn

        self.group_widgets.append({
            "group": group, "frame": gframe,
            "acc_box": None, "drama_widgets": [],
            "dramas_holder": None, "content_frame": content_frame,
        })

    def _toggle_incentive_group(self, content_frame, gframe, group):
        if content_frame._expanded:
            content_frame.pack_forget()
            content_frame._expand_btn.configure(text="▶ 展开")
            content_frame._expanded = False
        else:
            if not content_frame.winfo_children():
                self._build_incentive_group_content(content_frame, group)
            content_frame.pack(fill="x", padx=14, pady=(0, 12))
            content_frame._expand_btn.configure(text="▼ 收起")
            content_frame._expanded = True

    def _build_incentive_group_content(self, content_frame, group):
        gw = None
        for w in self.group_widgets:
            if w.get("content_frame") is content_frame:
                gw = w
                break
        if gw is None:
            return

        ctk.CTkLabel(content_frame, text="组名",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(6, 2))
        gname_var = tk.StringVar(value=group.get("group_name", ""))
        ctk.CTkEntry(content_frame, height=32, textvariable=gname_var,
                     font=ctk.CTkFont(family=F_UI, size=12),
                     fg_color="#ffffff", text_color=C_TEXT,
                     border_color=C_BORDER, border_width=1,
                     corner_radius=R_SM).pack(fill="x", pady=(0, 8))
        gw["gname_var"] = gname_var

        ctk.CTkLabel(content_frame, text="账号 ID（每行一个）",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(0, 2))
        acc_box = ctk.CTkTextbox(content_frame, height=70,
                                 font=ctk.CTkFont(family=F_MONO, size=12),
                                 fg_color="#ffffff", text_color=C_TEXT,
                                 border_color=C_BORDER, border_width=1,
                                 corner_radius=R_SM, wrap="none")
        acc_box.pack(fill="x", pady=(0, 10))
        acc_box.insert("1.0", "\n".join(group.get("account_ids") or []))
        gw["acc_box"] = acc_box

        link_vars = {}
        for label, fkey, ph in [
            ("点击监测链接 (click_url)", "click_url", "https://...action_type=0..."),
            ("展示监测链接 (show_url)", "show_url", "https://...action_type=1..."),
            ("有效播放监测链接 (play_url)", "play_url", "https://...action_type=6..."),
        ]:
            ctk.CTkLabel(content_frame, text=label,
                         font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                         text_color=C_TEXT_2).pack(anchor="w", pady=(0, 2))
            var = tk.StringVar(value=group.get(fkey, ""))
            ctk.CTkEntry(content_frame, height=32, textvariable=var,
                         font=ctk.CTkFont(family=F_MONO, size=11),
                         fg_color="#ffffff", text_color=C_TEXT,
                         border_color=C_BORDER, border_width=1,
                         corner_radius=R_SM).pack(fill="x", pady=(0, 8))
            link_vars[fkey] = var
        gw["link_vars"] = link_vars

    # ─────────────── group 渲染 ───────────────
    def _render_group(self, gi, group):
        gframe = ctk.CTkFrame(self._groups_holder, fg_color=C_SURFACE,
                              corner_radius=R_MD, border_width=1, border_color=C_BORDER)
        gframe.pack(fill="x", pady=8)

        ghead = ctk.CTkFrame(gframe, fg_color=C_SURFACE)
        ghead.pack(fill="x", padx=14, pady=(10, 4))

        acc_ids = group.get("account_ids") or []
        dramas = group.get("dramas") or []
        summary = f"📦  账号组 #{gi + 1}  ·  {len(acc_ids)} 个账号  ·  {len(dramas)} 部剧"

        title_lbl = ctk.CTkLabel(ghead, text=summary,
                     font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                     text_color=C_TEXT)
        title_lbl.pack(side="left")

        ctk.CTkButton(ghead, text="🗑 删除该组", height=26, width=86,
                      font=ctk.CTkFont(family=F_UI, size=11),
                      fg_color="transparent", hover_color="#fee2e2",
                      text_color=C_RED, border_width=1, border_color="#fecaca",
                      corner_radius=R_SM,
                      command=lambda g=group: self._remove_group(g)).pack(side="right")

        content_frame = ctk.CTkFrame(gframe, fg_color=C_SURFACE)
        content_frame._expanded = False

        expand_btn = ctk.CTkButton(ghead, text="▶ 展开", height=26, width=72,
                      font=ctk.CTkFont(family=F_UI, size=11),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_TEXT_2, border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM,
                      command=lambda cf=content_frame, eb=None, g=group, gf=gframe: self._toggle_group(cf, gf, g))
        expand_btn.pack(side="right", padx=(0, 6))
        content_frame._expand_btn = expand_btn

        self.group_widgets.append({
            "group": group, "frame": gframe,
            "acc_box": None, "drama_widgets": [],
            "dramas_holder": None, "content_frame": content_frame,
        })

    def _toggle_group(self, content_frame, gframe, group):
        if content_frame._expanded:
            content_frame.pack_forget()
            content_frame._expand_btn.configure(text="▶ 展开")
            content_frame._expanded = False
        else:
            if not content_frame.winfo_children():
                self._build_group_content(content_frame, group)
            content_frame.pack(fill="x", padx=14, pady=(0, 12))
            content_frame._expand_btn.configure(text="▼ 收起")
            content_frame._expanded = True

    def _build_group_content(self, content_frame, group):
        gw = None
        for w in self.group_widgets:
            if w.get("content_frame") is content_frame:
                gw = w
                break
        if gw is None:
            return

        ctk.CTkLabel(content_frame, text="账号 ID（每行一个）",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(anchor="w", pady=(6, 2))
        acc_box = ctk.CTkTextbox(content_frame, height=70,
                                 font=ctk.CTkFont(family=F_MONO, size=12),
                                 fg_color="#ffffff", text_color=C_TEXT,
                                 border_color=C_BORDER, border_width=1,
                                 corner_radius=R_SM, wrap="none")
        acc_box.pack(fill="x", pady=(0, 10))
        acc_box.insert("1.0", "\n".join(group.get("account_ids") or []))
        gw["acc_box"] = acc_box

        dramas_head = ctk.CTkFrame(content_frame, fg_color=C_SURFACE)
        dramas_head.pack(fill="x", pady=(2, 4))
        ctk.CTkLabel(dramas_head, text="剧 / 链接 / 素材",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left")
        ctk.CTkButton(dramas_head, text="＋ 添加剧", height=26, width=86,
                      font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                      fg_color=C_SURFACE, hover_color=C_HOVER,
                      text_color=C_PRIMARY, border_width=1, border_color=C_PRIMARY,
                      corner_radius=R_SM,
                      command=lambda g=group: self._add_drama(g)).pack(side="right")

        dramas_holder = ctk.CTkFrame(content_frame, fg_color=C_SURFACE)
        dramas_holder.pack(fill="x", pady=(0, 4))
        gw["dramas_holder"] = dramas_holder

        if not group.get("dramas"):
            group["dramas"] = [_empty_drama()]
        drama_widgets = []
        for di, d in enumerate(group["dramas"]):
            drama_widgets.append(self._render_drama(dramas_holder, di, d, group))
        gw["drama_widgets"] = drama_widgets

    def _render_drama(self, parent, di, drama, group):
        dframe = ctk.CTkFrame(parent, fg_color=C_CARD,
                              corner_radius=R_SM, border_width=1, border_color=C_BORDER)
        dframe.pack(fill="x", pady=4)

        head = ctk.CTkFrame(dframe, fg_color=C_CARD)
        head.pack(fill="x", padx=10, pady=(8, 2))
        ctk.CTkLabel(head, text=f"🎬  剧 #{di + 1}",
                     font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
                     text_color=C_TEXT).pack(side="left")
        ctk.CTkButton(head, text="✕ 删除", height=22, width=58,
                      font=ctk.CTkFont(family=F_UI, size=10),
                      fg_color="transparent", hover_color="#fee2e2",
                      text_color=C_RED, border_width=0,
                      command=lambda g=group, d=drama: self._remove_drama(g, d)).pack(side="right")

        # 字段：name / click / show / video
        vars_ = {}
        for label, key, ph in [
            ("剧名 / 备注", "name",  "例如：剧A"),
            ("点击链接 click", "click", "https://..."),
            ("展示链接 show",  "show",  "https://..."),
            ("视频链接 video", "video", "https://..."),
        ]:
            row = ctk.CTkFrame(dframe, fg_color=C_CARD)
            row.pack(fill="x", padx=10, pady=2)
            ctk.CTkLabel(row, text=label, width=120, anchor="w",
                         font=ctk.CTkFont(family=F_UI, size=11),
                         text_color=C_TEXT_2).pack(side="left")
            v = tk.StringVar(value=str(drama.get(key, "")))
            ent = ctk.CTkEntry(row, height=28, textvariable=v,
                               placeholder_text=ph,
                               font=ctk.CTkFont(family=F_MONO, size=11),
                               fg_color="#ffffff", text_color=C_TEXT,
                               border_color=C_BORDER, border_width=1,
                               corner_radius=R_SM)
            ent.pack(side="left", fill="x", expand=True)
            vars_[key] = v

        # material_ids
        ctk.CTkLabel(dframe, text="素材 ID（每行一个）",
                     font=ctk.CTkFont(family=F_UI, size=11, weight="bold"),
                     text_color=C_TEXT_2).pack(anchor="w", padx=10, pady=(6, 2))
        mat_box = ctk.CTkTextbox(dframe, height=58,
                                 font=ctk.CTkFont(family=F_MONO, size=11),
                                 fg_color="#ffffff", text_color=C_TEXT,
                                 border_color=C_BORDER, border_width=1,
                                 corner_radius=R_SM, wrap="none")
        mat_box.pack(fill="x", padx=10, pady=(0, 10))
        mat_box.insert("1.0", "\n".join(drama.get("material_ids") or []))

        return {"drama": drama, "frame": dframe, "vars": vars_, "mat_box": mat_box}

    # ─────────────── 增删改 ───────────────
    def _add_group(self):
        prof = self.cfg["profiles"][self.current_key]
        new_group = _empty_group()
        prof.setdefault("groups", []).append(new_group)
        # 增量渲染：只新增一个组的 header，不清空已有组
        gi = len(prof["groups"]) - 1
        self._render_group(gi, new_group)

    def _remove_group(self, group):
        self._collect_current()
        prof = self.cfg["profiles"][self.current_key]
        try:
            prof["groups"].remove(group)
        except ValueError:
            pass
        if not prof["groups"]:
            prof["groups"] = [_empty_group()]
        # 局部移除该组对应的 widget，不整页重渲染
        target = None
        for w in self.group_widgets:
            if w.get("group") is group:
                target = w
                break
        if target is not None:
            try:
                target["frame"].destroy()
            except Exception:
                pass
            self.group_widgets.remove(target)
            # 刷新剩余组的编号
            for idx, gw in enumerate(self.group_widgets):
                try:
                    ghead = gw["frame"].winfo_children()[0]
                    for child in ghead.winfo_children():
                        if isinstance(child, ctk.CTkLabel):
                            g = gw["group"]
                            acc_ids = g.get("account_ids") or []
                            dramas = g.get("dramas") or []
                            child.configure(
                                text=f"📦  账号组 #{idx + 1}  ·  {len(acc_ids)} 个账号  ·  {len(dramas)} 部剧"
                            )
                            break
                except Exception:
                    pass
            # 如果删完只剩一个空组，需要确保它被渲染出来
            if not self.group_widgets and prof["groups"]:
                self._render_group(0, prof["groups"][0])
        else:
            self._render_profile(self.current_key)

    def _add_drama(self, group):
        group.setdefault("dramas", []).append(_empty_drama())
        # 局部刷新：只重建当前组的内容区域，保持其他组折叠状态不变
        self._refresh_group_content(group)

    def _remove_drama(self, group, drama):
        self._collect_current()
        try:
            group["dramas"].remove(drama)
        except ValueError:
            pass
        if not group["dramas"]:
            group["dramas"] = [_empty_drama()]
        # 局部刷新：只重建当前组的内容区域
        self._refresh_group_content(group)

    def _refresh_group_content(self, group):
        """局部刷新指定组：只重建该组的内容区域，避免整页重渲染。"""
        gw = None
        for w in self.group_widgets:
            if w.get("group") is group:
                gw = w
                break
        if gw is None:
            # 找不到对应 widget，回退到整页渲染
            self._render_profile(self.current_key)
            return
        content_frame = gw.get("content_frame")
        if content_frame is None:
            self._render_profile(self.current_key)
            return
        # 清空旧内容
        for w in content_frame.winfo_children():
            w.destroy()
        gw["acc_box"] = None
        gw["dramas_holder"] = None
        gw["drama_widgets"] = []
        # 同步刷新 header 摘要（账号/剧数量）
        try:
            ghead = gw["frame"].winfo_children()[0]
            for child in ghead.winfo_children():
                if isinstance(child, ctk.CTkLabel):
                    acc_ids = group.get("account_ids") or []
                    dramas = group.get("dramas") or []
                    gi = self.group_widgets.index(gw)
                    child.configure(
                        text=f"📦  账号组 #{gi + 1}  ·  {len(acc_ids)} 个账号  ·  {len(dramas)} 部剧"
                    )
                    break
        except Exception:
            pass
        # 重建内容并保持展开
        self._build_group_content(content_frame, group)
        if not content_frame._expanded:
            content_frame.pack(fill="x", padx=14, pady=(0, 12))
            content_frame._expand_btn.configure(text="▼ 收起")
            content_frame._expanded = True

    # ─────────────── 整齐视图 ↔ 原始 ids 切换 ───────────────
    @staticmethod
    def _groups_to_raw_text(groups):
        """将 config groups 列表序列化为 ids.txt 原始文本格式。"""
        parts = []
        for g in groups:
            lines = []
            for aid in (g.get("account_ids") or []):
                aid = str(aid).strip()
                if aid:
                    lines.append(aid)
            lines.append("")  # 空行分隔 ID 和剧
            for d in (g.get("dramas") or []):
                name = (d.get("name") or "").strip()
                if not name:
                    continue
                lines.append(name)
                for lk in ("click", "show", "video"):
                    lines.append((d.get(lk) or "").strip())
                mats = [str(m).strip() for m in (d.get("material_ids") or []) if str(m).strip()]
                if mats:
                    lines.append(" ".join(mats))
                lines.append("")  # 剧之间空行
            parts.append("\n".join(lines).strip())
        return "\n===\n".join(parts)

    @staticmethod
    def _raw_text_to_groups(raw_text):
        """将原始 ids.txt 文本解析为 config groups 列表。"""
        raw_text = "\n".join(
            line for line in raw_text.splitlines()
            if not line.lstrip().startswith("#")
        )
        chunks = re.split(r"(?:═{3,}[^\n]*═{3,}|={3,})", raw_text.strip())
        _silent = logging.getLogger("_silent_raw")
        _silent.addHandler(logging.NullHandler())
        _silent.propagate = False
        groups = []
        for chunk in chunks:
            chunk = chunk.strip()
            if not chunk:
                continue
            ids, dramas = _parse_single_group(chunk, _silent)
            if ids or dramas:
                groups.append({
                    "account_ids": [str(x) for x in ids],
                    "dramas": [
                        {
                            "name": d.get("name", ""),
                            "click": d.get("click", ""),
                            "show": d.get("show", ""),
                            "video": d.get("video", ""),
                            "material_ids": list(d.get("material_ids", [])),
                        }
                        for d in dramas
                    ],
                })
        return groups if groups else [_empty_group()]

    def _toggle_raw_mode(self):
        """切换整齐视图 ↔ 原始 ids 文本编辑框。"""
        if not self._raw_mode:
            # 整齐 → 原始：先收集当前表单数据，再序列化到文本框
            self._collect_current()
            prof = self.cfg["profiles"].get(self.current_key) or {}
            raw = self._groups_to_raw_text(prof.get("groups") or [_empty_group()])
            self._raw_textbox.delete("1.0", "end")
            self._raw_textbox.insert("1.0", raw)
            # 隐藏整齐视图，显示原始文本
            self._groups_holder.pack_forget()
            self._groups_head_btn.pack_forget()
            self._raw_holder.pack(fill="both", expand=True, padx=18, pady=(0, 14))
            self._toggle_btn.configure(text="📋 整齐视图")
            self._raw_mode = True
        else:
            # 原始 → 整齐：解析文本框内容，写回 cfg，重新渲染
            raw = self._raw_textbox.get("1.0", "end").strip()
            new_groups = self._raw_text_to_groups(raw)
            self.cfg["profiles"][self.current_key]["groups"] = new_groups
            # 隐藏原始文本，显示整齐视图
            self._raw_holder.pack_forget()
            self._groups_holder.pack(fill="both", expand=True, padx=18, pady=(0, 14))
            self._groups_head_btn.pack(side="right")
            self._toggle_btn.configure(text="📝 原始设置")
            self._raw_mode = False
            # 清理并重新渲染 groups
            for w in self._groups_holder.winfo_children():
                w.destroy()
            self.group_widgets = []
            groups = new_groups
            self._render_token += 1
            token = self._render_token
            BATCH = 5
            if groups:
                first_batch = min(BATCH, len(groups))
                for i in range(first_batch):
                    self._render_group(i, groups[i])

                def _render_rest(idx):
                    if token != self._render_token:
                        return
                    if idx >= len(groups):
                        return
                    end = min(idx + BATCH, len(groups))
                    try:
                        for i in range(idx, end):
                            self._render_group(i, groups[i])
                    except Exception:
                        return
                    self.after(1, lambda: _render_rest(end))

                if len(groups) > first_batch:
                    self.after(1, lambda: _render_rest(first_batch))

    # ─────────────── 收集表单 → self.cfg ───────────────
    def _collect_current(self):
        if not hasattr(self, "base_vars"):
            return
        prof = self.cfg["profiles"].get(self.current_key)
        if not prof:
            return
        expected_strategy = ALL_PROFILES[self.current_key]["strategy"]
        form_strategy = self.base_vars["strategy"].get().strip()
        if form_strategy and form_strategy != expected_strategy:
            is_cur_inc = ALL_PROFILES.get(self.current_key, {}).get("build_mode") == "incentive"
            form_looks_inc = "激励" in form_strategy
            if is_cur_inc != form_looks_inc:
                return
        for f in PROFILE_EDITABLE_FIELDS:
            v = self.base_vars[f].get().strip()
            if f == "wait_scale":
                try:
                    prof[f] = float(v) if v else 1.0
                except ValueError:
                    prof[f] = 1.0
            else:
                prof[f] = v
        if self._raw_mode and self._raw_textbox:
            raw = self._raw_textbox.get("1.0", "end").strip()
            prof["groups"] = self._raw_text_to_groups(raw)
        else:
            is_inc = ALL_PROFILES.get(self.current_key, {}).get("build_mode") == "incentive"
            for gw in self.group_widgets:
                if gw["acc_box"] is None:
                    continue
                acc_text = gw["acc_box"].get("1.0", "end").strip()
                account_ids = [x.strip() for x in acc_text.splitlines() if x.strip()]
                if is_inc:
                    link_vars = gw.get("link_vars", {})
                    gname_var = gw.get("gname_var")
                    gw["group"].clear()
                    gw["group"].update({
                        "account_ids": account_ids,
                        "group_name": gname_var.get().strip() if gname_var else "",
                        "click_url": link_vars["click_url"].get().strip() if "click_url" in link_vars else "",
                        "show_url": link_vars["show_url"].get().strip() if "show_url" in link_vars else "",
                        "play_url": link_vars["play_url"].get().strip() if "play_url" in link_vars else "",
                        "dramas": [],
                    })
                else:
                    dramas = []
                    for dw in gw["drama_widgets"]:
                        d_obj = {
                            "name":  dw["vars"]["name"].get().strip(),
                            "click": dw["vars"]["click"].get().strip(),
                            "show":  dw["vars"]["show"].get().strip(),
                            "video": dw["vars"]["video"].get().strip(),
                            "material_ids": [
                                x.strip() for x in
                                dw["mat_box"].get("1.0", "end").strip().splitlines()
                                if x.strip()
                            ],
                        }
                        dramas.append(d_obj)
                    gw["group"].clear()
                    gw["group"].update({"account_ids": account_ids, "dramas": dramas})
            prof["groups"] = [gw["group"] for gw in self.group_widgets]
        # 公共
        self.cfg.setdefault("common", {})["cdp_endpoint"] = self.var_cdp.get().strip() or "http://localhost:9222"

    # ─────────────── 保存 / 恢复 / 关闭 ───────────────
    def _save(self):
        self._collect_current()
        try:
            save_config(self.cfg)
        except Exception as e:
            messagebox.showerror("保存失败", f"无法写入 config.json：\n{e}", parent=self)
            return
        messagebox.showinfo("已保存",
                            f"配置已写入：\n{CONFIG_FILE}\n\n下次启动构建时将自动使用最新设置。",
                            parent=self)

    def _restore_defaults(self):
        if not messagebox.askyesno(
            "恢复默认",
            f"将把方向「{self.current_key}」恢复为内置默认值（清空已录入的账号 / 链接 / 素材）。\n\n继续？",
            parent=self,
        ):
            return
        self.cfg.setdefault("profiles", {})[self.current_key] = _profile_defaults(self.current_key)
        self._render_profile(self.current_key)

    def _on_close(self):
        self._collect_current()
        try:
            save_config(self.cfg)
        except Exception as e:
            messagebox.showerror("保存失败", f"无法写入 config.json：\n{e}", parent=self)
            return
        if callable(self.on_close):
            try:
                self.on_close()
            except Exception:
                pass

    # ─────────────── 视图切换（设置 ↔ 工具）+ 过渡动画 ───────────────
    def _hide_current_view(self):
        current = {
            "settings": self._settings_view,
            "tool": self._tool_view,
            "promotion": self._promotion_tool_view,
            "promotion_split": self._promotion_split_tool_view,
            "search_push": self._search_push_tool_view,
            "inc_promo": self._inc_promo_view,
            "inc_push": self._inc_push_view,
            "inc_link": self._inc_link_view,
            "inc_split": self._inc_split_view,
            "mat_history": self._mat_history_view,
            "build_record": self._build_record_view,
        }.get(self._current_view)
        if current is not None:
            current.pack_forget()

    def _reset_all_tool_btns(self):
        for k, btn in self.nav_buttons.items():
            btn.configure(fg_color=C_CARD, text_color=C_TEXT_2)
        for btn in (self._tool_nav_btn, self._promotion_nav_btn,
                    self._promotion_split_nav_btn, self._search_push_nav_btn,
                    self._inc_promo_nav_btn, self._inc_push_nav_btn,
                    self._inc_link_nav_btn, self._inc_split_nav_btn,
                    self._mat_history_nav_btn,
                    self._build_record_nav_btn):
            btn.configure(fg_color=C_SURFACE, text_color=C_PRIMARY)

    def _show_tool_view(self):
        """切换到链接分配工具视图。"""
        if self._current_view == "tool":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._tool_view is None:
            self._tool_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._tool_frame = JuMingToolFrame(
                self._tool_view,
                profile_keys=self.profile_keys,
                on_add_to_profile=self._on_add_to_profile)
            self._tool_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._tool_view.pack(fill="both", expand=True)
        self._current_view = "tool"
        self._reset_all_tool_btns()
        self._tool_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_promotion_tool_view(self):
        """切换到推广链生成工具视图。"""
        if self._current_view == "promotion":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._promotion_tool_view is None:
            self._promotion_tool_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._promotion_tool_frame = PromotionChainToolFrame(self._promotion_tool_view)
            self._promotion_tool_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._promotion_tool_view.pack(fill="both", expand=True)
        self._current_view = "promotion"
        self._reset_all_tool_btns()
        self._promotion_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_promotion_split_tool_view(self):
        if self._current_view == "promotion_split":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._promotion_split_tool_view is None:
            self._promotion_split_tool_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._promotion_split_tool_frame = PromotionSplitToolFrame(self._promotion_split_tool_view)
            self._promotion_split_tool_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._promotion_split_tool_view.pack(fill="both", expand=True)
        self._current_view = "promotion_split"
        self._reset_all_tool_btns()
        self._promotion_split_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_search_push_tool_view(self):
        """切换到搜索剧名素材推送工具视图。"""
        if self._current_view == "search_push":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._search_push_tool_view is None:
            self._search_push_tool_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._search_push_tool_frame = SearchDramaMaterialPushToolFrame(self._search_push_tool_view)
            self._search_push_tool_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._search_push_tool_view.pack(fill="both", expand=True)
        self._current_view = "search_push"
        self._reset_all_tool_btns()
        self._search_push_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_inc_promo_view(self):
        if self._current_view == "inc_promo":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._inc_promo_view is None:
            self._inc_promo_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._inc_promo_frame = IncentivePromoChainToolFrame(self._inc_promo_view)
            self._inc_promo_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._inc_promo_view.pack(fill="both", expand=True)
        self._current_view = "inc_promo"
        self._reset_all_tool_btns()
        self._inc_promo_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_inc_push_view(self):
        if self._current_view == "inc_push":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._inc_push_view is None:
            self._inc_push_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._inc_push_frame = IncentivePushToolFrame(self._inc_push_view)
            self._inc_push_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._inc_push_view.pack(fill="both", expand=True)
        self._current_view = "inc_push"
        self._reset_all_tool_btns()
        self._inc_push_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_inc_link_view(self):
        if self._current_view == "inc_link":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._inc_link_view is None:
            self._inc_link_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._inc_link_frame = IncentiveLinkToolFrame(
                self._inc_link_view,
                on_add_to_profile=self._on_add_to_profile)
            self._inc_link_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._inc_link_view.pack(fill="both", expand=True)
        self._current_view = "inc_link"
        self._reset_all_tool_btns()
        self._inc_link_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_inc_split_view(self):
        if self._current_view == "inc_split":
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._inc_split_view is None:
            self._inc_split_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._inc_split_frame = IncentivePromotionSplitToolFrame(self._inc_split_view)
            self._inc_split_frame.pack(fill="both", expand=True, padx=20, pady=16)
        self._hide_current_view()
        self._inc_split_view.pack(fill="both", expand=True)
        self._current_view = "inc_split"
        self._reset_all_tool_btns()
        self._inc_split_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_mat_history_view(self):
        if self._current_view == "mat_history":
            if self._mat_history_frame:
                self._mat_history_frame.refresh()
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._mat_history_view is None:
            self._mat_history_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._mat_history_frame = MaterialHistoryFrame(self._mat_history_view)
            self._mat_history_frame.pack(fill="both", expand=True, padx=20, pady=16)
        else:
            self._mat_history_frame.refresh()
        self._hide_current_view()
        self._mat_history_view.pack(fill="both", expand=True)
        self._current_view = "mat_history"
        self._reset_all_tool_btns()
        self._mat_history_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_build_record_view(self):
        if self._current_view == "build_record":
            if self._build_record_frame:
                self._build_record_frame.refresh()
            return
        if self._current_view == "settings":
            self._collect_current()
        if self._build_record_view is None:
            self._build_record_view = ctk.CTkFrame(self._right_wrap, fg_color=C_BG)
            self._build_record_frame = BuildRecordFrame(self._build_record_view)
            self._build_record_frame.pack(fill="both", expand=True, padx=20, pady=16)
        else:
            self._build_record_frame.refresh()
        self._hide_current_view()
        self._build_record_view.pack(fill="both", expand=True)
        self._current_view = "build_record"
        self._reset_all_tool_btns()
        self._build_record_nav_btn.configure(fg_color=C_PRIMARY, text_color="#ffffff")

    def _show_settings_view(self):
        """切换回设置视图。"""
        if self._current_view == "settings":
            return
        self._hide_current_view()
        self._settings_view.pack(fill="both", expand=True)
        self._current_view = "settings"
        self._reset_all_tool_btns()
        self._highlight_nav(self.current_key)

    def _fade_switch(self, old_frame, new_frame, steps=6, interval=18):
        """淡出旧视图 → 淡入新视图的过渡动画。"""
        # 使用 place 覆盖实现淡入效果
        def _fade_out(step):
            if step <= 0:
                old_frame.pack_forget()
                _fade_in(0)
                return
            try:
                # 通过逐步降低透明度模拟淡出（CTk 不支持真透明，用 place + 覆盖层）
                old_frame.configure(fg_color=C_BG)
            except Exception:
                pass
            self.after(interval, lambda: _fade_out(step - 1))

        def _fade_in(step):
            if step == 0:
                new_frame.pack(fill="both", expand=True)
                try:
                    new_frame.configure(fg_color=C_BG)
                except Exception:
                    pass
            if step >= steps:
                return
            self.after(interval, lambda: _fade_in(step + 1))

        _fade_out(steps)

    def _on_add_to_profile(self, profile_key, new_groups):
        """工具视图的「添加到」回调：写入 cfg 并保存。"""
        self._collect_current()
        prof = self.cfg.setdefault("profiles", {}).setdefault(
            profile_key, _profile_defaults(profile_key))
        existing = prof.get("groups") or []
        if len(existing) == 1:
            g = existing[0]
            if not any(g.get("account_ids") or []) and all(
                not d.get("name") for d in (g.get("dramas") or [])):
                existing = []
        prof["groups"] = new_groups
        save_config(self.cfg)
        if profile_key == self.current_key:
            self._render_profile(profile_key)
        if callable(self.on_profile_added):
            self.on_profile_added(profile_key, len(new_groups))

    def _open_html_tool(self):
        """兼容旧调用 — 切换到工具视图。"""
        self._show_tool_view()

    # ─────────────── 卡片容器 ───────────────
    def _card(self, parent, title):
        wrap = ctk.CTkFrame(parent, fg_color=C_CARD,
                            corner_radius=R_LG, border_width=1, border_color=C_BORDER)
        wrap.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(wrap, text=title,
                     font=ctk.CTkFont(family=F_UI, size=14, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w", padx=18, pady=(14, 8))
        return wrap


class BuildApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("红果搭建工具")
        self.geometry("960x680")
        self.minsize(820, 560)
        self.configure(fg_color=C_BG)
        self._stop_event = threading.Event()
        self._thread = None
        self._log_queue = queue.Queue()
        self._running = False
        self._settings_frame = None
        self._main_frame = None
        self._build_ui()
        self._poll_log()

    # ── 界面构建 ──────────────────────────────────────────────
    def _build_ui(self):
        self._main_frame = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        self._main_frame.pack(fill="both", expand=True)

        # ===== 顶部导航栏 =====
        header = ctk.CTkFrame(self._main_frame, fg_color=C_CARD, corner_radius=0, height=64,
                              border_width=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        # 底部 1px 分割线
        sep = ctk.CTkFrame(header, fg_color=C_BORDER, height=1, corner_radius=0)
        sep.place(relx=0, rely=1.0, relwidth=1, anchor="sw")

        # 左侧：Logo + 标题
        left_box = ctk.CTkFrame(header, fg_color="transparent")
        left_box.pack(side="left", padx=24, pady=10)

        logo = ctk.CTkLabel(left_box, text="🍎", font=ctk.CTkFont(size=26))
        logo.pack(side="left", padx=(0, 10))

        title_box = ctk.CTkFrame(left_box, fg_color="transparent")
        title_box.pack(side="left")
        ctk.CTkLabel(title_box, text="红果搭建工具",
                     font=ctk.CTkFont(family=F_UI, size=16, weight="bold"),
                     text_color=C_TEXT).pack(anchor="w")
        ctk.CTkLabel(title_box, text="Hongguo Build Console",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(anchor="w")

        # 右侧：⚙ 设置按钮 + 状态徽章
        self.btn_settings = ctk.CTkButton(
            header, text="⚙  设置", width=78, height=28,
            font=ctk.CTkFont(family=F_UI, size=12, weight="bold"),
            fg_color=C_SURFACE, hover_color=C_HOVER,
            text_color=C_TEXT, border_width=1, border_color=C_BORDER,
            corner_radius=R_SM, command=self._open_settings,
        )
        self.btn_settings.pack(side="right", padx=(0, 24))

        self.status_badge = ctk.CTkFrame(header, fg_color="#ecfdf5",
                                         corner_radius=999, height=28,
                                         border_width=1, border_color="#a7f3d0")
        self.status_badge.pack(side="right", padx=(0, 10))
        self.status_dot = ctk.CTkLabel(self.status_badge, text="●",
                                       font=ctk.CTkFont(size=12),
                                       text_color=C_GREEN)
        self.status_dot.pack(side="left", padx=(12, 6), pady=4)
        self.status_label = ctk.CTkLabel(self.status_badge, text="就绪",
                                         font=ctk.CTkFont(family=F_UI, size=12,
                                                          weight="bold"),
                                         text_color="#065f46")
        self.status_label.pack(side="left", padx=(0, 14), pady=4)

        # ===== 主体内容区 =====
        body = ctk.CTkFrame(self._main_frame, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=20, pady=(16, 12))

        # ===== 配置卡片 =====
        config_card = ctk.CTkFrame(body, fg_color=C_CARD, corner_radius=R_LG,
                                   border_width=1, border_color=C_BORDER)
        config_card.pack(fill="x")

        # 卡片标题区
        cc_head = ctk.CTkFrame(config_card, fg_color="transparent")
        cc_head.pack(fill="x", padx=20, pady=(16, 4))
        ctk.CTkLabel(cc_head, text="构建配置",
                     font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                     text_color=C_TEXT).pack(side="left")
        self.config_hint = ctk.CTkLabel(
            cc_head, text="安卓 · 每留",
            font=ctk.CTkFont(family=F_UI, size=11),
            text_color=C_DIM)
        self.config_hint.pack(side="right")

        # 配置主行
        row = ctk.CTkFrame(config_card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(8, 18))

        # — 平台选择 —
        col_p = ctk.CTkFrame(row, fg_color="transparent")
        col_p.pack(side="left", padx=(0, 24))
        ctk.CTkLabel(col_p, text="PLATFORM",
                     font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 6))
        self.platform_var = ctk.StringVar(value="安卓")
        self.seg_platform = ctk.CTkSegmentedButton(
            col_p, values=["安卓", "IOS"], variable=self.platform_var,
            font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
            corner_radius=R_SM, height=36,
            fg_color=C_SURFACE,
            selected_color=C_ACCENT, selected_hover_color=C_ACCENT_H,
            unselected_color=C_SURFACE, unselected_hover_color=C_HOVER,
            text_color=C_TEXT, text_color_disabled=C_DIM, width=180)
        self.seg_platform.pack(anchor="w")

        # — 留存选择 —
        col_r = ctk.CTkFrame(row, fg_color="transparent")
        col_r.pack(side="left", padx=(0, 24))
        ctk.CTkLabel(col_r, text="RETENTION",
                     font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 6))
        self.retention_var = ctk.StringVar(value="每留")
        self.seg_retention = ctk.CTkSegmentedButton(
            col_r, values=["每留", "七留"], variable=self.retention_var,
            font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
            corner_radius=R_SM, height=36,
            fg_color=C_SURFACE,
            selected_color=C_ACCENT, selected_hover_color=C_ACCENT_H,
            unselected_color=C_SURFACE, unselected_hover_color=C_HOVER,
            text_color=C_TEXT, text_color_disabled=C_DIM, width=180)
        self.seg_retention.pack(anchor="w")

        # — 模式选择 —
        col_m = ctk.CTkFrame(row, fg_color="transparent")
        col_m.pack(side="left", padx=(0, 24))
        ctk.CTkLabel(col_m, text="MODE",
                     font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                     text_color=C_DIM).pack(anchor="w", pady=(0, 6))
        self.mode_var = ctk.StringVar(value="普通")
        self.seg_mode = ctk.CTkSegmentedButton(
            col_m, values=["普通", "激励"], variable=self.mode_var,
            font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
            corner_radius=R_SM, height=36,
            fg_color=C_SURFACE,
            selected_color="#ffffff", selected_hover_color="#ffffff",
            unselected_color=C_SURFACE, unselected_hover_color=C_HOVER,
            text_color=C_TEXT, text_color_disabled=C_DIM, width=180)
        self.seg_mode.pack(anchor="w")

        # — 操作按钮（右对齐） —
        col_btn = ctk.CTkFrame(row, fg_color="transparent")
        col_btn.pack(side="right")
        # 占位让按钮和分段按钮底部对齐
        ctk.CTkLabel(col_btn, text=" ",
                     font=ctk.CTkFont(family=F_UI, size=10)).pack(pady=(0, 6))
        btn_box = ctk.CTkFrame(col_btn, fg_color="transparent")
        btn_box.pack()
        self.btn_stop = ctk.CTkButton(
            btn_box, text="停止", command=self._on_stop, state="disabled",
            font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
            fg_color=C_CARD, hover_color=C_HOVER,
            text_color=C_TEXT_2,
            border_width=1, border_color=C_BORDER,
            corner_radius=R_SM, width=84, height=36)
        self.btn_stop.pack(side="left", padx=(0, 8))

        self.btn_start = ctk.CTkButton(
            btn_box, text="▶  开始搭建", command=self._on_start,
            font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
            fg_color=C_ACCENT, hover_color=C_ACCENT_H,
            text_color="#ffffff",
            corner_radius=R_SM, width=130, height=36)
        self.btn_start.pack(side="left")

        self.platform_var.trace_add("write", self._update_hint)
        self.retention_var.trace_add("write", self._update_hint)
        self.mode_var.trace_add("write", self._on_mode_changed)
        self._update_hint()

        # ===== 日志卡片 =====
        log_card = ctk.CTkFrame(body, fg_color=C_CARD, corner_radius=R_LG,
                                border_width=1, border_color=C_BORDER)
        log_card.pack(fill="both", expand=True, pady=(14, 0))

        log_header = ctk.CTkFrame(log_card, fg_color="transparent")
        log_header.pack(fill="x", padx=20, pady=(14, 10))

        # 左侧标题 + 计数徽章
        lh_left = ctk.CTkFrame(log_header, fg_color="transparent")
        lh_left.pack(side="left")
        ctk.CTkLabel(lh_left, text="运行日志",
                     font=ctk.CTkFont(family=F_UI, size=13, weight="bold"),
                     text_color=C_TEXT).pack(side="left")
        self.log_count_badge = ctk.CTkLabel(
            lh_left, text="0",
            font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
            text_color=C_TEXT_2,
            fg_color=C_SURFACE, corner_radius=999,
            width=28, height=18)
        self.log_count_badge.pack(side="left", padx=(8, 0))

        # 右侧操作按钮
        lh_right = ctk.CTkFrame(log_header, fg_color="transparent")
        lh_right.pack(side="right")
        ctk.CTkButton(lh_right, text="清空", command=self._clear_log,
                      font=ctk.CTkFont(family=F_UI, size=11),
                      fg_color="transparent", hover_color=C_HOVER,
                      text_color=C_TEXT_2,
                      border_width=1, border_color=C_BORDER,
                      corner_radius=R_SM, width=58, height=28).pack(side="right")

        # 深色日志面板（终端风）
        log_inner = ctk.CTkFrame(log_card, fg_color=C_LOG_BG,
                                 corner_radius=R_MD)
        log_inner.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        # 日志面板顶部：模拟 macOS 终端的 3 个圆点
        dots_bar = ctk.CTkFrame(log_inner, fg_color="transparent", height=24)
        dots_bar.pack(fill="x", padx=12, pady=(8, 0))
        for c in ("#ff5f56", "#ffbd2e", "#27c93f"):
            d = ctk.CTkLabel(dots_bar, text="●",
                             font=ctk.CTkFont(size=12),
                             text_color=c)
            d.pack(side="left", padx=(0, 6))
        ctk.CTkLabel(dots_bar, text="build.log",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_LOG_DIM).pack(side="left", padx=8)

        # 日志文本框
        log_body = ctk.CTkFrame(log_inner, fg_color="transparent")
        log_body.pack(fill="both", expand=True, padx=10, pady=(2, 10))

        self.log_text = tk.Text(
            log_body, wrap="word",
            font=(F_UI, 18),
            bg=C_LOG_BG, fg=C_LOG_FG,
            insertbackground=C_LOG_FG,
            selectbackground=C_LOG_SEL, selectforeground="#ffffff",
            relief="flat", borderwidth=0, padx=12, pady=8,
            state="disabled", spacing1=3, spacing3=3)
        scrollbar = ctk.CTkScrollbar(log_body, command=self.log_text.yview,
                                     button_color="#334155",
                                     button_hover_color="#475569",
                                     fg_color="transparent",
                                     width=10)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)

        # 日志文本标签颜色（终端配色）
        self.log_text.tag_configure("success", foreground="#34d399")
        self.log_text.tag_configure("error", foreground="#f87171")
        self.log_text.tag_configure("warning", foreground="#fbbf24")
        self.log_text.tag_configure("info", foreground=C_LOG_FG)
        self.log_text.tag_configure("separator", foreground=C_LOG_DIM)

        self._log_count = 0

        # ===== 底部状态栏 =====
        footer = ctk.CTkFrame(self._main_frame, fg_color=C_CARD, corner_radius=0, height=30,
                              border_width=0)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        # 顶部 1px 分割线
        fsep = ctk.CTkFrame(footer, fg_color=C_BORDER, height=1, corner_radius=0)
        fsep.place(relx=0, rely=0, relwidth=1, anchor="nw")

        ctk.CTkLabel(footer, text="v1.0",
                     font=ctk.CTkFont(family=F_UI, size=10, weight="bold"),
                     text_color=C_TEXT_2).pack(side="left", padx=(20, 8))
        ctk.CTkLabel(footer, text="·",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(side="left", padx=2)
        ctk.CTkLabel(footer, text="安卓 + IOS",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(side="left", padx=2)
        ctk.CTkLabel(footer, text="·",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(side="left", padx=2)
        ctk.CTkLabel(footer, text="七留 + 每留",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(side="left", padx=2)

        ctk.CTkLabel(footer, text="Ready",
                     font=ctk.CTkFont(family=F_UI, size=10),
                     text_color=C_DIM).pack(side="right", padx=20)

    # ── 辅助方法 ──────────────────────────────────────────────
    def _update_hint(self, *_):
        platform = self.platform_var.get()
        retention = self.retention_var.get()
        mode = self.mode_var.get()
        if mode == "激励":
            self.config_hint.configure(text=f"{platform} · 激励 · {retention}")
        else:
            self.config_hint.configure(text=f"{platform} · {retention}")

    def _get_profile_key(self):
        mode = self.mode_var.get()
        platform = self.platform_var.get()
        retention = self.retention_var.get()
        if mode == "激励":
            return f"{platform}-激励{retention}"
        return f"{platform}-{retention}"

    def _on_mode_changed(self, *_):
        if self.mode_var.get() == "激励":
            self.platform_var.set("安卓")
            self.seg_platform.configure(state="disabled")
        else:
            self.seg_platform.configure(state="normal")
        self._update_hint()

    def _get_log_tag(self, msg):
        if "✅" in msg or "🎉" in msg or "🏁" in msg:
            return "success"
        if "❌" in msg:
            return "error"
        if "⚠️" in msg:
            return "warning"
        if msg.startswith("="):
            return "separator"
        return "info"

    def _append_log(self, msg):
        bottom_before = True
        try:
            bottom_before = self.log_text.yview()[1] >= 0.995
        except Exception:
            pass
        self.log_text.config(state="normal")
        tag = self._get_log_tag(msg)
        self.log_text.insert("end", msg + "\n", tag)
        if bottom_before:
            self.log_text.see("end")
        self.log_text.config(state="disabled")
        self._log_count += 1
        self.log_count_badge.configure(text=str(self._log_count))

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")
        self._log_count = 0
        self.log_count_badge.configure(text="0")

    def _poll_log(self):
        while not self._log_queue.empty():
            try:
                msg = self._log_queue.get_nowait()
                self._append_log(msg)
            except queue.Empty:
                break
        self.after(100, self._poll_log)

    def _log_callback(self, msg):
        self._log_queue.put(msg)

    def _set_status(self, text, color):
        # 状态色映射 → 徽章背景 / 边框 / 文字 / 圆点
        # color 入参（C_GREEN / C_ORANGE / C_RED）决定整套配色
        palette = {
            C_GREEN:  ("#ecfdf5", "#a7f3d0", "#065f46"),  # 就绪 / 成功
            C_ORANGE: ("#fffbeb", "#fde68a", "#92400e"),  # 运行中
            C_RED:    ("#fef2f2", "#fecaca", "#991b1b"),  # 停止 / 错误
        }
        bg, border, fg = palette.get(color, ("#f1f5f9", C_BORDER, C_TEXT_2))
        # 去掉文本里的 ● 前缀，圆点交给独立 label 渲染
        clean = text.lstrip("● ").strip()
        self.status_badge.configure(fg_color=bg, border_color=border)
        self.status_dot.configure(text_color=color)
        self.status_label.configure(text=clean, text_color=fg)

    # ── 事件处理 ──────────────────────────────────────────────
    def _on_start(self):
        self._update_hint()
        key = self._get_profile_key()
        if self._running:
            self._append_log("当前已有搭建任务在运行，先点停止并等待结束后再开始")
            return
        if key not in ALL_PROFILES:
            messagebox.showerror("错误", f"未找到配置: {key}")
            return

        # 优先校验 config.json 是否已经录入了数据
        app_cfg = load_config()
        runtime_cfg = build_runtime_profile_config(key, app_cfg)
        groups = profile_groups_from_config(app_cfg, key)
        if not groups:
            if messagebox.askyesno(
                "尚未配置",
                f"该方向「{key}」还没有任何账号 ID / 链接 / 素材数据。\n\n"
                f"是否现在打开「⚙ 设置」录入？",
            ):
                self._open_settings()
            return

        self._running = True
        self._stop_event.clear()
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.seg_platform.configure(state="disabled")
        self.seg_retention.configure(state="disabled")
        self._set_status(f"● 运行中: {key}", C_ORANGE)
        self._append_log(f"本次读取配置：{len(groups)} 组")
        self._append_log(
            "本次运行变量："
            f"策略={runtime_cfg['strategy']} | "
            f"素材账号ID={runtime_cfg['material_account_id']} | "
            f"受众关键词={runtime_cfg['audience_keyword']} | "
            f"监控按钮={runtime_cfg['monitor_btn_text']} | "
            f"命名前缀={runtime_cfg['name_prefix']} | "
            f"等待倍率={runtime_cfg['wait_scale']}"
        )
        self._append_log(f"{'═'*50}")
        self._append_log(f"🚀 启动搭建: {key}")
        self._append_log(f"{'═'*50}")

        def _worker():
            try:
                if ALL_PROFILES.get(key, {}).get("build_mode") == "incentive":
                    run_build_incentive(key, log_callback=self._log_callback, stop_event=self._stop_event)
                else:
                    run_build(key, log_callback=self._log_callback, stop_event=self._stop_event)
            except Exception as e:
                self._log_callback(f"❌ 运行异常: {e}")
            finally:
                self._log_queue.put("🏁 搭建流程结束")
                self.after(0, self._on_done)

        self._thread = threading.Thread(target=_worker, daemon=True)
        self._thread.start()

    def _on_stop(self):
        if not self._running:
            return
        if self._stop_event.is_set():
            self._set_status("● 等待当前页面操作超时后停止", C_ORANGE)
            return
        self._stop_event.set()
        self._set_status("● 正在停止，最多等待当前页面操作 15 秒", C_RED)
        self._append_log("⏹ 正在停止，等待当前页面操作结束...")

    def _on_done(self):
        self._running = False
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.seg_platform.configure(state="normal")
        self.seg_retention.configure(state="normal")
        self._set_status("● 就绪", C_GREEN)

    # ─────────────────────────────────────────────────────────
    #  ⚙ 设置入口
    # ─────────────────────────────────────────────────────────
    def _open_settings(self):
        if self._settings_frame is None:
            self._settings_frame = SettingsFrame(
                self,
                on_close=self._on_settings_closed,
                on_profile_added=self._on_profile_added_from_settings)
        else:
            self._settings_frame.reload_config()
        self._main_frame.pack_forget()
        self._settings_frame.pack(fill="both", expand=True)
        self.geometry("1080x720")
        self.minsize(900, 600)

    def _on_settings_closed(self):
        self._settings_frame.pack_forget()
        self._main_frame.pack(fill="both", expand=True)
        self.geometry("960x680")
        self.minsize(820, 560)

    def _on_profile_added_from_settings(self, profile_key, group_count):
        if "-" in profile_key:
            platform, retention = profile_key.split("-", 1)
            if platform in ("安卓", "IOS"):
                self.platform_var.set(platform)
            if retention in ("七留", "每留"):
                self.retention_var.set(retention)
        self._on_settings_closed()
        self._set_status("● 配置已更新", C_GREEN)
        self._append_log(f"已覆盖 {group_count} 组数据到「{profile_key}」，当前已切换到该方向，可直接开始搭建")


def main():
    # 首次启动自动创建 4 个方向的目录与 ids.txt 模板
    created = init_data_dirs()
    # 首次启动若 config.json 不存在，则尝试把已有 ids.txt 内容迁入
    migrated = migrate_ids_txt_to_config()
    app = BuildApp()
    if created:
        # 通过 after 延迟弹窗，确保主窗已渲染
        def _notify():
            msg = "已为以下方向自动创建 ids.txt 模板，请编辑后再开始搭建：\n\n"
            msg += "\n".join("• " + p for p in created)
            messagebox.showinfo("首次初始化", msg)
        app.after(300, _notify)
    app.mainloop()


if __name__ == "__main__":
    main()
