import sys, os, glob, traceback
sys.path.insert(0, r'c:\Users\Administrator\Desktop\gui')
import openpyxl
from openpyxl import Workbook

PROMOTION_SPLIT_DEFAULT_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
PROMOTION_SPLIT_KEEP_COLS = [1, 7, 8, 9, 10, 11]
PROMOTION_SPLIT_GROUP_ORDER = ["IOS-每留", "IOS-七留", "Android-每留", "Android-七留"]

def classify_row(row):
    page = str(row[3]) if len(row) > 3 and row[3] else ""
    if "激励" in page:
        return None
    name = str(row[1]) if len(row) > 1 and row[1] else ""
    os_v = str(row[2]) if len(row) > 2 and row[2] else ""
    if "每留" in name:
        return os_v + "-每留"
    if "七留" in name:
        return os_v + "-七留"
    return None

target_dir = PROMOTION_SPLIT_DEFAULT_DIR
pattern = os.path.join(target_dir, "推广链统计_*.xlsx")
files = [f for f in glob.glob(pattern) if "_处理后" not in f and "_processed" not in f and "_样本" not in f and "_gui_test" not in f and "_激励拆分" not in f and "_拆分" not in f]

print(f"Found {len(files)} file(s)")
for src in files:
    print(f"\nProcessing: {os.path.basename(src)}")
    base = os.path.splitext(os.path.basename(src))[0]
    dst = os.path.join(target_dir, base + "_processed.xlsx")
    try:
        wb_src = openpyxl.load_workbook(src)
        ws_src = wb_src.active
        groups = {k: [] for k in PROMOTION_SPLIT_GROUP_ORDER}
        for row in ws_src.iter_rows(values_only=True):
            key = classify_row(row)
            if key in groups:
                groups[key].append(row)
        
        total = sum(len(v) for v in groups.values())
        print(f"  Classified: {total} rows")
        counts = {k: len(v) for k, v in groups.items()}
        print(f"  Groups: {counts}")
        
        if total == 0:
            print("  SKIP: no single-book data")
            continue
        
        wb = Workbook()
        ws = wb.active
        ws.title = "推广链统计"
        GAP = 2
        for i, key in enumerate(PROMOTION_SPLIT_GROUP_ORDER):
            ws.append([key])
            for row in groups[key]:
                ws.append([row[j] if len(row) > j else None for j in PROMOTION_SPLIT_KEEP_COLS])
            if i < len(PROMOTION_SPLIT_GROUP_ORDER) - 1:
                for _ in range(GAP):
                    ws.append([])
        
        wb.save(dst)
        print(f"  Saved: {os.path.basename(dst)}")
        
        def format_rows(rows):
            lines = []
            for row in rows:
                cells = []
                for idx in PROMOTION_SPLIT_KEEP_COLS:
                    value = row[idx] if len(row) > idx else ""
                    cells.append("" if value is None else str(value))
                lines.append("\t".join(cells))
            return "\n".join(lines)
        
        texts = {k: format_rows(v) for k, v in groups.items()}
        for k, t in texts.items():
            print(f"  {k}: {len(t)} chars")
        
        print("  SUCCESS")
    except Exception as e:
        print(f"  FAIL:")
        traceback.print_exc()
